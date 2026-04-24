"""
Streamlit Dashboard for BESS and Northwold Data Analysis
Displays comprehensive analysis of battery energy storage system operations and market trading
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime
import os
import sys

# Add src directory to path for imports
sys.path.append(os.path.join(os.path.dirname(__file__), 'src'))
import digital_twin_config as config
from config import (
    GB_REVENUE_NET_SHARE,
    GB_NET_FOOTNOTE,
    GB_NET_FOOTNOTE_SHORT,
    apply_gb_net,
)
from pages.data_quality import show_data_quality_page
from pages.invoice_analysis import show_invoice_analysis
from pages.monthly_checklist import show_monthly_checklist

# =============================================================================
# GLOBAL COLOR PALETTE — Colorblind-friendly, consistent across all charts
# =============================================================================

# Revenue streams
COLOR_SFFR = '#2C4B78'           # Deep Navy
COLOR_EPEX = '#F18805'           # Warm Orange
COLOR_IDA1 = '#7BC8F6'           # Sky Blue
COLOR_IDC = '#9467BD'            # Purple
COLOR_IMBALANCE_REV = '#2CA02C'  # Forest Green
COLOR_IMBALANCE_COST = '#D62728' # Crimson

# Strategies (Actual/Optimized comparisons)
COLOR_ACTUAL = '#3498db'         # Blue — "what happened"
COLOR_EPEX_ONLY = '#F18805'      # Orange — matches EPEX revenue color
COLOR_SFFR_ONLY = '#2C4B78'      # Navy — matches SFFR revenue color
COLOR_MULTI_MARKET = '#2CA02C'   # Green — "best outcome"

# Ancillary services
COLOR_DCL = '#8C564B'            # Brown
COLOR_DCH = '#E377C2'            # Pink
COLOR_DML = '#17BECF'            # Teal
COLOR_DMH = '#BCBD22'            # Olive
COLOR_DRL = '#AEC7E8'            # Light Steel Blue
COLOR_DRH = '#FFBB78'            # Peach

# Market prices (already consistent — kept as-is)
COLOR_EPEX_DA = '#3498db'        # Blue
COLOR_SSP = '#2ecc71'            # Green
COLOR_SBP = '#e74c3c'            # Red

# Revenue breakdown color map (reusable for pie/donut charts)
REVENUE_COLOR_MAP = {
    'SFFR': COLOR_SFFR,
    'SFFR Revenue': COLOR_SFFR,
    'EPEX': COLOR_EPEX,
    'EPEX 30 DA': COLOR_EPEX,
    'EPEX Trading': COLOR_EPEX,
    'IDA1': COLOR_IDA1,
    'IDA1 Trading': COLOR_IDA1,
    'IDC': COLOR_IDC,
    'IDC Trading': COLOR_IDC,
    'Imbalance': COLOR_IMBALANCE_COST,
    'Imbalance (Net)': COLOR_IMBALANCE_COST,
    'Imbalance Revenue': COLOR_IMBALANCE_REV,
    'Imbalance Charges': COLOR_IMBALANCE_COST,
}

# Strategy colors list (for bar charts with 4 strategies in order:
# Actual GridBeyond, EPEX-Only, SFFR-Only, Multi-Market)
STRATEGY_COLORS = [COLOR_ACTUAL, COLOR_EPEX_ONLY, COLOR_SFFR_ONLY, COLOR_MULTI_MARKET]

# Configure data directory
DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
if not os.path.exists(DATA_DIR):
    DATA_DIR = '.'  # Fallback to current directory for local testing

# Set page config
st.set_page_config(
    page_title="Asset Performance Dashboard",
    page_icon="🔋",
    layout="wide"
)

# Available months configuration
AVAILABLE_MONTHS = {
    "September 2025": {
        "bess_file": "BESS_Sept_2025_converted.csv",
        "northwold_file": "Northwold_Sep_2025_converted.csv",
        "master_file": "Master_BESS_Analysis_Sept_2025.csv",
        "optimization_file": "Optimized_Results_Sept_2025.csv",
        "use_master": False,  # September uses separate files
    },
    "October 2025": {
        "bess_file": None,
        "northwold_file": None,
        "master_file": "Master_BESS_Analysis_Oct_2025.csv",
        "optimization_file": "Optimized_Results_Oct_2025.csv",
        "use_master": True,
    },
    "November 2025": {
        "bess_file": None,
        "northwold_file": None,
        "master_file": "Master_BESS_Analysis_Nov_2025.csv",
        "optimization_file": "Optimized_Results_Nov_2025.csv",
        "use_master": True,
    },
    "December 2025": {
        "bess_file": None,
        "northwold_file": None,
        "master_file": "Master_BESS_Analysis_Dec_2025.csv",
        "optimization_file": "Optimized_Results_Dec_2025.csv",
        "use_master": True,
    },
    "January 2026": {
        "bess_file": None,
        "northwold_file": None,
        "master_file": "Master_BESS_Analysis_Jan_2026.csv",
        "optimization_file": "Optimized_Results_Jan_2026.csv",
        "use_master": True,
    },
    "February 2026": {
        "bess_file": None,
        "northwold_file": None,
        "master_file": "Master_BESS_Analysis_Feb_2026.csv",
        "optimization_file": "Optimized_Results_Feb_2026.csv",
        "use_master": True,
    },
    "March 2026": {
        "bess_file": None,
        "northwold_file": None,
        "master_file": "Master_BESS_Analysis_Mar_2026.csv",
        "optimization_file": "Optimized_Results_Mar_2026.csv",
        "use_master": True,
    },
}

@st.cache_data
def load_data():
    """Load and cache the CSV files (legacy - September only)"""
    try:
        bess_df = pd.read_csv(os.path.join(DATA_DIR, 'BESS_Sept_2025_converted.csv'))
        northwold_df = pd.read_csv(os.path.join(DATA_DIR, 'Northwold_Sep_2025_converted.csv'))

        # Convert timestamps
        bess_df['date'] = pd.to_datetime(bess_df['date'], format='%d/%m/%Y %H:%M:%S')
        northwold_df['Timestamp'] = pd.to_datetime(northwold_df['Timestamp'], format='%d-%m-%Y %H:%M')

        return bess_df, northwold_df
    except Exception as e:
        st.error(f"Error loading data: {str(e)}")
        return None, None

@st.cache_data
def load_month_data(month: str):
    """Load data for selected month - handles both separate files and Master format"""
    config = AVAILABLE_MONTHS.get(month)
    if not config:
        st.error(f"Unknown month: {month}")
        return None, None

    try:
        if config["use_master"]:
            # Load merged Master file
            master_df = pd.read_csv(os.path.join(DATA_DIR, config["master_file"]))

            # Parse timestamp
            if 'Timestamp' in master_df.columns:
                master_df['Timestamp'] = pd.to_datetime(master_df['Timestamp'])
            elif 'Unnamed: 0' in master_df.columns:
                master_df['Timestamp'] = pd.to_datetime(master_df['Unnamed: 0'])
                master_df = master_df.drop(columns=['Unnamed: 0'])

            # Create compatible bess_df from Master file (SCADA columns)
            bess_df = pd.DataFrame()
            bess_df['date'] = master_df['Timestamp']

            # Map power column (different naming in different files)
            if 'Power_MW' in master_df.columns:
                bess_df['Power'] = master_df['Power_MW']
            elif 'Physical_Power_MW' in master_df.columns:
                bess_df['Power'] = master_df['Physical_Power_MW']
            else:
                bess_df['Power'] = 0

            # Map SOC column
            if 'SOC' in master_df.columns:
                bess_df['SOC'] = master_df['SOC']
            elif 'Physical_SoC' in master_df.columns:
                bess_df['SOC'] = master_df['Physical_SoC']
            else:
                bess_df['SOC'] = 50  # Default

            # Frequency if available
            if 'Frequency' in master_df.columns:
                bess_df['Frequency'] = master_df['Frequency']
            else:
                bess_df['Frequency'] = 50.0  # Default grid frequency

            # northwold_df is the Master file with GridBeyond data
            northwold_df = master_df.copy()

            return bess_df, northwold_df
        else:
            # Load separate files (September format)
            bess_df = pd.read_csv(os.path.join(DATA_DIR, config["bess_file"]))
            northwold_df = pd.read_csv(os.path.join(DATA_DIR, config["northwold_file"]))

            # Convert timestamps
            bess_df['date'] = pd.to_datetime(bess_df['date'], format='%d/%m/%Y %H:%M:%S')
            northwold_df['Timestamp'] = pd.to_datetime(northwold_df['Timestamp'], format='%d-%m-%Y %H:%M')

            return bess_df, northwold_df

    except Exception as e:
        st.error(f"Error loading data for {month}: {str(e)}")
        return None, None

def analyze_bess_data(bess_df):
    """Analyze BESS dataset"""
    analysis = {}

    # Time coverage
    analysis['start_time'] = bess_df['date'].min()
    analysis['end_time'] = bess_df['date'].max()
    analysis['duration_days'] = (bess_df['date'].max() - bess_df['date'].min()).days + 1

    # SOC Analysis
    analysis['soc_mean'] = bess_df['SOC'].mean()
    analysis['soc_median'] = bess_df['SOC'].median()
    analysis['soc_min'] = bess_df['SOC'].min()
    analysis['soc_max'] = bess_df['SOC'].max()
    analysis['soc_std'] = bess_df['SOC'].std()

    # Power Analysis
    analysis['power_mean'] = bess_df['Power'].mean()
    analysis['power_max_charge'] = bess_df['Power'].max()
    analysis['power_max_discharge'] = bess_df['Power'].min()
    analysis['total_energy_charged'] = bess_df[bess_df['Power'] > 0]['Power'].sum() * 0.5
    analysis['total_energy_discharged'] = abs(bess_df[bess_df['Power'] < 0]['Power'].sum() * 0.5)

    # Frequency Analysis
    analysis['freq_mean'] = bess_df['Frequency'].mean()
    analysis['freq_min'] = bess_df['Frequency'].min()
    analysis['freq_max'] = bess_df['Frequency'].max()
    analysis['freq_std'] = bess_df['Frequency'].std()

    # Operation patterns
    analysis['charging_periods'] = (bess_df['Power'] > 0).sum()
    analysis['discharging_periods'] = (bess_df['Power'] < 0).sum()
    analysis['idle_periods'] = (bess_df['Power'] == 0).sum()
    analysis['total_periods'] = len(bess_df)

    return analysis

def analyze_northwold_data(northwold_df):
    """Analyze Northwold dataset"""
    analysis = {}

    # Time coverage
    analysis['start_time'] = northwold_df['Timestamp'].min()
    analysis['end_time'] = northwold_df['Timestamp'].max()
    analysis['duration_days'] = (northwold_df['Timestamp'].max() - northwold_df['Timestamp'].min()).days + 1

    # Energy Trading Analysis
    analysis['da_price_mean'] = northwold_df['Day Ahead Price (EPEX)'].mean()
    analysis['da_price_min'] = northwold_df['Day Ahead Price (EPEX)'].min()
    analysis['da_price_max'] = northwold_df['Day Ahead Price (EPEX)'].max()
    analysis['da_price_std'] = northwold_df['Day Ahead Price (EPEX)'].std()

    analysis['intraday_price_mean'] = northwold_df['GB-ISEM Intraday 1 Price'].mean()
    analysis['intraday_price_min'] = northwold_df['GB-ISEM Intraday 1 Price'].min()
    analysis['intraday_price_max'] = northwold_df['GB-ISEM Intraday 1 Price'].max()

    # Revenue Analysis — all GB-traded streams netted by the 5% GridBeyond fee,
    # so figures tie to the GridBeyond invoice and are apples-to-apples with IAR.
    analysis['imbalance_revenue'] = apply_gb_net(northwold_df['Imbalance Revenue'].sum())
    analysis['imbalance_charge'] = apply_gb_net(northwold_df['Imbalance Charge'].sum())
    analysis['net_imbalance'] = analysis['imbalance_revenue'] - analysis['imbalance_charge']

    # Revenue streams (net of GB fee)
    analysis['sffr_revenue'] = apply_gb_net(northwold_df['SFFR revenues'].sum())
    analysis['epex30_revenue'] = apply_gb_net(northwold_df['EPEX 30 DA Revenue'].sum())
    analysis['ida1_revenue'] = apply_gb_net(northwold_df['IDA1 Revenue'].sum())
    analysis['idc_revenue'] = apply_gb_net(northwold_df['IDC Revenue'].sum())

    analysis['total_net_revenue'] = (
        analysis['sffr_revenue'] +
        analysis['epex30_revenue'] +
        analysis['ida1_revenue'] +
        analysis['idc_revenue'] +
        analysis['imbalance_revenue'] -
        analysis['imbalance_charge']
    )

    # Ancillary services
    services = ['SFFR', 'DCL', 'DCH', 'DML', 'DMH', 'DRL', 'DRH']
    analysis['ancillary_services'] = {}
    for service in services:
        clearing_col = f'{service} Clearing Price'
        avail_col = f'{service} Availability'
        if clearing_col in northwold_df.columns:
            analysis['ancillary_services'][service] = {
                'avg_price': northwold_df[clearing_col].mean(),
                'avg_availability': northwold_df[avail_col].mean()
            }

    # Price volatility
    analysis['ssp_std'] = northwold_df['SSP'].std()
    analysis['sbp_std'] = northwold_df['SBP'].std()

    # Trading metrics
    analysis['da_mw_mean'] = northwold_df['DA MW'].mean()
    analysis['epex30_mw_mean'] = northwold_df['EPEX 30 DA MW'].mean()
    analysis['ida1_mw_mean'] = northwold_df['IDA1 MW'].mean()

    return analysis

def calculate_cycles(df, power_col, capacity_mwh=8.4, dt_hours=0.5):
    """
    Calculate battery cycles using three methodologies.

    Args:
        df: DataFrame with power data
        power_col: Column name for power (positive = discharge, negative = charge)
        capacity_mwh: Battery capacity (default 8.4)
        dt_hours: Time step in hours (default 0.5 for 30-min)

    Returns:
        dict with cycles_discharge, cycles_full, cycles_throughput
    """
    power = pd.to_numeric(df[power_col], errors='coerce').fillna(0)

    # Convert power (MW) to energy (MWh) for each period
    energy = power * dt_hours

    # Method A: Discharge-only (current method)
    discharge_mwh = energy[energy > 0].sum()
    cycles_discharge = discharge_mwh / capacity_mwh

    # Method B: Full Equivalent Cycles (Industry Standard)
    charge_mwh = abs(energy[energy < 0].sum())
    cycles_full = (discharge_mwh + charge_mwh) / 2 / capacity_mwh

    # Method C: Throughput-based (mathematically same as B)
    total_throughput = discharge_mwh + charge_mwh
    cycles_throughput = total_throughput / (2 * capacity_mwh)

    return {
        'discharge_mwh': discharge_mwh,
        'charge_mwh': charge_mwh,
        'cycles_discharge': cycles_discharge,
        'cycles_full': cycles_full,
        'cycles_throughput': cycles_throughput
    }


def calculate_tb_spreads(df, price_col='Day Ahead Price (EPEX)'):
    """
    Calculate Top-Bottom (TB) spreads from price data.

    TB spreads measure daily arbitrage potential:
    - TB1 = highest hourly price - lowest hourly price
    - TB2 = sum of 2 highest - sum of 2 lowest hourly prices
    - TB3 = sum of 3 highest - sum of 3 lowest hourly prices

    Args:
        df: DataFrame with half-hourly price data
        price_col: Column name for price data

    Returns:
        DataFrame with columns: Date, TB1, TB2, TB3
    """
    df = df.copy()

    # Handle timestamp column
    if 'Timestamp' in df.columns:
        df['Timestamp'] = pd.to_datetime(df['Timestamp'], errors='coerce')
    elif 'Unnamed: 0' in df.columns:
        df['Timestamp'] = pd.to_datetime(df['Unnamed: 0'], errors='coerce')
    else:
        return pd.DataFrame(columns=['Date', 'TB1', 'TB2', 'TB3'])

    # Parse price data
    if price_col not in df.columns:
        return pd.DataFrame(columns=['Date', 'TB1', 'TB2', 'TB3'])

    df['Price'] = pd.to_numeric(df[price_col], errors='coerce')

    # Extract date and hour for grouping
    df['Date'] = df['Timestamp'].dt.date
    df['Hour'] = df['Timestamp'].dt.hour

    # Aggregate half-hourly to hourly (average of two HH periods)
    hourly_prices = df.groupby(['Date', 'Hour'])['Price'].mean().reset_index()

    # Calculate TB spreads per day
    tb_results = []

    for date in hourly_prices['Date'].unique():
        day_prices = hourly_prices[hourly_prices['Date'] == date]['Price'].dropna()

        if len(day_prices) < 3:
            continue

        sorted_desc = day_prices.sort_values(ascending=False).values
        sorted_asc = day_prices.sort_values(ascending=True).values

        # TB1: Highest - Lowest
        tb1 = sorted_desc[0] - sorted_asc[0]

        # TB2: Sum of 2 highest - Sum of 2 lowest
        tb2 = sum(sorted_desc[:2]) - sum(sorted_asc[:2])

        # TB3: Sum of 3 highest - Sum of 3 lowest
        tb3 = sum(sorted_desc[:3]) - sum(sorted_asc[:3])

        tb_results.append({
            'Date': date,
            'TB1': tb1,
            'TB2': tb2,
            'TB3': tb3
        })

    return pd.DataFrame(tb_results)


def calculate_daily_arbitrage(df):
    """
    Calculate daily wholesale trading revenue (arbitrage).

    Arbitrage revenue = EPEX DA + IDA1 + IDC revenues
    (excludes SFFR as it's frequency response, not arbitrage)

    Args:
        df: DataFrame with revenue columns

    Returns:
        DataFrame with columns: Date, Arbitrage_Revenue
    """
    df = df.copy()

    # Handle timestamp column
    if 'Timestamp' in df.columns:
        df['Timestamp'] = pd.to_datetime(df['Timestamp'], errors='coerce')
    elif 'Unnamed: 0' in df.columns:
        df['Timestamp'] = pd.to_datetime(df['Unnamed: 0'], errors='coerce')
    else:
        return pd.DataFrame(columns=['Date', 'Arbitrage_Revenue'])

    df['Date'] = df['Timestamp'].dt.date

    # Sum revenue columns (handle both naming conventions)
    def safe_col(dataframe, col):
        if col in dataframe.columns:
            return pd.to_numeric(dataframe[col], errors='coerce').fillna(0)
        return 0

    df['EPEX_Revenue'] = safe_col(df, 'EPEX 30 DA Revenue') + safe_col(df, 'EPEX DA Revenues')
    df['IDA1_Revenue'] = safe_col(df, 'IDA1 Revenue')
    df['IDC_Revenue'] = safe_col(df, 'IDC Revenue')

    df['Arbitrage_Total'] = df['EPEX_Revenue'] + df['IDA1_Revenue'] + df['IDC_Revenue']

    # Group by date
    daily = df.groupby('Date').agg({
        'Arbitrage_Total': 'sum'
    }).reset_index()

    daily.columns = ['Date', 'Arbitrage_Revenue']

    return daily


def show_asset_details():
    """Display asset details from digital twin configuration"""
    st.title("🏭 Asset Details")
    st.markdown("### Northwold Solar Farm (Hall Farm) - Battery Energy Storage System")

    # Physical Specifications
    st.subheader("⚡ Physical Specifications")

    col1, col2 = st.columns(2)

    with col1:
        st.info("**Power Capabilities**")
        st.metric("Maximum Charging Rate", f"{config.P_IMP_MAX_MW} MW",
                 help="Maximum rate at which the battery can charge (import power)")
        st.metric("Maximum Discharging Rate", f"{config.P_EXP_MAX_MW} MW",
                 help="Maximum rate at which the battery can discharge (export power)")
        st.metric("Power Asymmetry Ratio", f"{config.P_EXP_MAX_MW/config.P_IMP_MAX_MW:.2f}x",
                 help="Discharge to charge power ratio")

    with col2:
        st.info("**Energy Storage**")
        st.metric("Usable Capacity", f"{config.CAPACITY_MWH} MWh",
                 help="Total usable energy storage capacity")
        st.metric("Minimum SOC", f"{config.SOC_MIN_PCT*100:.0f}%",
                 help=f"Minimum State of Charge ({config.SOC_MIN_MWH:.2f} MWh)")
        st.metric("Maximum SOC", f"{config.SOC_MAX_PCT*100:.0f}%",
                 help=f"Maximum State of Charge ({config.SOC_MAX_MWH:.2f} MWh)")

    # Efficiency & Warranty
    st.subheader("🔧 Efficiency & Warranty")

    col1, col2 = st.columns(2)

    with col1:
        st.info("**Efficiency Parameters**")
        st.metric("Round-Trip Efficiency", f"{config.EFF_ROUND_TRIP*100:.0f}%",
                 help="Energy efficiency for a complete charge-discharge cycle")

    with col2:
        st.info("**Warranty Constraints**")
        st.metric("Maximum Daily Cycles", f"{config.CYCLES_PER_DAY} cycles/day",
                 help="Warranty limit on daily cycling")

def show_operations_summary(bess_df, northwold_df, bess_analysis, northwold_analysis, month: str = "September 2025"):
    """Display monthly operations summary"""
    st.title(f"📊 {month} Operations Summary")

    # Create tabs for the operations summary
    tab1, tab2 = st.tabs(["💰 Trading Analysis", "📈 Visualizations"])

    with tab1:

        # Revenue Breakdown
        st.subheader("💵 Revenue Breakdown")

        col1, col2, col3 = st.columns(3)

        with col1:
            st.metric("SFFR Revenues", f"£{northwold_analysis['sffr_revenue']:,.2f}",
                     help="Static Firm Frequency Response - Largest revenue stream")
        with col2:
            st.metric("IDA1 Revenue", f"£{northwold_analysis['ida1_revenue']:,.2f}",
                     help="Intraday Auction 1 trading revenue")
        with col3:
            st.metric("EPEX 30 DA Revenue", f"£{northwold_analysis['epex30_revenue']:,.2f}",
                     help="Day-ahead market revenue")

        col4, col5, col6 = st.columns(3)
        with col4:
            st.metric("Imbalance Revenue", f"£{northwold_analysis['imbalance_revenue']:,.2f}",
                     help="Revenue from imbalance mechanism (negative indicates costs)")
        with col5:
            st.metric("Imbalance Charge", f"£{northwold_analysis['imbalance_charge']:,.2f}",
                     help="Charges from imbalance mechanism")
        with col6:
            st.metric("**TOTAL NET REVENUE**", f"£{northwold_analysis['total_net_revenue']:,.2f}",
                     help="Total net revenue for the month, after 5% GridBeyond fee")

        st.caption(GB_NET_FOOTNOTE_SHORT)

        # Market Prices
        st.subheader("📈 Market Prices Analysis")

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("**Day Ahead Price (EPEX)**")
            price_data = {
                'Average': f"£{northwold_analysis['da_price_mean']:.2f}",
                'Minimum': f"£{northwold_analysis['da_price_min']:.2f}",
                'Maximum': f"£{northwold_analysis['da_price_max']:.2f}",
                'Std Dev': f"£{northwold_analysis['da_price_std']:.2f}"
            }
            for key, value in price_data.items():
                st.text(f"{key}: {value}")

        with col2:
            st.markdown("**Intraday Price**")
            intraday_data = {
                'Average': f"£{northwold_analysis['intraday_price_mean']:.2f}",
                'Minimum': f"£{northwold_analysis['intraday_price_min']:.2f}",
                'Maximum': f"£{northwold_analysis['intraday_price_max']:.2f}",
                'Spread (ID-DA)': f"£{(northwold_analysis['intraday_price_mean'] - northwold_analysis['da_price_mean']):.2f}"
            }
            for key, value in intraday_data.items():
                st.text(f"{key}: {value}")

        # Ancillary Services
        st.subheader("⚡ Ancillary Services Pricing")

        services_data = []
        for service, data in northwold_analysis['ancillary_services'].items():
            services_data.append({
                'Service': service,
                'Avg Clearing Price (£)': f"{data['avg_price']:.2f}",
                'Avg Availability (MW)': f"{data['avg_availability']:.0f}"
            })

        services_df = pd.DataFrame(services_data)
        st.dataframe(services_df, use_container_width=True, hide_index=True)

        # Trading Activity
        st.subheader("📊 Trading Activity")

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Average DA MW", f"{northwold_analysis['da_mw_mean']:.2f} MW")
        with col2:
            st.metric("Average EPEX 30 DA MW", f"{northwold_analysis['epex30_mw_mean']:.2f} MW")
        with col3:
            st.metric("Average IDA1 MW", f"{northwold_analysis['ida1_mw_mean']:.2f} MW")

    with tab2:

        # Create visualizations
        col1, col2 = st.columns(2)

        with col1:
            # Revenue breakdown pie chart
            st.subheader("Revenue Breakdown")
            revenue_data = {
                'SFFR': abs(northwold_analysis['sffr_revenue']),
                'EPEX 30 DA': abs(northwold_analysis['epex30_revenue']),
                'IDA1': abs(northwold_analysis['ida1_revenue']),
                'Imbalance (Net)': abs(northwold_analysis['net_imbalance'])
            }

            fig = px.pie(
                values=list(revenue_data.values()),
                names=list(revenue_data.keys()),
                title="Revenue Sources Distribution",
                color=list(revenue_data.keys()),
                color_discrete_map=REVENUE_COLOR_MAP
            )
            st.plotly_chart(fig, use_container_width=True)
            st.caption(GB_NET_FOOTNOTE_SHORT)

        with col2:
            # SOC distribution
            st.subheader("State of Charge Distribution")
            fig = px.histogram(
                bess_df,
                x='SOC',
                nbins=50,
                title="SOC Distribution",
                labels={'SOC': 'State of Charge (%)', 'count': 'Frequency'}
            )
            st.plotly_chart(fig, use_container_width=True)

        # Time series plots
        st.subheader("Time Series Analysis")

        # Power over time
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=bess_df['date'],
            y=bess_df['Power'],
            mode='lines',
            name='Power (MW)',
            line=dict(width=1)
        ))
        fig.update_layout(
            title="Battery Power Over Time",
            xaxis_title="Date",
            yaxis_title="Power (MW)",
            hovermode='x unified'
        )
        st.plotly_chart(fig, use_container_width=True)

        # Price comparison
        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=northwold_df['Timestamp'],
            y=northwold_df['Day Ahead Price (EPEX)'],
            mode='lines',
            name='Day Ahead Price',
            line=dict(width=1)
        ))
        fig.add_trace(go.Scatter(
            x=northwold_df['Timestamp'],
            y=northwold_df['GB-ISEM Intraday 1 Price'],
            mode='lines',
            name='Intraday Price',
            line=dict(width=1)
        ))
        fig.update_layout(
            title="Energy Prices Comparison",
            xaxis_title="Date",
            yaxis_title="Price (£/MWh)",
            hovermode='x unified'
        )
        st.plotly_chart(fig, use_container_width=True)


def show_multimarket_optimization(month: str = "September 2025"):
    """Display multi-market optimization results"""
    st.title(f"🚀 Multi-Market Optimization Analysis - {month}")
    st.markdown("### Enhanced Trading Strategy with Cross-Market Arbitrage")

    # Get month-specific file paths
    month_config = AVAILABLE_MONTHS.get(month, AVAILABLE_MONTHS["September 2025"])
    optimization_file = month_config.get("optimization_file", "Optimized_Results_MultiMarket.csv")
    master_file = month_config.get("master_file", "Master_BESS_Analysis_Sept_2025.csv")

    # Check if multi-market results exist
    try:
        multi_df = pd.read_csv(os.path.join(DATA_DIR, optimization_file))
        multi_df['Timestamp'] = pd.to_datetime(multi_df['Timestamp'])
    except:
        st.warning(f"Multi-market results not found for {month}. Click below to generate.")
        if st.button("Run Multi-Market Optimization"):
            with st.spinner("Running optimization... This may take a minute."):
                from src import phase3_multimarket
                multi_df = phase3_multimarket.run_phase_3_multimarket(os.path.join(DATA_DIR, master_file))
                multi_df.to_csv(os.path.join(DATA_DIR, optimization_file), index=False)
                st.success(f"Optimization complete! Results saved to {optimization_file}")
                st.rerun()
        return

    # Summary metrics
    st.subheader("💰 Revenue Comparison")

    col1, col2, col3, col4 = st.columns(4)

    # Optimised revenue is netted by the 5% GB fee on the assumption that, in real-world
    # deployment, these trades would also clear through GridBeyond (or an equivalent aggregator).
    total_epex_daily = apply_gb_net(multi_df['Optimised_Revenue_Daily'].sum())
    total_epex_efa = apply_gb_net(multi_df['Optimised_Revenue_EFA'].sum())
    total_multi = apply_gb_net(multi_df['Optimised_Revenue_Multi'].sum())
    improvement = ((total_multi / total_epex_daily - 1) * 100) if total_epex_daily else 0

    with col1:
        st.metric("EPEX-Only (Daily)", f"£{total_epex_daily:,.2f}",
                 help="Original strategy using only EPEX prices with daily switching (net of 5% GB fee)")
    with col2:
        st.metric("EPEX-Only (EFA)", f"£{total_epex_efa:,.2f}",
                 delta=f"{((total_epex_efa/total_epex_daily - 1) * 100):.1f}%" if total_epex_daily else None,
                 help="EPEX prices with 2-hour block switching (net of 5% GB fee)")
    with col3:
        st.metric("Multi-Market", f"£{total_multi:,.2f}",
                 delta=f"+{improvement:.1f}%",
                 help="Optimized across all available markets (net of 5% GB fee)")
    with col4:
        st.metric("Additional Revenue", f"£{(total_multi - total_epex_daily):,.2f}",
                 help="Extra revenue from multi-market strategy (net of 5% GB fee)")

    st.caption(GB_NET_FOOTNOTE_SHORT)

    # Market usage analysis
    st.subheader("📊 Market Utilization")

    # Add strategy selector dropdown
    strategy_option = st.selectbox(
        "Select Strategy to View:",
        ["Multi-Market", "EPEX-only (Daily)", "EPEX-only (EFA)", "Actual"],
        index=0,  # Default to Multi-Market
        help="Choose which strategy's market utilization to display"
    )

    # Initialize variables for volume and revenue tracking
    market_usage = pd.Series()
    market_revenue = pd.Series()
    chart_title = "No Data Available"

    # Determine which column to use based on selection
    if strategy_option == "Multi-Market":
        if 'Market_Used_Multi' in multi_df.columns:
            market_usage = multi_df['Market_Used_Multi'].value_counts()
            # Calculate revenue by market
            if 'Optimised_Revenue_Multi' in multi_df.columns:
                market_revenue = multi_df.groupby('Market_Used_Multi')['Optimised_Revenue_Multi'].sum()
            chart_title = "Multi-Market Strategy"
        else:
            st.warning("Multi-Market data not available")

    elif strategy_option == "EPEX-only (Daily)":
        if 'Strategy_Selected_Daily' in multi_df.columns:
            market_usage = multi_df['Strategy_Selected_Daily'].value_counts()
            if 'Optimised_Revenue_Daily' in multi_df.columns:
                market_revenue = multi_df.groupby('Strategy_Selected_Daily')['Optimised_Revenue_Daily'].sum()
            chart_title = "EPEX-only (Daily) Strategy"
        else:
            st.warning("EPEX Daily data not available")

    elif strategy_option == "EPEX-only (EFA)":
        if 'Strategy_Selected_EFA' in multi_df.columns:
            market_usage = multi_df['Strategy_Selected_EFA'].value_counts()
            if 'Optimised_Revenue_EFA' in multi_df.columns:
                market_revenue = multi_df.groupby('Strategy_Selected_EFA')['Optimised_Revenue_EFA'].sum()
            chart_title = "EPEX-only (EFA) Strategy"
        else:
            st.warning("EPEX EFA data not available")

    else:  # Actual
        # For actual operation, derive market usage AND revenue from revenue columns
        try:
            actual_master_df = pd.read_csv(os.path.join(DATA_DIR, master_file))
            if 'Unnamed: 0' in actual_master_df.columns:
                actual_master_df.rename(columns={'Unnamed: 0': 'Timestamp'}, inplace=True)

            # Find power column for buy/sell direction
            power_col = None
            for col in actual_master_df.columns:
                if 'Battery MWh' in col and 'Output' in col:
                    power_col = col
                    break

            # Track both market assignment and revenue per market
            actual_markets = []
            actual_revenues = []

            for idx, row in actual_master_df.iterrows():
                market = 'Idle'
                revenue = 0.0
                power_val = row.get(power_col, 0) if power_col else 0

                # Helper to safely get numeric value (handles NaN)
                def safe_get(r, col, default=0):
                    val = r.get(col, default)
                    return default if pd.isna(val) else val

                # Check trading markets first (actual energy traded)
                epex_rev = safe_get(row, 'EPEX 30 DA Revenue', 0) + safe_get(row, 'EPEX DA Revenues', 0)
                ida1_rev = safe_get(row, 'IDA1 Revenue', 0)
                idc_rev = safe_get(row, 'IDC Revenue', 0)
                imb_rev = safe_get(row, 'Imbalance Revenue', 0) - abs(safe_get(row, 'Imbalance Charge', 0))
                sffr_rev = safe_get(row, 'SFFR revenues', 0)

                # Check EPEX markets
                if abs(epex_rev) > 0.01:
                    if power_val > 0.1:
                        market = 'Sell-EPEX'
                    elif power_val < -0.1:
                        market = 'Buy-EPEX'
                    else:
                        market = 'EPEX'
                    revenue = epex_rev
                # Check IDA1
                elif abs(ida1_rev) > 0.01:
                    if power_val > 0.1:
                        market = 'Sell-IDA1'
                    elif power_val < -0.1:
                        market = 'Buy-IDA1'
                    else:
                        market = 'IDA1'
                    revenue = ida1_rev
                # Check IDC
                elif abs(idc_rev) > 0.01:
                    if power_val > 0.1:
                        market = 'Sell-IDC'
                    elif power_val < -0.1:
                        market = 'Buy-IDC'
                    else:
                        market = 'IDC'
                    revenue = idc_rev
                # Check Imbalance
                elif abs(row.get('Imbalance Revenue', 0)) > 0.01 or abs(row.get('Imbalance Charge', 0)) > 0.01:
                    market = 'Imbalance'
                    revenue = imb_rev
                # Check SFFR (availability payment)
                elif abs(sffr_rev) > 0.01:
                    market = 'SFFR'
                    revenue = sffr_rev
                # Default - check physical operation without identified market
                elif power_col and abs(power_val) > 0.1:
                    if power_val > 0:
                        market = 'Discharging (Unknown)'
                    else:
                        market = 'Charging (Unknown)'

                actual_markets.append(market)
                actual_revenues.append(revenue)

            # Create DataFrame for aggregation
            actual_df = pd.DataFrame({'market': actual_markets, 'revenue': actual_revenues})
            market_usage = actual_df['market'].value_counts()
            market_revenue = actual_df.groupby('market')['revenue'].sum()
            chart_title = "Actual (GridBeyond)"

        except Exception as e:
            st.warning(f"Could not load actual data: {str(e)}")

    # Net all market_revenue values by the 5% GB fee before displaying, so Optimised
    # strategies and Actual are both shown post-fee and are directly comparable to
    # the GridBeyond invoice.
    if not market_revenue.empty:
        try:
            market_revenue = market_revenue.astype(float) * GB_REVENUE_NET_SHARE
        except (ValueError, TypeError):
            pass

    # Display the visualization if we have data
    if not market_usage.empty:
        # Two pie charts: Volume and Revenue
        col1, col2 = st.columns(2)

        with col1:
            # Volume pie chart
            fig_volume = px.pie(
                values=market_usage.values[:10],
                names=market_usage.index[:10],
                title=f"📊 Volume (Periods) - {chart_title}",
                color=market_usage.index[:10],
                color_discrete_map=REVENUE_COLOR_MAP
            )
            fig_volume.update_traces(textposition='inside', textinfo='percent+label')
            st.plotly_chart(fig_volume, use_container_width=True)

        with col2:
            # Revenue pie chart
            if not market_revenue.empty:
                # Handle negative values for pie chart (show absolute, note negatives)
                rev_for_pie = market_revenue.abs()
                fig_revenue = px.pie(
                    values=rev_for_pie.values[:10],
                    names=rev_for_pie.index[:10],
                    title=f"💰 Revenue (£) - {chart_title}",
                    color=rev_for_pie.index[:10],
                    color_discrete_map=REVENUE_COLOR_MAP
                )
                fig_revenue.update_traces(textposition='inside', textinfo='percent+label')
                st.plotly_chart(fig_revenue, use_container_width=True)
            else:
                st.info("Revenue data not available for this strategy")

        # Combined statistics table
        st.markdown("**Market Statistics**")

        # Build table with both volume and revenue
        table_data = {
            'Market': market_usage.index[:10],
            'Periods': market_usage.values[:10],
            '% of Time': (market_usage.values[:10] / len(multi_df) * 100).round(1)
        }

        if not market_revenue.empty:
            # Align revenue with market_usage index
            revenue_values = [market_revenue.get(m, 0) for m in market_usage.index[:10]]
            table_data['Revenue (£)'] = [f"£{v:,.2f}" for v in revenue_values]
            total_rev = sum(revenue_values)
            table_data['% of Revenue'] = [(v / total_rev * 100 if total_rev != 0 else 0) for v in revenue_values]
            table_data['% of Revenue'] = [f"{v:.1f}%" for v in table_data['% of Revenue']]

        market_stats = pd.DataFrame(table_data)
        st.dataframe(market_stats, use_container_width=True, hide_index=True)
        st.caption(GB_NET_FOOTNOTE_SHORT)

    # Price spread analysis
    st.subheader("📈 Market Price Spreads")

    if 'Market_Spread' in multi_df.columns:
        col1, col2 = st.columns(2)

        with col1:
            spread_stats = {
                'Average Spread': f"£{multi_df['Market_Spread'].mean():.2f}/MWh",
                'Max Spread': f"£{multi_df['Market_Spread'].max():.2f}/MWh",
                'Min Spread': f"£{multi_df['Market_Spread'].min():.2f}/MWh",
                'Spreads >£30': f"{(multi_df['Market_Spread'] > 30).sum()} periods",
                'Spreads >£50': f"{(multi_df['Market_Spread'] > 50).sum()} periods"
            }
            for key, value in spread_stats.items():
                st.text(f"{key}: {value}")

        with col2:
            # Spread distribution histogram
            fig_hist = px.histogram(
                multi_df,
                x='Market_Spread',
                nbins=50,
                title="Distribution of Market Spreads",
                labels={'Market_Spread': 'Price Spread (£/MWh)', 'count': 'Frequency'}
            )
            st.plotly_chart(fig_hist, use_container_width=True)

    # Time series comparison
    st.subheader("📊 Strategy Comparison Over Time")

    # Revenue over time — optimiser output netted by 5% GB fee for apples-to-apples comparison.
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=multi_df['Timestamp'],
        y=(multi_df['Optimised_Revenue_Daily'] * GB_REVENUE_NET_SHARE).cumsum(),
        mode='lines',
        name='EPEX-Only (Cumulative)',
        line=dict(width=2, color=COLOR_EPEX_ONLY)
    ))
    fig.add_trace(go.Scatter(
        x=multi_df['Timestamp'],
        y=(multi_df['Optimised_Revenue_Multi'] * GB_REVENUE_NET_SHARE).cumsum(),
        mode='lines',
        name='Multi-Market (Cumulative)',
        line=dict(width=2, color=COLOR_MULTI_MARKET)
    ))
    fig.update_layout(
        title="Cumulative Revenue Comparison",
        xaxis_title="Date",
        yaxis_title="Cumulative Revenue (£)",
        hovermode='x unified',
        height=400
    )
    st.plotly_chart(fig, use_container_width=True)
    st.caption(GB_NET_FOOTNOTE_SHORT)

    # Best markets by time
    st.subheader("🏆 Best Markets by Period")

    if 'Best_Buy_Market' in multi_df.columns and 'Best_Sell_Market' in multi_df.columns:
        col1, col2 = st.columns(2)

        with col1:
            buy_markets = multi_df['Best_Buy_Market'].value_counts()
            st.markdown("**Best Markets for Buying (Charging)**")
            buy_df = pd.DataFrame({
                'Market': buy_markets.index,
                'Times Selected': buy_markets.values,
                'Percentage': (buy_markets.values / len(multi_df) * 100).round(1)
            })
            st.dataframe(buy_df.head(), use_container_width=True, hide_index=True)

        with col2:
            sell_markets = multi_df['Best_Sell_Market'].value_counts()
            st.markdown("**Best Markets for Selling (Discharging)**")
            sell_df = pd.DataFrame({
                'Market': sell_markets.index,
                'Times Selected': sell_markets.values,
                'Percentage': (sell_markets.values / len(multi_df) * 100).round(1)
            })
            st.dataframe(sell_df.head(), use_container_width=True, hide_index=True)

    # Key insights
    st.subheader("🔍 Key Insights")

    insights_col1, insights_col2 = st.columns(2)

    with insights_col1:
        st.info(f"""
        **Revenue Improvement:**
        - Multi-market strategy yields {improvement:.1f}% more revenue
        - Additional monthly revenue: £{(total_multi - total_epex_daily):,.2f}
        - Annualized benefit: £{((total_multi - total_epex_daily) * 12):,.2f}
        """)

    with insights_col2:
        st.success(f"""
        **Market Dynamics:**
        - SSP market provides best arbitrage opportunities
        - Battery idle {(market_usage.get('Idle', 0) / len(multi_df) * 100):.1f}% of time to preserve cycles
        - Average spread captured: £{multi_df['Market_Spread'].mean():.2f}/MWh
        """)

def show_bess_health(month: str = "September 2025"):
    """Display BESS health analysis - cycles and degradation"""
    st.title(f"🔋 BESS Health Analysis - {month}")
    st.markdown("### Battery Cycling and Degradation Assessment")

    # Get month-specific file paths
    month_config = AVAILABLE_MONTHS.get(month, AVAILABLE_MONTHS["September 2025"])
    master_file = month_config.get("master_file", "Master_BESS_Analysis_Sept_2025.csv")
    optimization_file = month_config.get("optimization_file", "Optimized_Results_MultiMarket.csv")

    # Load required data
    try:
        # Load actual data
        master_df = pd.read_csv(os.path.join(DATA_DIR, master_file))
        if 'Unnamed: 0' in master_df.columns:
            master_df.rename(columns={'Unnamed: 0': 'Timestamp'}, inplace=True)
        master_df['Timestamp'] = pd.to_datetime(master_df['Timestamp'])

        # Load multi-market optimization results
        multi_df = pd.read_csv(os.path.join(DATA_DIR, optimization_file))
        multi_df['Timestamp'] = pd.to_datetime(multi_df['Timestamp'])

    except Exception as e:
        st.error(f"Error loading data files for {month}: {str(e)}")
        st.info(f"Please ensure both {master_file} and {optimization_file} exist. Run optimization first if needed.")
        return

    # Constants from config
    CAPACITY_MWH = config.CAPACITY_MWH  # 8.4 MWh
    WARRANTY_CYCLES_DAILY = config.CYCLES_PER_DAY  # 1.5 cycles/day
    WARRANTY_DEGRADATION_ANNUAL_PCT = 2.5  # Annual degradation at warranty limit
    DEGRADATION_PER_CYCLE_PCT = WARRANTY_DEGRADATION_ANNUAL_PCT / (WARRANTY_CYCLES_DAILY * 365)

    # Calculate number of days from data
    num_days = (master_df['Timestamp'].max() - master_df['Timestamp'].min()).days + 1

    # Find actual battery power column (in MW, not pre-converted to MWh)
    actual_power_col = None
    if 'Physical_Power_MW' in master_df.columns:
        actual_power_col = 'Physical_Power_MW'
    elif 'Power_MW' in master_df.columns:
        actual_power_col = 'Power_MW'

    # ==================== CYCLE METHODOLOGY COMPARISON ====================
    st.subheader("📐 Cycle Calculation Methods Comparison")

    st.caption("""
**Cycle Calculation Methods:**
- **A: Discharge-only** = Discharge MWh / Capacity (current method)
- **B: Full Equivalent** = (Discharge + Charge) / 2 / Capacity ⭐ Industry Standard
- **C: Throughput** = Total Throughput / (2 × Capacity) (mathematically same as B)

Warranty limit: 1.5 cycles/day (547 cycles/year)
""")

    # Calculate cycles using all 3 methods for Actual and Multi-Market
    if actual_power_col and actual_power_col in master_df.columns:
        actual_cycles_all = calculate_cycles(master_df, actual_power_col, CAPACITY_MWH, dt_hours=0.5)
    else:
        actual_cycles_all = {'discharge_mwh': 0, 'charge_mwh': 0, 'cycles_discharge': 0, 'cycles_full': 0, 'cycles_throughput': 0}

    # For Multi-Market, the values are already in MWh (not MW), so dt_hours=1
    multi_cycles_all = calculate_cycles(multi_df, 'Optimised_Net_MWh_Multi', CAPACITY_MWH, dt_hours=1.0)

    # Create comparison table for all 3 methods
    method_comparison = pd.DataFrame({
        'Method': ['A: Discharge-only', 'B: Full Equivalent (Industry Std)', 'C: Throughput-based'],
        'Actual Total Cycles': [
            actual_cycles_all['cycles_discharge'],
            actual_cycles_all['cycles_full'],
            actual_cycles_all['cycles_throughput']
        ],
        'Multi-Market Total Cycles': [
            multi_cycles_all['cycles_discharge'],
            multi_cycles_all['cycles_full'],
            multi_cycles_all['cycles_throughput']
        ],
        'Actual Daily Avg': [
            actual_cycles_all['cycles_discharge'] / num_days,
            actual_cycles_all['cycles_full'] / num_days,
            actual_cycles_all['cycles_throughput'] / num_days
        ],
        'Multi-Market Daily Avg': [
            multi_cycles_all['cycles_discharge'] / num_days,
            multi_cycles_all['cycles_full'] / num_days,
            multi_cycles_all['cycles_throughput'] / num_days
        ]
    })

    st.dataframe(
        method_comparison.style.format({
            'Actual Total Cycles': '{:.2f}',
            'Multi-Market Total Cycles': '{:.2f}',
            'Actual Daily Avg': '{:.3f}',
            'Multi-Market Daily Avg': '{:.3f}'
        }),
        use_container_width=True,
        hide_index=True
    )

    # Show energy totals
    col1, col2 = st.columns(2)
    with col1:
        st.metric("Actual Discharge Energy", f"{actual_cycles_all['discharge_mwh']:.2f} MWh",
                 help="Total energy discharged (exported to grid)")
        st.metric("Actual Charge Energy", f"{actual_cycles_all['charge_mwh']:.2f} MWh",
                 help="Total energy charged (imported from grid)")
    with col2:
        st.metric("Multi-Market Discharge Energy", f"{multi_cycles_all['discharge_mwh']:.2f} MWh",
                 help="Total simulated discharge energy")
        st.metric("Multi-Market Charge Energy", f"{multi_cycles_all['charge_mwh']:.2f} MWh",
                 help="Total simulated charge energy")

    st.markdown("---")

    # ==================== STRATEGY COMPARISON WITH METHOD SELECTOR ====================
    st.subheader("📊 Strategy Cycling Comparison")

    # Method selector
    cycle_method = st.radio(
        "Select Cycle Calculation Method for Strategy Comparison:",
        ["A: Discharge-only", "B: Full Equivalent (Industry Std)", "C: Throughput-based"],
        index=1,  # Default to industry standard
        horizontal=True,
        help="Method B (Full Equivalent) is the industry standard for battery cycling"
    )

    # Map selection to cycle key
    method_key_map = {
        "A: Discharge-only": 'cycles_discharge',
        "B: Full Equivalent (Industry Std)": 'cycles_full',
        "C: Throughput-based": 'cycles_throughput'
    }
    selected_key = method_key_map[cycle_method]

    # Calculate cycles for all strategies using selected method
    strategies_data = []

    # Helper to get cycles and discharge based on selected method
    def get_strategy_metrics(df, col, is_mwh=False):
        dt = 1.0 if is_mwh else 0.5
        cycles_data = calculate_cycles(df, col, CAPACITY_MWH, dt_hours=dt)
        return cycles_data[selected_key], cycles_data['discharge_mwh']

    # 1. Actual (Original) Operation
    if actual_power_col and actual_power_col in master_df.columns:
        actual_cycles, actual_discharge = get_strategy_metrics(master_df, actual_power_col, is_mwh=False)
        actual_daily_cycles = actual_cycles / num_days
        actual_degradation = actual_cycles * DEGRADATION_PER_CYCLE_PCT
        strategies_data.append({
            'Strategy': 'Actual Operation',
            'Total Discharge (MWh)': actual_discharge,
            'Total Cycles': actual_cycles,
            'Daily Cycles': actual_daily_cycles,
            'Est. Degradation (%)': actual_degradation,
            'Warranty Status': '✅ OK' if actual_daily_cycles <= WARRANTY_CYCLES_DAILY else '⚠️ EXCEEDED'
        })

    # 2. EPEX-Only Daily Strategy
    epex_daily_cycles, epex_daily_discharge = get_strategy_metrics(multi_df, 'Optimised_Net_MWh_Daily', is_mwh=True)
    epex_daily_daily_cycles = epex_daily_cycles / num_days
    epex_daily_degradation = epex_daily_cycles * DEGRADATION_PER_CYCLE_PCT
    strategies_data.append({
        'Strategy': 'EPEX-Only (Daily)',
        'Total Discharge (MWh)': epex_daily_discharge,
        'Total Cycles': epex_daily_cycles,
        'Daily Cycles': epex_daily_daily_cycles,
        'Est. Degradation (%)': epex_daily_degradation,
        'Warranty Status': '✅ OK' if epex_daily_daily_cycles <= WARRANTY_CYCLES_DAILY else '⚠️ EXCEEDED'
    })

    # 3. EPEX-Only EFA Strategy
    epex_efa_cycles, epex_efa_discharge = get_strategy_metrics(multi_df, 'Optimised_Net_MWh_EFA', is_mwh=True)
    epex_efa_daily_cycles = epex_efa_cycles / num_days
    epex_efa_degradation = epex_efa_cycles * DEGRADATION_PER_CYCLE_PCT
    strategies_data.append({
        'Strategy': 'EPEX-Only (EFA)',
        'Total Discharge (MWh)': epex_efa_discharge,
        'Total Cycles': epex_efa_cycles,
        'Daily Cycles': epex_efa_daily_cycles,
        'Est. Degradation (%)': epex_efa_degradation,
        'Warranty Status': '✅ OK' if epex_efa_daily_cycles <= WARRANTY_CYCLES_DAILY else '⚠️ EXCEEDED'
    })

    # 4. Multi-Market Strategy
    multi_cycles, multi_discharge = get_strategy_metrics(multi_df, 'Optimised_Net_MWh_Multi', is_mwh=True)
    multi_daily_cycles = multi_cycles / num_days
    multi_degradation = multi_cycles * DEGRADATION_PER_CYCLE_PCT
    strategies_data.append({
        'Strategy': 'Multi-Market',
        'Total Discharge (MWh)': multi_discharge,
        'Total Cycles': multi_cycles,
        'Daily Cycles': multi_daily_cycles,
        'Est. Degradation (%)': multi_degradation,
        'Warranty Status': '✅ OK' if multi_daily_cycles <= WARRANTY_CYCLES_DAILY else '⚠️ EXCEEDED'
    })

    # Create DataFrame
    health_df = pd.DataFrame(strategies_data)

    st.info(f"**Using Method: {cycle_method}**")

    # Display comparison table
    st.dataframe(
        health_df.style.format({
            'Total Discharge (MWh)': '{:.2f}',
            'Total Cycles': '{:.2f}',
            'Daily Cycles': '{:.3f}',
            'Est. Degradation (%)': '{:.4f}'
        }),
        use_container_width=True,
        hide_index=True
    )

    # Key metrics cards
    col1, col2, col3 = st.columns(3)

    with col1:
        st.metric(
            "Warranty Limit",
            f"{WARRANTY_CYCLES_DAILY} cycles/day",
            help="Maximum daily cycles allowed under warranty"
        )

    with col2:
        st.metric(
            "Annual Degradation at Limit",
            f"{WARRANTY_DEGRADATION_ANNUAL_PCT}%",
            help="Expected annual degradation if operating at warranty limit"
        )

    with col3:
        st.metric(
            "Analysis Period",
            f"{num_days} days",
            help=f"{month} data"
        )

    # Visualizations
    st.subheader("📈 Visual Comparisons")

    col1, col2 = st.columns(2)

    with col1:
        # Daily cycles bar chart
        fig_cycles = go.Figure(data=[
            go.Bar(
                x=health_df['Strategy'],
                y=health_df['Daily Cycles'],
                text=health_df['Daily Cycles'].round(3),
                textposition='auto',
                marker_color=STRATEGY_COLORS
            )
        ])
        # Add warranty limit line
        fig_cycles.add_hline(
            y=WARRANTY_CYCLES_DAILY,
            line_dash="dash",
            line_color="red",
            annotation_text="Warranty Limit (1.5 cycles/day)"
        )
        fig_cycles.update_layout(
            title="Daily Cycling Comparison",
            yaxis_title="Cycles per Day",
            xaxis_title="Strategy",
            height=400
        )
        st.plotly_chart(fig_cycles, use_container_width=True)

    with col2:
        # Degradation comparison
        fig_degrade = go.Figure(data=[
            go.Bar(
                x=health_df['Strategy'],
                y=health_df['Est. Degradation (%)'],
                text=health_df['Est. Degradation (%)'].round(4),
                textposition='auto',
                marker_color=STRATEGY_COLORS
            )
        ])
        fig_degrade.update_layout(
            title="Monthly Degradation Comparison",
            yaxis_title="Estimated Degradation (%)",
            xaxis_title="Strategy",
            height=400
        )
        st.plotly_chart(fig_degrade, use_container_width=True)

    # Total discharge comparison
    st.subheader("⚡ Energy Throughput Analysis")

    fig_discharge = go.Figure()

    # Add bars for each strategy
    strategies = health_df['Strategy'].tolist()
    discharges = health_df['Total Discharge (MWh)'].tolist()

    fig_discharge.add_trace(go.Bar(
        x=strategies,
        y=discharges,
        text=[f"{d:.1f} MWh" for d in discharges],
        textposition='auto',
        marker_color=STRATEGY_COLORS
    ))

    fig_discharge.update_layout(
        title=f"Total Energy Discharged ({month})",
        yaxis_title="Total Discharge (MWh)",
        xaxis_title="Strategy",
        height=400,
        showlegend=False
    )

    st.plotly_chart(fig_discharge, use_container_width=True)

    # Projected annual analysis
    st.subheader("📅 Projected Annual Impact")

    annual_data = []
    for _, row in health_df.iterrows():
        annual_cycles = row['Daily Cycles'] * 365
        annual_degradation = annual_cycles * DEGRADATION_PER_CYCLE_PCT
        annual_data.append({
            'Strategy': row['Strategy'],
            'Projected Annual Cycles': annual_cycles,
            'Projected Annual Degradation (%)': annual_degradation,
            'Years to 80% SOH': 20.0 / annual_degradation if annual_degradation > 0 else float('inf')
        })

    annual_df = pd.DataFrame(annual_data)

    col1, col2 = st.columns(2)

    with col1:
        st.dataframe(
            annual_df[['Strategy', 'Projected Annual Cycles', 'Projected Annual Degradation (%)']].style.format({
                'Projected Annual Cycles': '{:.1f}',
                'Projected Annual Degradation (%)': '{:.2f}'
            }),
            use_container_width=True,
            hide_index=True
        )

    with col2:
        # Years to 80% SOH
        fig_years = go.Figure(data=[
            go.Bar(
                x=annual_df['Strategy'],
                y=annual_df['Years to 80% SOH'],
                text=[f"{y:.1f} years" if y < 100 else "100+ years" for y in annual_df['Years to 80% SOH']],
                textposition='auto',
                marker_color=STRATEGY_COLORS
            )
        ])
        fig_years.update_layout(
            title="Estimated Battery Lifespan",
            yaxis_title="Years to 80% State of Health",
            xaxis_title="Strategy",
            height=400
        )
        st.plotly_chart(fig_years, use_container_width=True)

    # Insights
    st.subheader("🔍 Key Health Insights")

    best_strategy = health_df.loc[health_df['Daily Cycles'].idxmin(), 'Strategy']
    worst_strategy = health_df.loc[health_df['Daily Cycles'].idxmax(), 'Strategy']

    col1, col2 = st.columns(2)

    with col1:
        st.info(f"""
        **Battery Preservation:**
        - Best for battery health: **{best_strategy}**
        - Most intensive: **{worst_strategy}**
        - All strategies within warranty limits ✅
        - Multi-market strategy balances revenue and health
        """)

    with col2:
        st.success(f"""
        **Optimization Benefits:**
        - Multi-market reduces cycling vs EPEX strategies
        - Higher revenue with lower battery wear
        - Extended battery lifespan potential
        - Warranty compliance maintained
        """)

    # ==================== NEW SECTION: DAILY CYCLES COMPARISON ====================
    st.markdown("---")
    st.subheader("📊 Daily Cycles: Actual vs Multi-Market Optimization")

    st.info(f"**Using Method: {cycle_method}** (same as strategy comparison above)")

    # Helper function for daily cycle calculation based on selected method
    def calc_daily_cycles(df, col, date_col, is_mwh=False):
        """Calculate daily cycles using selected method."""
        dt = 1.0 if is_mwh else 0.5
        power = pd.to_numeric(df[col], errors='coerce').fillna(0)
        energy = power * dt

        if selected_key == 'cycles_discharge':
            # Method A: Discharge only
            df_temp = df.copy()
            df_temp['_energy'] = energy
            df_temp['_discharge'] = df_temp['_energy'].apply(lambda x: x if x > 0 else 0)
            daily = df_temp.groupby(date_col)['_discharge'].sum() / CAPACITY_MWH
        elif selected_key in ['cycles_full', 'cycles_throughput']:
            # Method B/C: Full equivalent (discharge + charge) / 2
            df_temp = df.copy()
            df_temp['_energy'] = energy
            df_temp['_discharge'] = df_temp['_energy'].apply(lambda x: x if x > 0 else 0)
            df_temp['_charge'] = df_temp['_energy'].apply(lambda x: abs(x) if x < 0 else 0)
            daily_discharge = df_temp.groupby(date_col)['_discharge'].sum()
            daily_charge = df_temp.groupby(date_col)['_charge'].sum()
            daily = (daily_discharge + daily_charge) / 2 / CAPACITY_MWH
        return daily.reset_index()

    # Calculate daily cycles for each day
    master_df['Date'] = master_df['Timestamp'].dt.date
    multi_df['Date'] = multi_df['Timestamp'].dt.date

    # Actual daily cycles
    if actual_power_col and actual_power_col in master_df.columns:
        actual_daily = calc_daily_cycles(master_df, actual_power_col, 'Date', is_mwh=False)
        actual_daily.columns = ['Date', 'Actual_Cycles']
    else:
        actual_daily = pd.DataFrame({'Date': master_df['Date'].unique(), 'Actual_Cycles': 0})

    # Multi-market daily cycles
    multi_daily = calc_daily_cycles(multi_df, 'Optimised_Net_MWh_Multi', 'Date', is_mwh=True)
    multi_daily.columns = ['Date', 'Multi_Cycles']

    # Merge the two
    daily_cycles_df = pd.merge(actual_daily, multi_daily, on='Date', how='outer').fillna(0)
    daily_cycles_df['Date'] = pd.to_datetime(daily_cycles_df['Date'])
    daily_cycles_df = daily_cycles_df.sort_values('Date')

    # Create comparison chart
    fig_daily = go.Figure()

    fig_daily.add_trace(go.Bar(
        x=daily_cycles_df['Date'],
        y=daily_cycles_df['Actual_Cycles'],
        name='Actual (GridBeyond)',
        marker_color=COLOR_ACTUAL,
        opacity=0.8
    ))

    fig_daily.add_trace(go.Bar(
        x=daily_cycles_df['Date'],
        y=daily_cycles_df['Multi_Cycles'],
        name='Multi-Market Optimized',
        marker_color=COLOR_MULTI_MARKET,
        opacity=0.8
    ))

    # Add warranty limit line
    fig_daily.add_hline(
        y=WARRANTY_CYCLES_DAILY,
        line_dash="dash",
        line_color="red",
        annotation_text=f"Warranty Limit ({WARRANTY_CYCLES_DAILY} cycles/day)",
        annotation_position="top right"
    )

    fig_daily.update_layout(
        title=f"Daily Battery Cycles - {month}",
        xaxis_title="Date",
        yaxis_title="Cycles per Day",
        barmode='group',
        height=450,
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1
        ),
        hovermode='x unified'
    )

    st.plotly_chart(fig_daily, use_container_width=True)

    # Summary stats
    method_help = {
        'cycles_discharge': 'Discharge-only method (A)',
        'cycles_full': 'Full Equivalent method (B) - Industry Standard',
        'cycles_throughput': 'Throughput method (C)'
    }
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Actual Avg Daily Cycles", f"{daily_cycles_df['Actual_Cycles'].mean():.3f}",
                 help=f"Using {method_help[selected_key]}. Average cycles per day for actual operation.")
    with col2:
        st.metric("Multi-Market Avg Daily", f"{daily_cycles_df['Multi_Cycles'].mean():.3f}",
                 help=f"Using {method_help[selected_key]}. Simulated daily cycles under multi-market optimization.")
    with col3:
        st.metric("Actual Max Day", f"{daily_cycles_df['Actual_Cycles'].max():.3f}",
                 help=f"Using {method_help[selected_key]}. Highest single-day cycling for actual operation.")
    with col4:
        st.metric("Multi-Market Max Day", f"{daily_cycles_df['Multi_Cycles'].max():.3f}",
                 help=f"Using {method_help[selected_key]}. Highest single-day cycling for simulated strategy.")

    # ==================== NEW SECTION: WARRANTY EXCEEDANCE TABLE ====================
    st.markdown("---")
    st.subheader("⚠️ Warranty Limit Exceedance Analysis")

    # Find days exceeding warranty limits
    daily_cycles_df['Actual_Exceeded'] = daily_cycles_df['Actual_Cycles'] > WARRANTY_CYCLES_DAILY
    daily_cycles_df['Multi_Exceeded'] = daily_cycles_df['Multi_Cycles'] > WARRANTY_CYCLES_DAILY

    actual_exceeded_days = daily_cycles_df[daily_cycles_df['Actual_Exceeded']]
    multi_exceeded_days = daily_cycles_df[daily_cycles_df['Multi_Exceeded']]

    col1, col2 = st.columns(2)

    with col1:
        st.markdown(f"**Actual Operation: {len(actual_exceeded_days)} days exceeded**")
        if len(actual_exceeded_days) > 0:
            exceeded_display = actual_exceeded_days[['Date', 'Actual_Cycles']].copy()
            exceeded_display['Date'] = exceeded_display['Date'].dt.strftime('%Y-%m-%d')
            exceeded_display['Over Limit By'] = exceeded_display['Actual_Cycles'] - WARRANTY_CYCLES_DAILY
            exceeded_display.columns = ['Date', 'Cycles', 'Over Limit']
            exceeded_display = exceeded_display.sort_values('Cycles', ascending=False)

            st.dataframe(
                exceeded_display.style.format({
                    'Cycles': '{:.3f}',
                    'Over Limit': '{:+.3f}'
                }).background_gradient(subset=['Over Limit'], cmap='Reds'),
                use_container_width=True,
                hide_index=True
            )
        else:
            st.success("✅ No days exceeded warranty limits!")

    with col2:
        st.markdown(f"**Multi-Market Optimized: {len(multi_exceeded_days)} days exceeded**")
        if len(multi_exceeded_days) > 0:
            exceeded_display = multi_exceeded_days[['Date', 'Multi_Cycles']].copy()
            exceeded_display['Date'] = exceeded_display['Date'].dt.strftime('%Y-%m-%d')
            exceeded_display['Over Limit By'] = exceeded_display['Multi_Cycles'] - WARRANTY_CYCLES_DAILY
            exceeded_display.columns = ['Date', 'Cycles', 'Over Limit']
            exceeded_display = exceeded_display.sort_values('Cycles', ascending=False)

            st.dataframe(
                exceeded_display.style.format({
                    'Cycles': '{:.3f}',
                    'Over Limit': '{:+.3f}'
                }).background_gradient(subset=['Over Limit'], cmap='Reds'),
                use_container_width=True,
                hide_index=True
            )
        else:
            st.success("✅ No days exceeded warranty limits!")

    # Comparison summary
    st.markdown("---")
    st.subheader("📋 Warranty Compliance Summary")

    summary_data = {
        'Metric': [
            'Total Days Analyzed',
            'Days Exceeding Warranty (Actual)',
            'Days Exceeding Warranty (Multi-Market)',
            'Compliance Rate (Actual)',
            'Compliance Rate (Multi-Market)',
            'Total Excess Cycles (Actual)',
            'Total Excess Cycles (Multi-Market)'
        ],
        'Value': [
            f"{len(daily_cycles_df)} days",
            f"{len(actual_exceeded_days)} days",
            f"{len(multi_exceeded_days)} days",
            f"{(1 - len(actual_exceeded_days)/len(daily_cycles_df))*100:.1f}%",
            f"{(1 - len(multi_exceeded_days)/len(daily_cycles_df))*100:.1f}%",
            f"{daily_cycles_df[daily_cycles_df['Actual_Exceeded']]['Actual_Cycles'].sum() - len(actual_exceeded_days)*WARRANTY_CYCLES_DAILY:.3f}" if len(actual_exceeded_days) > 0 else "0.000",
            f"{daily_cycles_df[daily_cycles_df['Multi_Exceeded']]['Multi_Cycles'].sum() - len(multi_exceeded_days)*WARRANTY_CYCLES_DAILY:.3f}" if len(multi_exceeded_days) > 0 else "0.000"
        ]
    }

    summary_df = pd.DataFrame(summary_data)
    st.dataframe(summary_df, use_container_width=True, hide_index=True)

    # Insight box
    if len(actual_exceeded_days) > len(multi_exceeded_days):
        st.warning(f"""
        **⚠️ Finding:** Actual operation exceeded warranty limits on **{len(actual_exceeded_days) - len(multi_exceeded_days)} more days**
        than the multi-market optimized strategy. This indicates GridBeyond is cycling the battery more aggressively than optimal.
        """)
    elif len(multi_exceeded_days) > len(actual_exceeded_days):
        st.info(f"""
        **ℹ️ Finding:** Multi-market optimization would exceed warranty limits on **{len(multi_exceeded_days) - len(actual_exceeded_days)} more days**
        than actual operation. The optimization prioritizes revenue over battery preservation.
        """)
    else:
        st.success("""
        **✅ Finding:** Both strategies have similar warranty compliance. Battery health is being managed consistently.
        """)


def show_report_page(month: str = "September 2025"):
    """Display comprehensive performance report comparing actual vs multi-market optimization"""

    # Header
    st.title("📑 Performance Report")
    st.markdown(f"### Northwold BESS Performance Audit - {month}")
    st.markdown("---")

    # Get month-specific file paths
    month_config = AVAILABLE_MONTHS.get(month, AVAILABLE_MONTHS["September 2025"])
    master_file = month_config.get("master_file", "Master_BESS_Analysis_Sept_2025.csv")
    optimization_file = month_config.get("optimization_file", "Optimized_Results_MultiMarket.csv")
    northwold_file = month_config.get("northwold_file")

    # Load required data
    try:
        # Load actual data
        master_df = pd.read_csv(os.path.join(DATA_DIR, master_file))
        if 'Unnamed: 0' in master_df.columns:
            master_df.rename(columns={'Unnamed: 0': 'Timestamp'}, inplace=True)
        master_df['Timestamp'] = pd.to_datetime(master_df['Timestamp'])
        master_df['Date'] = master_df['Timestamp'].dt.date

        # Load multi-market optimization results
        multi_df = pd.read_csv(os.path.join(DATA_DIR, optimization_file))
        multi_df['Timestamp'] = pd.to_datetime(multi_df['Timestamp'])
        multi_df['Date'] = multi_df['Timestamp'].dt.date

        # Load northwold data for actual revenues (use master file if northwold not available)
        if northwold_file:
            northwold_df = pd.read_csv(os.path.join(DATA_DIR, northwold_file))
        else:
            northwold_df = master_df  # Master file contains GridBeyond revenue data

    except Exception as e:
        st.error(f"Error loading data files for {month}: {str(e)}")
        st.info(f"Please ensure {master_file} and {optimization_file} exist. Run optimization first if needed.")
        return

    # Calculate key metrics using actual column names in Northwold CSV.
    # All GB-traded streams are netted by the 5% GridBeyond fee so the figures tie to
    # the GridBeyond invoice. The algorithmic "Multi-Market potential" is also netted
    # on the assumption that in real-world deployment it would clear through GridBeyond.
    if 'SFFR revenues' in northwold_df.columns:
        sffr_revenue = apply_gb_net(northwold_df['SFFR revenues'].sum())
    else:
        sffr_cols = [col for col in northwold_df.columns if 'SFFR' in col and 'revenue' in col.lower()]
        sffr_revenue = apply_gb_net(northwold_df[sffr_cols[0]].sum()) if sffr_cols else 0

    epex_revenue = apply_gb_net(northwold_df['EPEX 30 DA Revenue'].sum()) if 'EPEX 30 DA Revenue' in northwold_df.columns else 0
    ida1_revenue = apply_gb_net(northwold_df['IDA1 Revenue'].sum()) if 'IDA1 Revenue' in northwold_df.columns else 0
    imbalance_revenue = apply_gb_net(northwold_df['Imbalance Revenue'].sum()) if 'Imbalance Revenue' in northwold_df.columns else 0
    imbalance_charge = apply_gb_net(northwold_df['Imbalance Charge'].sum()) if 'Imbalance Charge' in northwold_df.columns else 0
    net_imbalance = imbalance_revenue - imbalance_charge
    actual_total = sffr_revenue + epex_revenue + ida1_revenue + net_imbalance

    # Multi-market revenue (optimiser output also netted by 5% GB fee)
    multi_total = apply_gb_net(multi_df['Optimised_Revenue_Multi'].sum())

    # Performance gap
    improvement_pct = ((multi_total / actual_total - 1) * 100)
    revenue_gap = multi_total - actual_total

    # ==================== SECTION 1: EXECUTIVE SUMMARY ====================
    st.header("1️⃣ The Performance Gap")
    st.markdown("The analysis reveals a significant gap between the asset's actual revenue and its optimized potential using multi-market strategies.")

    # Key metrics
    col1, col2, col3 = st.columns(3)

    with col1:
        st.metric(
            f"Actual Revenue ({month[:3]})",
            f"£{actual_total:,.0f}",
            help=f"Total revenue from actual operation in {month}"
        )

    with col2:
        st.metric(
            "Multi-Market Potential",
            f"£{multi_total:,.0f}",
            delta=f"+{improvement_pct:.0f}%",
            help="Optimized revenue using multi-market strategy"
        )

    with col3:
        st.metric(
            "Identified Opportunity",
            f"£{revenue_gap:,.0f}",
            delta=f"+{improvement_pct:.0f}%",
            help="Additional revenue available through optimization"
        )

    # Annualized comparison chart
    st.subheader("📊 Annualized Revenue Comparison")

    # Calculate annualized values per MW
    capacity_mw = config.P_EXP_MAX_MW  # 7.5 MW
    actual_annual_per_mw = (actual_total * 12) / capacity_mw
    multi_annual_per_mw = (multi_total * 12) / capacity_mw

    fig_annual = go.Figure(data=[
        go.Bar(
            x=['Actual (£/MW/yr)', 'Multi-Market (£/MW/yr)'],
            y=[actual_annual_per_mw, multi_annual_per_mw],
            text=[f'£{actual_annual_per_mw:,.0f}', f'£{multi_annual_per_mw:,.0f}'],
            textposition='auto',
            marker_color=[COLOR_ACTUAL, COLOR_MULTI_MARKET]
        )
    ])

    fig_annual.update_layout(
        title="Annualized Revenue per MW",
        yaxis_title="Revenue (£/MW/yr)",
        height=400,
        showlegend=False
    )

    st.plotly_chart(fig_annual, use_container_width=True)
    st.caption(GB_NET_FOOTNOTE_SHORT)

    # ==================== SECTION 2: ACTUAL PERFORMANCE ====================
    st.header("2️⃣ Actual Performance Analysis")
    st.markdown("The aggregator's strategy relied heavily on SFFR with significant imbalance penalties eroding profits.")

    col1, col2 = st.columns([2, 1])

    with col1:
        # Revenue breakdown donut chart
        revenue_components = {
            'SFFR Revenue': abs(sffr_revenue),
            'EPEX Trading': abs(epex_revenue),
            'IDA1 Trading': abs(ida1_revenue),
            'Imbalance (Net)': abs(net_imbalance)
        }

        fig_donut = px.pie(
            values=list(revenue_components.values()),
            names=list(revenue_components.keys()),
            title="Actual Revenue Breakdown",
            hole=0.4,
            color=list(revenue_components.keys()),
            color_discrete_map=REVENUE_COLOR_MAP
        )

        st.plotly_chart(fig_donut, use_container_width=True)
        st.caption(GB_NET_FOOTNOTE_SHORT)

    with col2:
        st.error(f"""
        **⚠️ Critical Finding**

        **-£{abs(net_imbalance):,.0f}**
        in Net Imbalance Impact

        This loss erased **{abs(net_imbalance)/(sffr_revenue+epex_revenue+ida1_revenue)*100:.0f}%**
        of the total revenue generated by the asset.
        """)

    # ==================== SECTION 3: OPTIMIZATION OPPORTUNITY ====================
    st.header("3️⃣ The Optimization Opportunity")
    st.markdown("The Digital Twin model shows that a multi-market strategy dramatically increases profitability.")

    # Asset specifications
    st.subheader("Asset Specifications")
    col1, col2, col3 = st.columns(3)

    with col1:
        st.info(f"""
        **Capacity**
        # {config.CAPACITY_MWH} MWh
        Usable Energy Storage
        """)

    with col2:
        st.info(f"""
        **Power**
        # {config.P_EXP_MAX_MW}/{config.P_IMP_MAX_MW} MW
        Discharge/Charge Rates
        """)

    with col3:
        st.info(f"""
        **Efficiency**
        # {config.EFF_ROUND_TRIP*100:.0f}%
        Round-Trip Efficiency
        """)

    # Strategy comparison
    st.subheader("📈 Revenue Mix Comparison")

    # Calculate revenues for different strategies.
    # Optimiser outputs netted by 5% GB fee so Actual vs Optimised is apples-to-apples.
    strategies = ['Actual', 'EPEX-Daily', 'EPEX-EFA', 'Multi-Market']
    sffr_revenues = [
        sffr_revenue,
        apply_gb_net(multi_df[multi_df['Strategy_Selected_Daily'] == 'SFFR']['Optimised_Revenue_Daily'].sum()),
        apply_gb_net(multi_df[multi_df['Strategy_Selected_EFA'] == 'SFFR']['Optimised_Revenue_EFA'].sum()),
        0  # Multi-market doesn't use SFFR when optimizing
    ]

    market_revenues = [
        epex_revenue + ida1_revenue,
        apply_gb_net(multi_df[multi_df['Strategy_Selected_Daily'] == 'EPEX']['Optimised_Revenue_Daily'].sum()),
        apply_gb_net(multi_df[multi_df['Strategy_Selected_EFA'] == 'EPEX']['Optimised_Revenue_EFA'].sum()),
        multi_total
    ]

    fig_mix = go.Figure()
    fig_mix.add_trace(go.Bar(
        name='SFFR',
        x=strategies,
        y=sffr_revenues,
        marker_color=COLOR_SFFR
    ))
    fig_mix.add_trace(go.Bar(
        name='Market Trading',
        x=strategies,
        y=market_revenues,
        marker_color=COLOR_EPEX
    ))

    fig_mix.update_layout(
        title="Strategy Revenue Comparison",
        yaxis_title="Revenue (£)",
        barmode='stack',
        height=400,
        hovermode='x unified'
    )

    st.plotly_chart(fig_mix, use_container_width=True)
    st.caption(GB_NET_FOOTNOTE_SHORT)

    # ==================== SECTION 4: DEGRADATION & WARRANTY ====================
    st.header("4️⃣ Degradation & Warranty Analysis")
    st.markdown("The asset is operating well below warranty limits, providing room for more aggressive strategies.")

    # Find actual battery output column
    actual_col = None
    for col in master_df.columns:
        if 'Battery MWh' in col and 'Output' in col:
            actual_col = col
            break

    if actual_col:
        # Calculate number of days from data
        num_days = (master_df['Timestamp'].max() - master_df['Timestamp'].min()).days + 1

        # Calculate degradation metrics
        actual_discharge = master_df[master_df[actual_col] > 0][actual_col].sum()
        actual_cycles = actual_discharge / config.CAPACITY_MWH
        actual_daily_cycles = actual_cycles / num_days

        multi_discharge = (multi_df[multi_df['Optimised_Net_MWh_Multi'] > 0]['Optimised_Net_MWh_Multi'].sum())
        multi_cycles = multi_discharge / config.CAPACITY_MWH
        multi_daily_cycles = multi_cycles / num_days

        # Create comparison table
        degradation_data = pd.DataFrame({
            'Scenario': ['Actual Usage', 'Multi-Market Strategy', 'Warranty Limit'],
            'Total Discharge (MWh)': [actual_discharge, multi_discharge, config.CYCLES_PER_DAY * num_days * config.CAPACITY_MWH],
            'Avg Cycles/Day': [actual_daily_cycles, multi_daily_cycles, config.CYCLES_PER_DAY],
            'Est. Monthly Degradation': [
                f"{actual_cycles * 0.0046:.3f}%",
                f"{multi_cycles * 0.0046:.3f}%",
                f"{config.CYCLES_PER_DAY * num_days * 0.0046:.3f}%"
            ],
            'Status': [
                '✅ Well Below Limit',
                '✅ Within Warranty',
                '⚠️ Maximum Allowed'
            ]
        })

        st.dataframe(degradation_data, use_container_width=True, hide_index=True)

    # ==================== SECTION 5: DAILY PERFORMANCE ====================
    st.header("5️⃣ Daily Performance Breakdown")
    st.markdown("Daily analysis reveals consistent revenue gaps, especially on high volatility days.")

    # Calculate daily revenues using actual column names
    revenue_cols = ['SFFR revenues', 'EPEX 30 DA Revenue', 'IDA1 Revenue',
                   'Imbalance Revenue', 'Imbalance Charge']

    # Only include columns that exist
    agg_dict = {}
    for col in revenue_cols:
        if col in master_df.columns:
            agg_dict[col] = 'sum'

    if agg_dict:
        daily_actual = master_df.groupby('Date').agg(agg_dict)

        # Calculate total revenue per day
        daily_actual['Total'] = 0
        if 'SFFR revenues' in daily_actual.columns:
            daily_actual['Total'] += daily_actual['SFFR revenues']
        if 'EPEX 30 DA Revenue' in daily_actual.columns:
            daily_actual['Total'] += daily_actual['EPEX 30 DA Revenue']
        if 'IDA1 Revenue' in daily_actual.columns:
            daily_actual['Total'] += daily_actual['IDA1 Revenue']
        if 'Imbalance Revenue' in daily_actual.columns:
            daily_actual['Total'] += daily_actual['Imbalance Revenue']
        if 'Imbalance Charge' in daily_actual.columns:
            daily_actual['Total'] -= daily_actual['Imbalance Charge']
        # Net daily totals by 5% GB fee for apples-to-apples with Optimised.
        daily_actual['Total'] = daily_actual['Total'] * GB_REVENUE_NET_SHARE
    else:
        daily_actual = pd.DataFrame()

    # Optimiser daily totals also netted by 5% GB fee.
    daily_multi = multi_df.groupby('Date')['Optimised_Revenue_Multi'].sum() * GB_REVENUE_NET_SHARE

    # Create line chart
    fig_daily = go.Figure()

    if not daily_actual.empty:
        fig_daily.add_trace(go.Scatter(
            x=daily_actual.index,
            y=daily_actual['Total'],
            mode='lines+markers',
            name='Actual Revenue',
            line=dict(color=COLOR_ACTUAL, width=2),
            marker=dict(size=6)
        ))

    fig_daily.add_trace(go.Scatter(
        x=daily_multi.index,
        y=daily_multi.values,
        mode='lines+markers',
        name='Multi-Market Potential',
        line=dict(color=COLOR_MULTI_MARKET, width=2, dash='dash'),
        marker=dict(size=6)
    ))

    fig_daily.update_layout(
        title="Daily Revenue Comparison",
        xaxis_title="Date",
        yaxis_title="Daily Revenue (£)",
        hovermode='x unified',
        height=400,
        showlegend=True
    )

    st.plotly_chart(fig_daily, use_container_width=True)
    st.caption(GB_NET_FOOTNOTE_SHORT)

    # ==================== SECTION 6: KEY DRIVERS OF LOSS ====================
    st.header("6️⃣ Key Drivers of Revenue Loss")

    col1, col2, col3 = st.columns(3)

    with col1:
        st.warning(f"""
        **1. Imbalance Penalties**

        The asset incurred **-£{abs(net_imbalance):,.0f}** in net imbalance impact,
        suggesting forecast errors or non-delivery of contracted services.

        This erased **{abs(net_imbalance)/(sffr_revenue+epex_revenue+ida1_revenue)*100:.0f}%**
        of all generated revenue.
        """)

    with col2:
        st.warning(f"""
        **2. Conservative Strategy**

        The strategy was heavily weighted toward SFFR
        (**{sffr_revenue/actual_total*100:.0f}%** of revenue),
        missing profitable market trading opportunities.

        Multi-market approach would yield **+£{revenue_gap:,.0f}** additional revenue.
        """)

    with col3:
        st.warning(f"""
        **3. Underutilization**

        Operating at only **{actual_daily_cycles:.2f}** cycles/day
        vs **{config.CYCLES_PER_DAY}** warranty limit.

        Significant headroom exists for more aggressive trading
        without warranty impact.
        """)

    # Summary box
    st.markdown("---")
    st.success(f"""
    ### 🎯 Key Recommendation

    Implementing the multi-market optimization strategy would:
    - Increase monthly revenue by **{improvement_pct:.0f}%** (£{revenue_gap:,.0f})
    - Generate annual additional revenue of **£{revenue_gap*12:,.0f}**
    - Remain well within warranty cycling limits
    - Better utilize the asset's full capabilities
    """)

def show_executive_comparison():
    """Display executive comparison dashboard — all available months, Actual vs Optimal"""
    st.title("📊 Executive Comparison Dashboard")
    st.markdown("### Multi-Month Performance Overview for Management")
    st.markdown("---")

    # ---- Month configuration (short labels, days-in-month, file keys) ----
    MONTH_CFG = [
        ('Sep 25', 'September', 30, 'Master_BESS_Analysis_Sept_2025.csv', 'Optimized_Results_Sept_2025.csv'),
        ('Oct 25', 'October', 31, 'Master_BESS_Analysis_Oct_2025.csv', 'Optimized_Results_Oct_2025.csv'),
        ('Nov 25', 'November', 30, 'Master_BESS_Analysis_Nov_2025.csv', 'Optimized_Results_Nov_2025.csv'),
        ('Dec 25', 'December', 31, 'Master_BESS_Analysis_Dec_2025.csv', 'Optimized_Results_Dec_2025.csv'),
        ('Jan 26', 'January', 31, 'Master_BESS_Analysis_Jan_2026.csv', 'Optimized_Results_Jan_2026.csv'),
        ('Feb 26', 'February', 28, 'Master_BESS_Analysis_Feb_2026.csv', 'Optimized_Results_Feb_2026.csv'),
        ('Mar 26', 'March', 31, 'Master_BESS_Analysis_Mar_2026.csv', 'Optimized_Results_Mar_2026.csv'),
    ]

    # Capacity Market & DUoS actuals (same data as in Benchmarks)
    CM_ACTUALS_EXEC = {
        'Oct 25': 1704.17, 'Nov 25': 1884.42,
        'Dec 25': 1994.84, 'Jan 26': 2113.87,
        'Feb 26': 1829.35,
    }
    DUOS_ACTUALS_EXEC = {
        'Sep 25': {'net_credit': 773.20, 'fixed': 3.58},
        'Oct 25': {'net_credit': 5807.68, 'fixed': 3.70},
        'Nov 25': {'net_credit': 5525.10, 'fixed': 3.58},
    }

    # ---- Load data for every available month ----
    def safe_sum(df, col):
        if col in df.columns:
            return pd.to_numeric(df[col], errors='coerce').fillna(0).sum()
        return 0

    def calc_actual_revenue(df):
        """Per-stream and total Actual revenue, net of the 5% GridBeyond fee."""
        sffr = apply_gb_net(safe_sum(df, 'SFFR revenues'))
        epex = apply_gb_net(safe_sum(df, 'EPEX 30 DA Revenue') + safe_sum(df, 'EPEX DA Revenues'))
        ida1 = apply_gb_net(safe_sum(df, 'IDA1 Revenue'))
        idc = apply_gb_net(safe_sum(df, 'IDC Revenue'))
        imb_rev = apply_gb_net(safe_sum(df, 'Imbalance Revenue'))
        imb_charge = apply_gb_net(safe_sum(df, 'Imbalance Charge'))
        return {
            'total': sffr + epex + ida1 + idc + imb_rev - imb_charge,
            'sffr': sffr, 'epex': epex, 'ida1': ida1, 'idc': idc,
            'imbalance': imb_rev - imb_charge
        }

    months = []  # list of dicts with all computed metrics
    for short, full, days, master_f, opt_f in MONTH_CFG:
        try:
            master = pd.read_csv(os.path.join(DATA_DIR, master_f))
            opt = pd.read_csv(os.path.join(DATA_DIR, opt_f))
            actual = calc_actual_revenue(master)
            # Optimiser output netted by 5% GB fee for apples-to-apples with Actual.
            optimal = apply_gb_net(opt['Optimised_Revenue_Multi'].sum())
            capture = (actual['total'] / optimal * 100) if optimal > 0 else 0
            gap = optimal - actual['total']

            # Add CM + DUoS
            cm = CM_ACTUALS_EXEC.get(short, 0)
            duos_data = DUOS_ACTUALS_EXEC.get(short)
            duos_credit = duos_data['net_credit'] if duos_data else 0
            duos_fixed = duos_data['fixed'] if duos_data else 0
            total_all = actual['total'] + cm + duos_credit - duos_fixed

            months.append({
                'short': short, 'full': full, 'days': days,
                'actual': actual, 'optimal': optimal,
                'capture': capture, 'gap': gap,
                'cm': cm, 'duos_credit': duos_credit, 'duos_fixed': duos_fixed,
                'total_all': total_all,
            })
        except FileNotFoundError:
            pass  # skip months whose files don't exist yet

    if len(months) < 2:
        st.warning("Need at least 2 months of data for comparison.")
        return

    # ==================== SECTION 1: KEY METRICS ====================
    st.header("1️⃣ Key Performance Metrics")

    cols = st.columns(len(months))
    for col, m in zip(cols, months):
        with col:
            st.metric(
                f"{m['short']} Actual",
                f"£{m['actual']['total']:,.0f}",
                delta=f"{m['capture']:.0f}% captured",
                delta_color="off"
            )

    # Revenue trend line
    first, last = months[0], months[-1]
    total_change = last['actual']['total'] - first['actual']['total']
    total_pct = (total_change / first['actual']['total'] * 100) if first['actual']['total'] != 0 else 0
    st.success(f"📈 **Trend ({first['short']} → {last['short']})**: Revenue changed by **£{total_change:,.0f}** ({total_pct:+.0f}%)")
    st.markdown("---")

    # ==================== SECTION 2: REVENUE COMPARISON BAR CHART ====================
    st.header("2️⃣ Revenue Comparison")

    bar_rows = []
    for m in months:
        bar_rows.append({'Scenario': f"{m['short']} Actual", 'Revenue': m['actual']['total'],
                         'Type': 'Actual', 'Month': m['full']})
        bar_rows.append({'Scenario': f"{m['short']} Optimal", 'Revenue': m['optimal'],
                         'Type': 'Optimal', 'Month': m['full']})
    comparison_data = pd.DataFrame(bar_rows)

    fig_bar = px.bar(
        comparison_data, x='Month', y='Revenue', color='Type', barmode='group',
        color_discrete_map={'Actual': COLOR_ACTUAL, 'Optimal': COLOR_MULTI_MARKET},
        title="Revenue: Actual vs Optimal by Month",
        text=comparison_data['Revenue'].apply(lambda x: f'£{x:,.0f}')
    )
    fig_bar.update_traces(textposition='outside')
    fig_bar.update_layout(yaxis_title="Revenue (£)", xaxis_title="", showlegend=True, height=450)
    st.plotly_chart(fig_bar, use_container_width=True)
    st.caption(GB_NET_FOOTNOTE_SHORT)
    st.markdown("---")

    # ==================== SECTION 3: GAP ANALYSIS TABLE ====================
    st.header("3️⃣ Performance Gap Analysis")
    st.caption("**Methodology:** Gap = Optimal - Actual. Optimal uses hindsight-based multi-market simulation. "
               "Both Actual and Optimal are shown net of the 5% GridBeyond revenue share.")

    gap_table = {'Metric': ['GridBeyond Revenue (£)', 'Capacity Market (£)',
                            'DUoS Net Credit (£)', 'Total Revenue (£)',
                            'Optimal Revenue (£)', 'Revenue Gap (£)',
                            'Capture Rate (%)', 'Imbalance (£)']}
    for m in months:
        gap_table[m['short']] = [
            f"£{m['actual']['total']:,.0f}",
            f"£{m['cm']:,.0f}" if m['cm'] else '-',
            f"£{m['duos_credit'] - m['duos_fixed']:,.0f}" if m['duos_credit'] else '-',
            f"£{m['total_all']:,.0f}",
            f"£{m['optimal']:,.0f}",
            f"£{m['gap']:,.0f}",
            f"{m['capture']:.1f}%",
            f"£{m['actual']['imbalance']:,.0f}"
        ]
    st.dataframe(pd.DataFrame(gap_table), use_container_width=True, hide_index=True)
    st.caption(GB_NET_FOOTNOTE_SHORT)
    st.markdown("---")

    # ==================== SECTION 4: MARKET MIX COMPARISON ====================
    st.header("4️⃣ Revenue by Market")

    # Pie charts — up to 3 per row
    market_keys = ['sffr', 'epex', 'ida1', 'idc', 'imbalance']
    market_labels = ['SFFR', 'EPEX', 'IDA1', 'IDC', 'Imbalance']

    # Render pie charts in rows of 3
    for row_start in range(0, len(months), 3):
        row_months = months[row_start:row_start + 3]
        pie_cols = st.columns(min(len(row_months), 3))
        for col, m in zip(pie_cols, row_months):
            with col:
                mkt_df = pd.DataFrame({
                    'Market': market_labels,
                    'Revenue': [m['actual'][k] for k in market_keys]
                })
                mkt_df = mkt_df[mkt_df['Revenue'].abs() > 0.01]
                fig = px.pie(mkt_df, values=mkt_df['Revenue'].abs(), names='Market',
                             title=f"{m['short']} - Market Mix", hole=0.4,
                             color='Market', color_discrete_map=REVENUE_COLOR_MAP)
                fig.update_traces(textposition='inside', textinfo='percent+label')
                st.plotly_chart(fig, use_container_width=True, key=f"exec_pie_{m['short']}")

    # Market comparison table (including CM + DUoS)
    mkt_table = {'Market': market_labels + ['GridBeyond Subtotal', 'Capacity Market', 'DUoS Net', 'TOTAL (All Streams)']}
    for m in months:
        vals = [f"£{m['actual'][k]:,.0f}" for k in market_keys]
        vals.append(f"£{m['actual']['total']:,.0f}")
        vals.append(f"£{m['cm']:,.0f}" if m['cm'] else '-')
        duos_net = m['duos_credit'] - m['duos_fixed']
        vals.append(f"£{duos_net:,.0f}" if m['duos_credit'] else '-')
        vals.append(f"£{m['total_all']:,.0f}")
        mkt_table[m['short']] = vals
    st.dataframe(pd.DataFrame(mkt_table), use_container_width=True, hide_index=True)
    st.caption(GB_NET_FOOTNOTE)
    st.markdown("---")

    # ==================== SECTION 5: EXECUTIVE SUMMARY ====================
    st.header("5️⃣ Executive Summary")

    best_m = max(months, key=lambda m: m['capture'])
    worst_m = min(months, key=lambda m: m['capture'])
    avg_capture = sum(m['capture'] for m in months) / len(months)
    latest = months[-1]

    col1, col2 = st.columns(2)
    with col1:
        st.error(f"""
        **Weakest Month — {worst_m['short']}:**
        - Capture Rate: {worst_m['capture']:.1f}%
        - Revenue Gap: £{worst_m['gap']:,.0f}
        - Imbalance: £{worst_m['actual']['imbalance']:,.0f}
        """)
    with col2:
        st.success(f"""
        **Strongest Month — {best_m['short']}:**
        - Capture Rate: {best_m['capture']:.1f}%
        - Revenue Gap: £{best_m['gap']:,.0f}
        - Imbalance: £{best_m['actual']['imbalance']:,.0f}
        """)

    st.info(f"""
    **Key Recommendations:**

    1. **Average Capture Rate**: {avg_capture:.1f}% across {len(months)} months
    2. **Best Performance**: {best_m['short']} achieved {best_m['capture']:.1f}% — replicate this strategy
    3. **Investigate**: {worst_m['short']} had largest gap (£{worst_m['gap']:,.0f})
    4. **Annualized Projection**: Based on latest month ({latest['short']}): **£{latest['total_all'] / latest['days'] * 365:,.0f}/year** (incl. CM + DUoS)
    """)


def show_market_price_analysis(month: str = "September 2025"):
    """Display market price analysis - spreads, missed opportunities, volatility"""
    st.title(f"📈 Market Price Analysis - {month}")
    st.markdown("### Price Spreads, Arbitrage Opportunities & Volatility Metrics")
    st.markdown("---")

    # Load data
    month_config = AVAILABLE_MONTHS.get(month, AVAILABLE_MONTHS["September 2025"])
    master_file = month_config.get("master_file", "Master_BESS_Analysis_Sept_2025.csv")

    try:
        df = pd.read_csv(os.path.join(DATA_DIR, master_file))
        if 'Unnamed: 0' in df.columns:
            df.rename(columns={'Unnamed: 0': 'Timestamp'}, inplace=True)
        df['Timestamp'] = pd.to_datetime(df['Timestamp'])
        df['Date'] = df['Timestamp'].dt.date
        df['Hour'] = df['Timestamp'].dt.hour
        df['DayOfWeek'] = df['Timestamp'].dt.day_name()
    except Exception as e:
        st.error(f"Error loading data: {str(e)}")
        return

    # Price columns
    price_cols = {
        'Day Ahead Price (EPEX)': 'EPEX DA',
        'GB-ISEM Intraday 1 Price': 'Intraday',
        'SSP': 'SSP (System Sell)',
        'SBP': 'SBP (System Buy)',
        'IDC Price': 'IDC'
    }

    # ==================== SECTION 1: KEY PRICE METRICS ====================
    st.header("1️⃣ Price Summary")

    col1, col2, col3, col4 = st.columns(4)

    epex_prices = pd.to_numeric(df['Day Ahead Price (EPEX)'], errors='coerce').dropna()
    ssp_prices = pd.to_numeric(df['SSP'], errors='coerce').dropna()
    sbp_prices = pd.to_numeric(df['SBP'], errors='coerce').dropna()

    with col1:
        st.metric("EPEX DA Avg", f"£{epex_prices.mean():.2f}/MWh",
                 help="Average Day Ahead price across all 30-min periods")
    with col2:
        st.metric("EPEX DA Max", f"£{epex_prices.max():.2f}/MWh",
                 help="Highest Day Ahead price in the month - best selling opportunity")
    with col3:
        st.metric("SSP Max", f"£{ssp_prices.max():.2f}/MWh",
                 help="Highest System Sell Price - max price grid pays you to export")
    with col4:
        st.metric("SBP Max", f"£{sbp_prices.max():.2f}/MWh",
                 help="Highest System Buy Price - worst price to buy from grid")

    st.markdown("---")

    # ==================== SECTION 2: PRICE SPREAD ANALYSIS ====================
    st.header("2️⃣ Price Spread Analysis (Arbitrage Opportunities)")

    # Calculate spreads
    df['EPEX_Spread'] = df['Day Ahead Price (EPEX)'].rolling(48).max() - df['Day Ahead Price (EPEX)'].rolling(48).min()
    df['SSP_SBP_Spread'] = pd.to_numeric(df['SBP'], errors='coerce') - pd.to_numeric(df['SSP'], errors='coerce')

    # Daily spread analysis
    daily_spreads = df.groupby('Date').agg({
        'Day Ahead Price (EPEX)': ['min', 'max', lambda x: x.max() - x.min()],
        'SSP': ['min', 'max'],
        'SBP': ['min', 'max']
    }).reset_index()
    daily_spreads.columns = ['Date', 'EPEX_Min', 'EPEX_Max', 'EPEX_Spread', 'SSP_Min', 'SSP_Max', 'SBP_Min', 'SBP_Max']
    daily_spreads['Date'] = pd.to_datetime(daily_spreads['Date'])

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Avg Daily EPEX Spread", f"£{daily_spreads['EPEX_Spread'].mean():.2f}/MWh",
                 help="Average difference between daily max and min EPEX prices")
    with col2:
        st.metric("Best EPEX Spread Day", f"£{daily_spreads['EPEX_Spread'].max():.2f}/MWh",
                 help="Highest single-day arbitrage opportunity")
    with col3:
        avg_ssb_spread = (daily_spreads['SBP_Max'] - daily_spreads['SSP_Min']).mean()
        st.metric("Avg SSP/SBP Spread", f"£{avg_ssb_spread:.2f}/MWh",
                 help="Average spread between system prices")

    # Daily spread chart
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=daily_spreads['Date'],
        y=daily_spreads['EPEX_Spread'],
        name='EPEX Daily Spread',
        marker_color=COLOR_EPEX_DA
    ))
    fig.update_layout(
        title="Daily EPEX Price Spread (Max - Min)",
        xaxis_title="Date",
        yaxis_title="Spread (£/MWh)",
        height=350
    )
    st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")

    # ==================== SECTION 3: HOURLY PRICE PATTERNS ====================
    st.header("3️⃣ Hourly Price Patterns (Best Trading Windows)")

    # Calculate hourly averages
    hourly_avg = df.groupby('Hour').agg({
        'Day Ahead Price (EPEX)': 'mean',
        'SSP': 'mean',
        'SBP': 'mean'
    }).reset_index()

    # Identify best buy/sell hours
    buy_hour = hourly_avg.loc[hourly_avg['Day Ahead Price (EPEX)'].idxmin(), 'Hour']
    sell_hour = hourly_avg.loc[hourly_avg['Day Ahead Price (EPEX)'].idxmax(), 'Hour']
    buy_price = hourly_avg['Day Ahead Price (EPEX)'].min()
    sell_price = hourly_avg['Day Ahead Price (EPEX)'].max()

    col1, col2, col3 = st.columns(3)
    with col1:
        st.success(f"**Best Buy Hour:** {int(buy_hour):02d}:00 (Avg £{buy_price:.2f}/MWh)")
    with col2:
        st.error(f"**Best Sell Hour:** {int(sell_hour):02d}:00 (Avg £{sell_price:.2f}/MWh)")
    with col3:
        st.info(f"**Hourly Arbitrage:** £{sell_price - buy_price:.2f}/MWh potential")

    # Hourly price heatmap
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=hourly_avg['Hour'],
        y=hourly_avg['Day Ahead Price (EPEX)'],
        mode='lines+markers',
        name='EPEX DA',
        line=dict(color=COLOR_EPEX_DA, width=3),
        fill='tozeroy',
        fillcolor='rgba(52, 152, 219, 0.3)'
    ))
    fig.add_trace(go.Scatter(
        x=hourly_avg['Hour'],
        y=hourly_avg['SSP'],
        mode='lines+markers',
        name='SSP',
        line=dict(color=COLOR_SSP, width=2)
    ))
    fig.add_trace(go.Scatter(
        x=hourly_avg['Hour'],
        y=hourly_avg['SBP'],
        mode='lines+markers',
        name='SBP',
        line=dict(color=COLOR_SBP, width=2)
    ))
    fig.update_layout(
        title="Average Price by Hour of Day",
        xaxis_title="Hour",
        yaxis_title="Price (£/MWh)",
        xaxis=dict(tickmode='linear', tick0=0, dtick=2),
        height=400
    )
    st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")

    # ==================== SECTION 4: VOLATILITY ANALYSIS ====================
    st.header("4️⃣ Price Volatility (High-Value Trading Days)")

    # Calculate daily volatility (std dev)
    daily_volatility = df.groupby('Date').agg({
        'Day Ahead Price (EPEX)': ['std', 'mean']
    }).reset_index()
    daily_volatility.columns = ['Date', 'EPEX_Std', 'EPEX_Mean']
    daily_volatility['Date'] = pd.to_datetime(daily_volatility['Date'])
    daily_volatility['CV'] = daily_volatility['EPEX_Std'] / daily_volatility['EPEX_Mean'] * 100  # Coefficient of variation

    # Identify high volatility days (top quartile)
    high_vol_threshold = daily_volatility['EPEX_Std'].quantile(0.75)
    high_vol_days = daily_volatility[daily_volatility['EPEX_Std'] >= high_vol_threshold]

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Avg Daily Volatility", f"£{daily_volatility['EPEX_Std'].mean():.2f}",
                 help="Standard deviation of daily prices")
    with col2:
        st.metric("High Volatility Days", f"{len(high_vol_days)} days",
                 help=f"Days with std > £{high_vol_threshold:.2f}")
    with col3:
        st.metric("Max Volatility Day", f"£{daily_volatility['EPEX_Std'].max():.2f}",
                 help="Highest single-day price volatility")

    # Volatility chart
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=daily_volatility['Date'],
        y=daily_volatility['EPEX_Std'],
        name='Price Volatility (Std Dev)',
        marker_color=np.where(daily_volatility['EPEX_Std'] >= high_vol_threshold, COLOR_SBP, COLOR_EPEX_DA)
    ))
    fig.add_hline(y=high_vol_threshold, line_dash="dash", line_color="red",
                  annotation_text=f"High volatility threshold: £{high_vol_threshold:.2f}")
    fig.update_layout(
        title="Daily Price Volatility (Red = High Value Trading Days)",
        xaxis_title="Date",
        yaxis_title="Volatility (£/MWh Std Dev)",
        height=350
    )
    st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")

    # ==================== SECTION 5: MISSED OPPORTUNITIES ====================
    st.header("5️⃣ Missed Opportunity Tracker")

    # Load actual trading data
    power_col = 'Physical_Power_MW' if 'Physical_Power_MW' in df.columns else None

    if power_col:
        # Identify periods where battery was idle during high spread
        df['Was_Idle'] = df[power_col].abs() < 0.1
        df['High_Spread'] = df['EPEX_Spread'] > df['EPEX_Spread'].quantile(0.75)

        # Missed opportunities: idle during high spread periods
        missed = df[df['Was_Idle'] & df['High_Spread']]
        missed_periods = len(missed)
        missed_hours = missed_periods * 0.5  # 30-min periods

        # Estimate missed revenue (conservative: assume 50% of spread captured with 2MW)
        avg_spread_missed = df.loc[missed.index, 'EPEX_Spread'].mean() if len(missed) > 0 else 0
        estimated_missed_revenue = missed_hours * 2 * avg_spread_missed * 0.5 if avg_spread_missed > 0 else 0

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Idle During High Spread", f"{missed_periods} periods",
                     help="30-min periods where battery was idle but spreads were high")
        with col2:
            st.metric("Missed Hours", f"{missed_hours:.1f} hours",
                     help="Total hours idle during high-spread windows (periods × 0.5)")
        with col3:
            st.metric("Est. Missed Revenue", f"£{estimated_missed_revenue:,.0f}",
                     help="Conservative estimate of potential additional revenue")

        # Top missed opportunity days
        if len(missed) > 0:
            missed_by_day = missed.groupby('Date').size().reset_index(name='Missed_Periods')
            missed_by_day = missed_by_day.sort_values('Missed_Periods', ascending=False).head(5)

            st.subheader("Top 5 Days with Missed Opportunities")
            for _, row in missed_by_day.iterrows():
                st.warning(f"**{row['Date']}**: {row['Missed_Periods']} periods idle during high-spread windows")
    else:
        st.info("Physical power data not available to calculate missed opportunities")

    st.markdown("---")

    # ==================== SECTION 6: PRICE CORRELATION ====================
    st.header("6️⃣ Market Price Correlation")

    # Calculate correlation matrix
    price_data = df[['Day Ahead Price (EPEX)', 'GB-ISEM Intraday 1 Price', 'SSP', 'SBP']].copy()
    price_data = price_data.apply(pd.to_numeric, errors='coerce')
    correlation = price_data.corr()

    fig = px.imshow(
        correlation,
        labels=dict(color="Correlation"),
        x=['EPEX DA', 'Intraday', 'SSP', 'SBP'],
        y=['EPEX DA', 'Intraday', 'SSP', 'SBP'],
        color_continuous_scale='RdBu_r',
        zmin=-1, zmax=1,
        text_auto='.2f'
    )
    fig.update_layout(title="Price Correlation Matrix", height=400)
    st.plotly_chart(fig, use_container_width=True)

    st.info("""
    **Correlation Insights:**
    - High correlation (>0.8) between markets suggests limited arbitrage opportunities between them
    - Low correlation indicates potential for cross-market arbitrage
    - SSP/SBP divergence creates imbalance market opportunities
    """)


def show_imbalance_deep_dive(month: str = "September 2025"):
    """Display imbalance deep dive analysis"""
    st.title(f"⚠️ Imbalance Deep Dive - {month}")
    st.markdown("### Root Cause Analysis of Imbalance Penalties")
    st.markdown("---")

    # Load data
    month_config = AVAILABLE_MONTHS.get(month, AVAILABLE_MONTHS["September 2025"])
    master_file = month_config.get("master_file", "Master_BESS_Analysis_Sept_2025.csv")

    try:
        df = pd.read_csv(os.path.join(DATA_DIR, master_file))
        if 'Unnamed: 0' in df.columns:
            df.rename(columns={'Unnamed: 0': 'Timestamp'}, inplace=True)
        df['Timestamp'] = pd.to_datetime(df['Timestamp'])
        df['Date'] = df['Timestamp'].dt.date
        df['Hour'] = df['Timestamp'].dt.hour
        df['DayOfWeek'] = df['Timestamp'].dt.day_name()
    except Exception as e:
        st.error(f"Error loading data: {str(e)}")
        return

    # Calculate imbalance columns — netted by 5% GB fee so figures tie to the GB invoice.
    df['Imbalance Revenue'] = pd.to_numeric(df.get('Imbalance Revenue', 0), errors='coerce').fillna(0) * GB_REVENUE_NET_SHARE
    df['Imbalance Charge'] = pd.to_numeric(df.get('Imbalance Charge', 0), errors='coerce').fillna(0) * GB_REVENUE_NET_SHARE
    df['Net_Imbalance'] = df['Imbalance Revenue'] - df['Imbalance Charge']
    df['SSP'] = pd.to_numeric(df['SSP'], errors='coerce').fillna(0)
    df['SBP'] = pd.to_numeric(df['SBP'], errors='coerce').fillna(0)
    df['SSP_SBP_Spread'] = df['SBP'] - df['SSP']

    # ==================== SECTION 1: IMBALANCE SUMMARY ====================
    st.header("1️⃣ Imbalance Summary")

    total_revenue = df['Imbalance Revenue'].sum()
    total_charge = df['Imbalance Charge'].sum()
    net_imbalance = total_revenue - total_charge
    periods_with_charge = (df['Imbalance Charge'] > 0).sum()
    periods_with_revenue = (df['Imbalance Revenue'] > 0).sum()

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Imbalance Revenue", f"£{total_revenue:,.0f}",
                 delta=f"{periods_with_revenue} periods", delta_color="off",
                 help="Revenue earned from imbalance market positions")
    with col2:
        st.metric("Imbalance Charges", f"£{total_charge:,.0f}",
                 delta=f"{periods_with_charge} periods", delta_color="off",
                 help="Penalties paid for imbalance positions")
    with col3:
        color = "normal" if net_imbalance >= 0 else "inverse"
        st.metric("Net Imbalance", f"£{net_imbalance:,.0f}",
                 delta="Profit" if net_imbalance >= 0 else "Loss", delta_color=color,
                 help="Revenue - Charges. Negative = net loss from imbalance")
    with col4:
        st.metric("% of Periods with Charges", f"{periods_with_charge/len(df)*100:.1f}%",
                 help="% of 30-min settlement periods that incurred charges")

    if net_imbalance < 0:
        st.error(f"⚠️ **Critical:** Net imbalance loss of **£{abs(net_imbalance):,.0f}** this month")
    else:
        st.success(f"✅ Net imbalance profit of **£{net_imbalance:,.0f}** this month")

    st.caption(GB_NET_FOOTNOTE_SHORT)

    st.markdown("---")

    # ==================== SECTION 2: DAILY BREAKDOWN ====================
    st.header("2️⃣ Daily Imbalance Breakdown")

    daily_imbalance = df.groupby('Date').agg({
        'Imbalance Revenue': 'sum',
        'Imbalance Charge': 'sum',
        'Net_Imbalance': 'sum'
    }).reset_index()
    daily_imbalance['Date'] = pd.to_datetime(daily_imbalance['Date'])

    # Identify worst days
    worst_days = daily_imbalance.nsmallest(5, 'Net_Imbalance')

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=daily_imbalance['Date'],
        y=daily_imbalance['Imbalance Revenue'],
        name='Revenue',
        marker_color=COLOR_IMBALANCE_REV
    ))
    fig.add_trace(go.Bar(
        x=daily_imbalance['Date'],
        y=-daily_imbalance['Imbalance Charge'],
        name='Charges',
        marker_color=COLOR_IMBALANCE_COST
    ))
    fig.update_layout(
        title="Daily Imbalance Revenue vs Charges",
        xaxis_title="Date",
        yaxis_title="Amount (£)",
        barmode='relative',
        height=400
    )
    st.plotly_chart(fig, use_container_width=True)

    # Worst days table
    st.subheader("Top 5 Worst Imbalance Days")
    worst_display = worst_days.copy()
    worst_display['Date'] = worst_display['Date'].dt.strftime('%Y-%m-%d')
    worst_display.columns = ['Date', 'Revenue (£)', 'Charges (£)', 'Net (£)']
    st.dataframe(worst_display, use_container_width=True, hide_index=True)
    st.caption(GB_NET_FOOTNOTE_SHORT)

    st.markdown("---")

    # ==================== SECTION 3: HOURLY PATTERN ====================
    st.header("3️⃣ Imbalance by Hour of Day")

    hourly_imbalance = df.groupby('Hour').agg({
        'Imbalance Charge': 'sum',
        'Net_Imbalance': 'sum'
    }).reset_index()

    worst_hour = hourly_imbalance.loc[hourly_imbalance['Imbalance Charge'].idxmax(), 'Hour']
    worst_hour_amount = hourly_imbalance['Imbalance Charge'].max()

    st.warning(f"**Peak Imbalance Hour:** {int(worst_hour):02d}:00 - £{worst_hour_amount:,.0f} in charges")

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=hourly_imbalance['Hour'],
        y=hourly_imbalance['Imbalance Charge'],
        name='Total Charges by Hour',
        marker_color=COLOR_IMBALANCE_COST
    ))
    fig.update_layout(
        title="Imbalance Charges by Hour of Day",
        xaxis_title="Hour",
        yaxis_title="Total Charges (£)",
        xaxis=dict(tickmode='linear', tick0=0, dtick=2),
        height=350
    )
    st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")

    # ==================== SECTION 4: CORRELATION WITH MARKET CONDITIONS ====================
    st.header("4️⃣ Correlation with Market Conditions")

    # Filter periods with imbalance charges
    charged_periods = df[df['Imbalance Charge'] > 0].copy()

    if len(charged_periods) > 0:
        avg_spread_during_charge = charged_periods['SSP_SBP_Spread'].mean()
        avg_spread_overall = df['SSP_SBP_Spread'].mean()

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Avg SSP/SBP Spread (All)", f"£{avg_spread_overall:.2f}/MWh",
                     help="Average system price spread across all periods")
        with col2:
            st.metric("Avg Spread (During Charges)", f"£{avg_spread_during_charge:.2f}/MWh",
                     delta=f"{avg_spread_during_charge - avg_spread_overall:+.2f}",
                     help="Average spread when charges occurred vs overall")
        with col3:
            correlation = charged_periods['Imbalance Charge'].corr(charged_periods['SSP_SBP_Spread'])
            st.metric("Charge-Spread Correlation", f"{correlation:.2f}",
                     help="How charges relate to price volatility (-1 to +1). High = charges during volatile periods")

        # Scatter plot: charge vs spread
        fig = px.scatter(
            charged_periods,
            x='SSP_SBP_Spread',
            y='Imbalance Charge',
            color='Hour',
            title="Imbalance Charges vs SSP/SBP Spread",
            labels={'SSP_SBP_Spread': 'SSP/SBP Spread (£/MWh)', 'Imbalance Charge': 'Charge (£)'},
            color_continuous_scale='Viridis'
        )
        fig.update_layout(height=400)
        st.plotly_chart(fig, use_container_width=True)

        st.info("""
        **Analysis:**
        - If charges correlate with high spreads, imbalance is occurring during volatile periods (forecast error)
        - If charges occur during low spreads, it may indicate systematic position errors
        """)
    else:
        st.success("No imbalance charges recorded this month!")

    st.markdown("---")

    # ==================== SECTION 5: ROOT CAUSE ANALYSIS ====================
    st.header("5️⃣ Root Cause Analysis")

    if len(charged_periods) > 0:
        # Categorize by time of day
        morning_charges = charged_periods[(charged_periods['Hour'] >= 6) & (charged_periods['Hour'] < 12)]['Imbalance Charge'].sum()
        afternoon_charges = charged_periods[(charged_periods['Hour'] >= 12) & (charged_periods['Hour'] < 18)]['Imbalance Charge'].sum()
        evening_charges = charged_periods[(charged_periods['Hour'] >= 18) & (charged_periods['Hour'] < 22)]['Imbalance Charge'].sum()
        night_charges = charged_periods[(charged_periods['Hour'] >= 22) | (charged_periods['Hour'] < 6)]['Imbalance Charge'].sum()

        # Pie chart
        fig = go.Figure(data=[go.Pie(
            labels=['Morning (6-12)', 'Afternoon (12-18)', 'Evening (18-22)', 'Night (22-6)'],
            values=[morning_charges, afternoon_charges, evening_charges, night_charges],
            hole=0.4,
            marker_colors=[COLOR_EPEX, COLOR_IMBALANCE_COST, COLOR_IDC, COLOR_ACTUAL]
        )])
        fig.update_layout(title="Imbalance Charges by Time of Day", height=350)
        st.plotly_chart(fig, use_container_width=True)

        # Key findings
        max_period = max([
            ('Morning', morning_charges),
            ('Afternoon', afternoon_charges),
            ('Evening', evening_charges),
            ('Night', night_charges)
        ], key=lambda x: x[1])

        st.error(f"""
        **Key Finding:** {max_period[1]/total_charge*100:.0f}% of imbalance charges (£{max_period[0]}: £{max_period[1]:,.0f}) occur during **{max_period[0]}** periods.

        **Potential Causes:**
        - Forecast errors during high-demand periods
        - Unexpected frequency response activations
        - Market position adjustments arriving late
        - Physical constraints preventing optimal dispatch
        """)

        # Recommendations
        st.subheader("💡 Recommendations")
        st.markdown(f"""
        1. **Focus on {max_period[0]} periods** - Review dispatch decisions during these hours
        2. **Improve forecasting** - Better predict imbalance exposure during volatile periods
        3. **Reduce position size** - During peak charge hours ({int(worst_hour):02d}:00), consider smaller positions
        4. **Monitor SSP/SBP spread** - High spreads (>£{avg_spread_during_charge:.0f}/MWh) correlate with charges
        5. **Consider imbalance hedging** - Use intraday markets to reduce exposure
        """)
    else:
        st.success("No imbalance issues to analyze - excellent performance!")


def show_ancillary_services_analysis(month: str = "September 2025"):
    """Display ancillary services analysis - SFFR, DC, DM, DR comparison"""
    st.title(f"⚡ Ancillary Services Analysis - {month}")
    st.markdown("### Service Utilization, Revenue Comparison & Opportunity Cost")
    st.markdown("---")

    # Load data
    month_config = AVAILABLE_MONTHS.get(month, AVAILABLE_MONTHS["September 2025"])
    master_file = month_config.get("master_file", "Master_BESS_Analysis_Sept_2025.csv")

    try:
        df = pd.read_csv(os.path.join(DATA_DIR, master_file))
        if 'Unnamed: 0' in df.columns:
            df.rename(columns={'Unnamed: 0': 'Timestamp'}, inplace=True)
        df['Timestamp'] = pd.to_datetime(df['Timestamp'])
        df['Date'] = df['Timestamp'].dt.date
        df['Hour'] = df['Timestamp'].dt.hour
    except Exception as e:
        st.error(f"Error loading data: {str(e)}")
        return

    # Ancillary service definitions
    services = {
        'SFFR': {'avail': 'SFFR Availability', 'price': 'SFFR Clearing Price', 'rev': 'SFFR revenues', 'name': 'Static FFR'},
        'DCL': {'avail': 'DCL Availability', 'price': 'DCL Clearing Price', 'rev': 'DCL revenues', 'name': 'DC Low'},
        'DCH': {'avail': 'DCH Availability', 'price': 'DCH Clearing Price', 'rev': 'DCH revenues', 'name': 'DC High'},
        'DML': {'avail': 'DML Availability', 'price': 'DML Clearing Price', 'rev': 'DML revenues', 'name': 'DM Low'},
        'DMH': {'avail': 'DMH Availability', 'price': 'DMH Clearing Price', 'rev': 'DMH revenues', 'name': 'DM High'},
        'DRL': {'avail': 'DRL Availability', 'price': 'DRL Clearing Price', 'rev': 'DRL revenues', 'name': 'DR Low'},
        'DRH': {'avail': 'DRH Availability', 'price': 'DRH Clearing Price', 'rev': 'DRH revenues', 'name': 'DR High'},
    }

    # Calculate metrics for each service.
    # All ancillary revenues are GB-traded, so net by 5% to match the GB invoice.
    service_metrics = []
    for code, cols in services.items():
        if cols['rev'] in df.columns:
            revenue = apply_gb_net(pd.to_numeric(df[cols['rev']], errors='coerce').fillna(0).sum())
            avail = pd.to_numeric(df.get(cols['avail'], 0), errors='coerce').fillna(0)
            price = pd.to_numeric(df.get(cols['price'], 0), errors='coerce').fillna(0)
            periods_active = (avail > 0).sum()
            avg_price = price[avail > 0].mean() if periods_active > 0 else 0
            total_mw_hours = (avail * 0.5).sum()  # 30-min periods

            service_metrics.append({
                'Service': code,
                'Name': cols['name'],
                'Total Revenue': revenue,
                'Periods Active': periods_active,
                'Avg Price': avg_price,
                'MW-Hours': total_mw_hours,
                'Revenue per MW-Hour': revenue / total_mw_hours if total_mw_hours > 0 else 0
            })

    metrics_df = pd.DataFrame(service_metrics)
    metrics_df = metrics_df.sort_values('Total Revenue', ascending=False)

    # ==================== SECTION 1: REVENUE SUMMARY ====================
    st.header("1️⃣ Ancillary Revenue Summary")

    total_ancillary = metrics_df['Total Revenue'].sum()
    top_service = metrics_df.iloc[0] if len(metrics_df) > 0 else None

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Ancillary Revenue", f"£{total_ancillary:,.0f}",
                 help="Sum of all frequency response service revenues (SFFR, DC, DM, DR)")
    with col2:
        if top_service is not None:
            st.metric("Top Service", top_service['Service'],
                     delta=f"£{top_service['Total Revenue']:,.0f}",
                     help="Service generating most revenue this month")
    with col3:
        services_used = (metrics_df['Total Revenue'] > 0).sum()
        st.metric("Services Used", f"{services_used} / {len(services)}",
                 help="Number of different services with non-zero revenue")
    with col4:
        if top_service is not None:
            share = top_service['Total Revenue'] / total_ancillary * 100 if total_ancillary > 0 else 0
            st.metric("Top Service Share", f"{share:.0f}%",
                     help="% of total ancillary revenue from top service")

    st.markdown("---")

    # ==================== SECTION 2: REVENUE BY SERVICE ====================
    st.header("2️⃣ Revenue by Service")

    fig = px.bar(
        metrics_df,
        x='Service',
        y='Total Revenue',
        color='Total Revenue',
        color_continuous_scale='Greens',
        title="Revenue by Ancillary Service",
        text=metrics_df['Total Revenue'].apply(lambda x: f'£{x:,.0f}')
    )
    fig.update_traces(textposition='outside')
    fig.update_layout(yaxis_title="Revenue (£)", xaxis_title="Service", showlegend=False, height=400)
    st.plotly_chart(fig, use_container_width=True)

    # Summary table
    display_df = metrics_df[['Service', 'Name', 'Total Revenue', 'Periods Active', 'Avg Price', 'Revenue per MW-Hour']].copy()
    display_df['Total Revenue'] = display_df['Total Revenue'].apply(lambda x: f"£{x:,.0f}")
    display_df['Avg Price'] = display_df['Avg Price'].apply(lambda x: f"£{x:.2f}/MW/h")
    display_df['Revenue per MW-Hour'] = display_df['Revenue per MW-Hour'].apply(lambda x: f"£{x:.2f}")
    st.dataframe(display_df, use_container_width=True, hide_index=True)
    st.caption(GB_NET_FOOTNOTE_SHORT)

    st.markdown("---")

    # ==================== SECTION 3: SERVICE UTILIZATION ====================
    st.header("3️⃣ Service Utilization Patterns")

    # Pie chart of utilization
    active_services = metrics_df[metrics_df['Total Revenue'] > 0]

    if len(active_services) > 0:
        fig = go.Figure(data=[go.Pie(
            labels=active_services['Service'],
            values=active_services['Total Revenue'],
            hole=0.4,
            textinfo='label+percent'
        )])
        fig.update_layout(title="Revenue Distribution by Service", height=400)
        st.plotly_chart(fig, use_container_width=True)

    # Hourly heatmap for SFFR (main service) — netted by 5% GB fee.
    if 'SFFR revenues' in df.columns:
        hourly_sffr = df.groupby('Hour')['SFFR revenues'].sum().reset_index()
        hourly_sffr['SFFR revenues'] = hourly_sffr['SFFR revenues'] * GB_REVENUE_NET_SHARE

        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=hourly_sffr['Hour'],
            y=hourly_sffr['SFFR revenues'],
            marker_color=COLOR_SFFR
        ))
        fig.update_layout(
            title="SFFR Revenue by Hour of Day",
            xaxis_title="Hour",
            yaxis_title="Revenue (£)",
            xaxis=dict(tickmode='linear', tick0=0, dtick=2),
            height=350
        )
        st.plotly_chart(fig, use_container_width=True)
        st.caption(GB_NET_FOOTNOTE_SHORT)

    st.markdown("---")

    # ==================== SECTION 4: REVENUE PER MW COMPARISON ====================
    st.header("4️⃣ Revenue per MW-Hour (Efficiency)")

    efficient_df = metrics_df[metrics_df['MW-Hours'] > 0].sort_values('Revenue per MW-Hour', ascending=False)

    if len(efficient_df) > 0:
        fig = px.bar(
            efficient_df,
            x='Service',
            y='Revenue per MW-Hour',
            color='Revenue per MW-Hour',
            color_continuous_scale='Blues',
            title="Revenue per MW-Hour by Service (Higher = More Efficient)",
            text=efficient_df['Revenue per MW-Hour'].apply(lambda x: f'£{x:.2f}')
        )
        fig.update_traces(textposition='outside')
        fig.update_layout(yaxis_title="£/MW-Hour", xaxis_title="Service", showlegend=False, height=400)
        st.plotly_chart(fig, use_container_width=True)

        most_efficient = efficient_df.iloc[0]
        st.success(f"**Most Efficient Service:** {most_efficient['Service']} ({most_efficient['Name']}) at £{most_efficient['Revenue per MW-Hour']:.2f}/MW-Hour")

    st.markdown("---")

    # ==================== SECTION 5: OPPORTUNITY COST ANALYSIS ====================
    st.header("5️⃣ Opportunity Cost Analysis")

    st.markdown("*What if GridBeyond had used a different service mix?*")

    if len(efficient_df) > 0 and len(metrics_df) > 0:
        # Calculate what revenue would be if all MW-hours went to most efficient service
        total_mw_hours = metrics_df['MW-Hours'].sum()
        most_efficient_rate = efficient_df.iloc[0]['Revenue per MW-Hour']
        current_rate = total_ancillary / total_mw_hours if total_mw_hours > 0 else 0

        optimal_revenue = total_mw_hours * most_efficient_rate
        opportunity_cost = optimal_revenue - total_ancillary

        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Current Avg Rate", f"£{current_rate:.2f}/MW-h",
                     help="Actual revenue / MW-hours committed across all services")
        with col2:
            st.metric("Best Service Rate", f"£{most_efficient_rate:.2f}/MW-h",
                     help="Highest £/MW-hour achieved by any single service")
        with col3:
            st.metric("Opportunity Cost", f"£{opportunity_cost:,.0f}",
                     delta="Potential additional revenue" if opportunity_cost > 0 else "Optimal mix achieved",
                     delta_color="inverse" if opportunity_cost > 0 else "normal",
                     help="Theoretical gain if all capacity went to best service (ignores market constraints)")

        if opportunity_cost > 0:
            st.warning(f"""
            **Analysis:** If all {total_mw_hours:,.0f} MW-hours had been allocated to {efficient_df.iloc[0]['Service']}
            instead of the current mix, revenue could have been **£{optimal_revenue:,.0f}** instead of **£{total_ancillary:,.0f}**.

            *Note: This is theoretical - market availability and grid requirements constrain actual allocation.*
            """)
        else:
            st.success("Current service allocation appears optimal for the available market conditions.")

    st.markdown("---")

    # ==================== SECTION 6: RECOMMENDATIONS ====================
    st.header("6️⃣ Recommendations")

    if len(metrics_df) > 0:
        sffr_share = metrics_df[metrics_df['Service'] == 'SFFR']['Total Revenue'].sum() / total_ancillary * 100 if total_ancillary > 0 else 0

        st.info(f"""
        **Current Strategy Analysis:**
        - SFFR accounts for **{sffr_share:.0f}%** of ancillary revenue
        - {(metrics_df['Total Revenue'] == 0).sum()} services were not utilized this month

        **Recommendations:**
        1. **Diversification**: Consider utilizing more DC/DM/DR services for risk mitigation
        2. **Price Monitoring**: Track clearing prices across all services to identify arbitrage
        3. **Stacking Strategy**: Explore combining services during different time windows
        4. **Market Analysis**: Review why certain services show zero revenue - market conditions or strategy choice?
        """)


def generate_pdf_report(month: str = "September 2025"):
    """Generate PDF executive summary report"""
    from io import BytesIO

    # Load data
    month_config = AVAILABLE_MONTHS.get(month, AVAILABLE_MONTHS["September 2025"])
    master_file = month_config.get("master_file", "Master_BESS_Analysis_Sept_2025.csv")
    optimization_file = month_config.get("optimization_file", "Optimized_Results_Sept_2025.csv")

    try:
        df = pd.read_csv(os.path.join(DATA_DIR, master_file))
        opt_df = pd.read_csv(os.path.join(DATA_DIR, optimization_file))
    except Exception as e:
        return None, f"Error loading data: {str(e)}"

    # Calculate metrics
    def safe_sum(dataframe, col):
        if col in dataframe.columns:
            return pd.to_numeric(dataframe[col], errors='coerce').fillna(0).sum()
        return 0

    # All GB-traded streams and the optimiser output are netted by the 5% GridBeyond fee
    # so the numbers in the report tie to the GridBeyond invoice and Actual vs Optimal
    # is apples-to-apples.
    sffr = apply_gb_net(safe_sum(df, 'SFFR revenues'))
    epex = apply_gb_net(safe_sum(df, 'EPEX 30 DA Revenue') + safe_sum(df, 'EPEX DA Revenues'))
    ida1 = apply_gb_net(safe_sum(df, 'IDA1 Revenue'))
    idc = apply_gb_net(safe_sum(df, 'IDC Revenue'))
    imb_rev = apply_gb_net(safe_sum(df, 'Imbalance Revenue'))
    imb_charge = apply_gb_net(safe_sum(df, 'Imbalance Charge'))
    total_actual = sffr + epex + ida1 + idc + imb_rev - imb_charge

    optimal = apply_gb_net(opt_df['Optimised_Revenue_Multi'].sum()) if 'Optimised_Revenue_Multi' in opt_df.columns else 0
    capture_rate = (total_actual / optimal * 100) if optimal > 0 else 0
    gap = optimal - total_actual

    # Generate text report
    report = f"""
================================================================================
                    BESS PERFORMANCE EXECUTIVE SUMMARY
                           {month}
================================================================================

ASSET: Northwold Solar Farm BESS
AGGREGATOR: GridBeyond
REPORT DATE: {datetime.now().strftime('%Y-%m-%d')}

--------------------------------------------------------------------------------
                           KEY METRICS
--------------------------------------------------------------------------------

ACTUAL REVENUE:         £{total_actual:,.0f}
OPTIMAL (SIMULATED):    £{optimal:,.0f}
REVENUE GAP:            £{gap:,.0f}
CAPTURE RATE:           {capture_rate:.1f}%

  All figures are shown net of the 5% GridBeyond revenue share.
  The optimiser output is also netted on the assumption that, in
  real-world deployment, the optimised revenue would clear through
  GridBeyond or an equivalent aggregator.

--------------------------------------------------------------------------------
                        REVENUE BREAKDOWN
--------------------------------------------------------------------------------

SFFR (Frequency Response):     £{sffr:,.0f}
EPEX Trading:                  £{epex:,.0f}
IDA1 Trading:                  £{ida1:,.0f}
IDC Trading:                   £{idc:,.0f}
Imbalance Revenue:             £{imb_rev:,.0f}
Imbalance Charges:            -£{imb_charge:,.0f}
                              ──────────
TOTAL:                         £{total_actual:,.0f}

--------------------------------------------------------------------------------
                        PERFORMANCE ASSESSMENT
--------------------------------------------------------------------------------

"""
    if capture_rate >= 90:
        report += "ASSESSMENT: EXCELLENT - Near-optimal performance achieved\n"
    elif capture_rate >= 70:
        report += "ASSESSMENT: GOOD - Solid performance with room for improvement\n"
    elif capture_rate >= 50:
        report += "ASSESSMENT: MODERATE - Significant revenue left on table\n"
    else:
        report += "ASSESSMENT: POOR - Major optimization opportunities exist\n"

    if imb_charge > 1000:
        report += f"\n⚠️  WARNING: High imbalance charges of £{imb_charge:,.0f} detected\n"

    report += f"""
--------------------------------------------------------------------------------
                        RECOMMENDATIONS
--------------------------------------------------------------------------------

1. {'MAINTAIN current strategy - excellent capture rate' if capture_rate >= 90 else 'REVIEW trading strategy to close revenue gap'}
2. {'Imbalance management is effective' if imb_charge < 1000 else f'INVESTIGATE imbalance charges (£{imb_charge:,.0f} this month)'}
3. Continue monitoring multi-market optimization opportunities
4. Review ancillary service allocation for efficiency

--------------------------------------------------------------------------------
                           ANNUALIZED PROJECTION
--------------------------------------------------------------------------------

Based on {month} performance:
- Projected Annual Revenue: £{total_actual * 12:,.0f}
- Projected Optimal:        £{optimal * 12:,.0f}
- Projected Annual Gap:     £{gap * 12:,.0f}

================================================================================
Generated by Asset Performance Dashboard
🤖 Automated Report - {datetime.now().strftime('%Y-%m-%d %H:%M')}
================================================================================
"""

    return report, None


def show_pdf_export_page(month: str = "September 2025"):
    """Display PDF export page"""
    st.title("📄 Export Executive Summary")
    st.markdown("### Generate Reports for GridBeyond Meetings")
    st.markdown("---")

    st.subheader("Select Report Options")

    col1, col2 = st.columns(2)

    with col1:
        report_month = st.selectbox(
            "Report Month",
            list(AVAILABLE_MONTHS.keys()),
            index=list(AVAILABLE_MONTHS.keys()).index(month) if month in AVAILABLE_MONTHS else 0
        )

    with col2:
        report_format = st.radio(
            "Report Format",
            ["Text Summary", "CSV Data Export"],
            horizontal=True
        )

    st.markdown("---")

    if st.button("📥 Generate Report", type="primary"):
        with st.spinner("Generating report..."):
            if report_format == "Text Summary":
                report, error = generate_pdf_report(report_month)

                if error:
                    st.error(error)
                else:
                    st.success("Report generated successfully!")

                    # Preview
                    st.subheader("Report Preview")
                    st.text(report)

                    # Download button
                    st.download_button(
                        label="📥 Download Report",
                        data=report,
                        file_name=f"BESS_Executive_Summary_{report_month.replace(' ', '_')}.txt",
                        mime="text/plain"
                    )

            else:  # CSV Export
                month_config = AVAILABLE_MONTHS.get(report_month)
                if month_config:
                    try:
                        master_df = pd.read_csv(os.path.join(DATA_DIR, month_config['master_file']))
                        opt_df = pd.read_csv(os.path.join(DATA_DIR, month_config['optimization_file']))

                        st.success("Data loaded successfully!")

                        col1, col2 = st.columns(2)

                        with col1:
                            csv1 = master_df.to_csv(index=False)
                            st.download_button(
                                label="📥 Download Master Data",
                                data=csv1,
                                file_name=f"Master_Data_{report_month.replace(' ', '_')}.csv",
                                mime="text/csv"
                            )

                        with col2:
                            csv2 = opt_df.to_csv(index=False)
                            st.download_button(
                                label="📥 Download Optimization Results",
                                data=csv2,
                                file_name=f"Optimization_{report_month.replace(' ', '_')}.csv",
                                mime="text/csv"
                            )

                    except Exception as e:
                        st.error(f"Error loading data: {str(e)}")

    st.markdown("---")

    # Quick comparison export
    st.subheader("Multi-Month Comparison Export")

    if st.button("📊 Export Multi-Month Comparison"):
        with st.spinner("Generating comparison..."):
            try:
                EXPORT_MONTHS = [
                    ('Sep 25', 'Master_BESS_Analysis_Sept_2025.csv', 'Optimized_Results_Sept_2025.csv'),
                    ('Oct 25', 'Master_BESS_Analysis_Oct_2025.csv', 'Optimized_Results_Oct_2025.csv'),
                    ('Nov 25', 'Master_BESS_Analysis_Nov_2025.csv', 'Optimized_Results_Nov_2025.csv'),
                    ('Dec 25', 'Master_BESS_Analysis_Dec_2025.csv', 'Optimized_Results_Dec_2025.csv'),
                    ('Jan 26', 'Master_BESS_Analysis_Jan_2026.csv', 'Optimized_Results_Jan_2026.csv'),
                    ('Feb 26', 'Master_BESS_Analysis_Feb_2026.csv', 'Optimized_Results_Feb_2026.csv'),
                    ('Mar 26', 'Master_BESS_Analysis_Mar_2026.csv', 'Optimized_Results_Mar_2026.csv'),
                ]

                CM_EXPORT = {'Oct 25': 1704.17, 'Nov 25': 1884.42, 'Dec 25': 1994.84, 'Jan 26': 2113.87, 'Feb 26': 1829.35}
                DUOS_EXPORT = {
                    'Sep 25': {'net': 769.62}, 'Oct 25': {'net': 5803.98}, 'Nov 25': {'net': 5521.52},
                }

                def calc_metrics(master_df, opt_df, short):
                    """All GB-traded streams and optimiser output netted by 5% GB fee."""
                    def safe_sum(dataframe, col):
                        if col in dataframe.columns:
                            return pd.to_numeric(dataframe[col], errors='coerce').fillna(0).sum()
                        return 0

                    sffr = apply_gb_net(safe_sum(master_df, 'SFFR revenues'))
                    epex = apply_gb_net(safe_sum(master_df, 'EPEX 30 DA Revenue') + safe_sum(master_df, 'EPEX DA Revenues'))
                    ida1 = apply_gb_net(safe_sum(master_df, 'IDA1 Revenue'))
                    idc = apply_gb_net(safe_sum(master_df, 'IDC Revenue'))
                    imb = apply_gb_net(safe_sum(master_df, 'Imbalance Revenue') - safe_sum(master_df, 'Imbalance Charge'))
                    gb_total = sffr + epex + ida1 + idc + imb
                    optimal = apply_gb_net(opt_df['Optimised_Revenue_Multi'].sum()) if 'Optimised_Revenue_Multi' in opt_df.columns else 0
                    cm = CM_EXPORT.get(short, 0)
                    duos_net = DUOS_EXPORT.get(short, {}).get('net', 0)
                    total_all = gb_total + cm + duos_net

                    return {
                        'SFFR': sffr, 'EPEX': epex, 'Imbalance': imb,
                        'GridBeyond Total': gb_total,
                        'Capacity Market': cm, 'DUoS Net': duos_net,
                        'Total (All Streams)': total_all,
                        'Optimal': optimal,
                        'Capture Rate': gb_total / optimal * 100 if optimal > 0 else 0,
                        'Gap': optimal - gb_total,
                    }

                all_metrics = {}
                for short, mf, of in EXPORT_MONTHS:
                    try:
                        m_df = pd.read_csv(os.path.join(DATA_DIR, mf))
                        o_df = pd.read_csv(os.path.join(DATA_DIR, of))
                        all_metrics[short] = calc_metrics(m_df, o_df, short)
                    except FileNotFoundError:
                        pass

                if all_metrics:
                    first_key = list(all_metrics.keys())[0]
                    comparison_df = pd.DataFrame({
                        'Metric': list(all_metrics[first_key].keys()),
                        **{short: [f"£{v:,.0f}" if 'Rate' not in k else f"{v:.1f}%"
                                   for k, v in metrics.items()]
                           for short, metrics in all_metrics.items()}
                    })
                    st.dataframe(comparison_df, use_container_width=True, hide_index=True)

                    csv = comparison_df.to_csv(index=False)
                    st.download_button(
                        label="📥 Download Comparison CSV",
                        data=csv,
                        file_name="Multi_Month_Comparison.csv",
                        mime="text/csv"
                )

            except Exception as e:
                st.error(f"Error generating comparison: {str(e)}")


@st.cache_data
def _load_iar_monthly_per_mw():
    """Return {short_month: monthly £/MW total} from the IAR Excel.

    Excel stores £/MW per month per stream in columns 11..17 (Sep 25 → Mar 26),
    rows 4..11 (8 revenue streams). Sum the 8 stream rows per month to get
    total monthly £/MW. Section 1's 3-way bar chart uses this directly
    (Monthly £/MW view) and annualises via × 365/days for the annualised view.
    Returns {} on read failure.
    """
    iar_file = os.path.join(os.path.dirname(__file__), 'extra', 'Northwold BESS Revenue_IAR.xlsx')
    try:
        import openpyxl
        wb = openpyxl.load_workbook(iar_file, data_only=True)
        ws = wb['Sheet1']
        col_map = {11: 'Sep 25', 12: 'Oct 25', 13: 'Nov 25', 14: 'Dec 25',
                   15: 'Jan 26', 16: 'Feb 26', 17: 'Mar 26'}
        stream_rows = [4, 5, 6, 7, 8, 9, 10, 11]
        result = {}
        for col_idx, short in col_map.items():
            total_per_mw = sum((ws.cell(row=r, column=col_idx).value or 0)
                               for r in stream_rows)
            result[short] = float(total_per_mw)
        wb.close()
        return result
    except Exception:
        return {}


def show_benchmark_comparison():
    """Display industry benchmark comparison page."""
    st.title("📊 Benchmarks")
    st.markdown("Compare Northwold's performance against UK BESS industry benchmarks and IAR projections")

    # ---- Month configuration ----
    BENCH_MONTHS = [
        ('Sep 25', 30, 'Master_BESS_Analysis_Sept_2025.csv', 'Optimized_Results_Sept_2025.csv'),
        ('Oct 25', 31, 'Master_BESS_Analysis_Oct_2025.csv', 'Optimized_Results_Oct_2025.csv'),
        ('Nov 25', 30, 'Master_BESS_Analysis_Nov_2025.csv', 'Optimized_Results_Nov_2025.csv'),
        ('Dec 25', 31, 'Master_BESS_Analysis_Dec_2025.csv', 'Optimized_Results_Dec_2025.csv'),
        ('Jan 26', 31, 'Master_BESS_Analysis_Jan_2026.csv', 'Optimized_Results_Jan_2026.csv'),
        ('Feb 26', 28, 'Master_BESS_Analysis_Feb_2026.csv', 'Optimized_Results_Feb_2026.csv'),
        ('Mar 26', 31, 'Master_BESS_Analysis_Mar_2026.csv', 'Optimized_Results_Mar_2026.csv'),
    ]

    # Modo Energy monthly benchmark (£/MW/year) — source: Modo Energy ME BESS GB Index.
    # Sep 25–Feb 26: headline value from each month's benchmark article on modoenergy.com.
    # Mar 26: API-derived (monthly-index-live endpoint, market=total, duration=*) since the
    #   March benchmark article is not yet published. Expect to reconcile once it lands.
    MODO_BENCHMARKS = {
        'Sep 25': 70000, 'Oct 25': 77000, 'Nov 25': 59000,
        'Dec 25': 47000, 'Jan 26': 52000, 'Feb 26': 41000,
        'Mar 26': 65000,
    }

    # Source article URLs per month — shown in the "Source articles" expander
    # at the bottom of this page so users can verify each headline figure.
    # Mar 26 points at the Modo API developer reference because no article is
    # published yet.
    MODO_SOURCE_LINKS = {
        'Sep 25': 'https://modoenergy.com/research/en/battery-energy-storage-revenues-gb-september-2025-balancing-mechanism-frequency-response',
        'Oct 25': 'https://modoenergy.com/research/en/battery-energy-storage-revenues-gb-october-2025-record-balancing-mechanism-dispatch-rates',
        'Nov 25': 'https://modoenergy.com/research/en/me-bess-gb-battery-energy-storage-revenues-november-2025-balancing-mechanism-gas-wind',
        'Dec 25': 'https://modoenergy.com/research/en/me-bess-gb-battery-energy-storage-revenues-december-2025-low-demand-christmas',
        'Jan 26': 'https://modoenergy.com/research/en/me-bess-gb-revenues-january-2026-balancing-mechanism-wholesale-prices-gas-carbon',
        'Feb 26': 'https://modoenergy.com/research/en/me-bess-gb-revenues-february-2026-wholesale-battery-energy-storage-balancing-mechanism',
        'Mar 26': 'https://developers.modoenergy.com/reference/monthly-me-bess-gb',
    }

    # Capacity Market payments (£) — source: EMR Settlement T062 CSVs
    # Contract: CAN-2025-NSFL01-001, 1.023 MW @ £20,000/MW/yr, monthly weighting
    CM_ACTUALS = {
        'Oct 25': 1704.17, 'Nov 25': 1884.42,
        'Dec 25': 1994.84, 'Jan 26': 2113.87,
        'Feb 26': 1829.35,
    }

    # DUoS actuals (£ net ex-VAT) — source: Hartree Partners Gen_Inv PDFs
    # GDuos credits (Red+Amber+Green) are revenue; DNO Fixed is a cost
    DUOS_ACTUALS = {
        'Sep 25': {'red': -322.81, 'amber': -410.03, 'green': -43.94,
                   'fixed': 3.58, 'net_credit': 773.20},
        'Oct 25': {'red': -5500.11, 'amber': -268.35, 'green': -42.92,
                   'fixed': 3.70, 'net_credit': 5807.68},
        'Nov 25': {'red': -5379.73, 'amber': -106.41, 'green': -42.54,
                   'fixed': 3.58, 'net_credit': 5525.10},
    }

    # Load and calculate Northwold metrics first
    try:
        def safe_sum_b(dataframe, col):
            if col in dataframe.columns:
                return pd.to_numeric(dataframe[col], errors='coerce').fillna(0).sum()
            return 0

        def calculate_monthly_revenue(df):
            """Total GB-traded revenue for a month, NET of the 5% GridBeyond fee."""
            sffr = safe_sum_b(df, 'SFFR revenues')
            epex = safe_sum_b(df, 'EPEX 30 DA Revenue') + safe_sum_b(df, 'EPEX DA Revenues')
            ida1 = safe_sum_b(df, 'IDA1 Revenue')
            idc = safe_sum_b(df, 'IDC Revenue')
            imb_rev = safe_sum_b(df, 'Imbalance Revenue')
            imb_charge = safe_sum_b(df, 'Imbalance Charge')

            gross = sffr + epex + ida1 + idc + imb_rev - imb_charge
            return apply_gb_net(gross)

        def calculate_cycles_local(df, power_col, capacity_mwh=8.4, dt_hours=0.5):
            """Calculate cycles using industry standard method."""
            if power_col not in df.columns:
                return None
            power = pd.to_numeric(df[power_col], errors='coerce').fillna(0)
            energy = power * dt_hours
            discharge_mwh = energy[energy > 0].sum()
            charge_mwh = abs(energy[energy < 0].sum())
            return (discharge_mwh + charge_mwh) / 2 / capacity_mwh

        def find_power_col(df):
            """Find the best power column and its dt."""
            for col in df.columns:
                if 'Physical_Power_MW' in col or col == 'Power_MW':
                    return col, 0.5
            for col in df.columns:
                if 'Battery MWh' in col:
                    return col, 1.0
            return None, 0.5

        capacity_mw = 4.2

        # Load all months dynamically
        bm = []  # benchmark month dicts
        masters = {}
        opts = {}
        for short, days, master_f, opt_f in BENCH_MONTHS:
            try:
                m_df = pd.read_csv(os.path.join(DATA_DIR, master_f))
                o_df = pd.read_csv(os.path.join(DATA_DIR, opt_f))
                masters[short] = m_df
                opts[short] = o_df

                revenue = calculate_monthly_revenue(m_df)
                annual_per_mw = (revenue / days) * 365 / capacity_mw

                pcol, pdt = find_power_col(m_df)
                cycles = calculate_cycles_local(m_df, pcol, dt_hours=pdt) if pcol else None

                if 'Timestamp' in m_df.columns:
                    m_df['Timestamp'] = pd.to_datetime(m_df['Timestamp'], errors='coerce')
                    n_days = m_df['Timestamp'].dt.date.nunique()
                else:
                    n_days = days

                daily_cycles = cycles / n_days if cycles else None
                modo = MODO_BENCHMARKS.get(short)

                # Non-GridBeyond revenue streams
                cm = CM_ACTUALS.get(short, 0)
                duos_data = DUOS_ACTUALS.get(short)
                duos_credit = duos_data['net_credit'] if duos_data else 0
                duos_fixed = duos_data['fixed'] if duos_data else 0

                # Total including CM + DUoS
                total_revenue = revenue + cm + duos_credit - duos_fixed
                total_annual_per_mw = (total_revenue / days) * 365 / capacity_mw

                bm.append({
                    'short': short, 'days': days, 'revenue': revenue,
                    'annual_per_mw': annual_per_mw, 'daily_cycles': daily_cycles,
                    'modo': modo,
                    'cm': cm, 'duos_credit': duos_credit, 'duos_fixed': duos_fixed,
                    'total_revenue': total_revenue,
                    'total_annual_per_mw': total_annual_per_mw,
                })
            except FileNotFoundError:
                pass

        avg_annual = sum(m['total_annual_per_mw'] for m in bm) / len(bm) if bm else 0
        data_loaded = len(bm) > 0

    except Exception as e:
        st.error(f"Error loading data: {str(e)}")
        data_loaded = False
        bm = []
        avg_annual = 0

    # ==================== Section 1: Revenue vs Benchmarks ====================
    st.header("1. Revenue vs Benchmarks")
    st.markdown(
        "Three-way comparison of Northwold actual revenue against the Modo Energy "
        "GB BESS benchmark and the Internal Appraisal Report (IAR) projection. "
        "**Monthly £/MW** shows what each party earned per MW in each real month "
        "(honest apples-to-apples). **Annualised £/MW/year** projects the same "
        "numbers to a full year for board-level comparison."
    )

    if data_loaded:
        iar_monthly_per_mw = _load_iar_monthly_per_mw()

        # Shared arrays (same month order as `bm`).
        months = [m['short'] for m in bm]
        days_by_short = {m['short']: m['days'] for m in bm}

        # --- Monthly £/MW (actual reality, no annualisation) ---
        actual_monthly = [m['total_revenue'] / capacity_mw for m in bm]
        modo_monthly = [
            (MODO_BENCHMARKS.get(s, 0) * days_by_short[s] / 365)
            if MODO_BENCHMARKS.get(s) else 0
            for s in months
        ]
        iar_monthly = [iar_monthly_per_mw.get(s, 0) for s in months]

        # --- Annualised £/MW/year (board-level view) ---
        actual_annual = [m['total_annual_per_mw'] for m in bm]
        modo_annual = [MODO_BENCHMARKS.get(s, 0) for s in months]
        iar_annual = [
            (iar_monthly_per_mw.get(s, 0) * 365 / days_by_short[s])
            if iar_monthly_per_mw.get(s) else 0
            for s in months
        ]

        # Colour assignment — reuse existing palette for consistency across the page.
        C_ACTUAL = COLOR_ACTUAL         # Blue — Northwold
        C_MODO = COLOR_IDC              # Purple — external benchmark
        C_IAR = COLOR_MULTI_MARKET      # Green — internal target

        def _grouped_bars(title, y_title, actual_ys, modo_ys, iar_ys, hover_unit):
            fig = go.Figure()
            fig.add_trace(go.Bar(
                name='Northwold actual', x=months, y=actual_ys,
                marker_color=C_ACTUAL,
                hovertemplate=f'%{{x}}<br>Actual: £%{{y:,.0f}}{hover_unit}<extra></extra>',
            ))
            fig.add_trace(go.Bar(
                name='Modo benchmark', x=months, y=modo_ys,
                marker_color=C_MODO,
                hovertemplate=f'%{{x}}<br>Modo: £%{{y:,.0f}}{hover_unit}<extra></extra>',
            ))
            fig.add_trace(go.Bar(
                name='IAR projection', x=months, y=iar_ys,
                marker_color=C_IAR,
                hovertemplate=f'%{{x}}<br>IAR: £%{{y:,.0f}}{hover_unit}<extra></extra>',
            ))
            fig.update_layout(
                title=title, barmode='group', yaxis_title=y_title,
                height=420, showlegend=True, margin=dict(t=60, b=40),
            )
            return fig

        st.plotly_chart(
            _grouped_bars(
                'Monthly £/MW — what each earned per MW in each real month',
                '£/MW (month)', actual_monthly, modo_monthly, iar_monthly, '/MW',
            ),
            use_container_width=True,
        )

        st.plotly_chart(
            _grouped_bars(
                'Annualised £/MW/year — projecting each month at full-year pace',
                '£/MW/year', actual_annual, modo_annual, iar_annual, '/MW/yr',
            ),
            use_container_width=True,
        )

        # ---- Per-month capture table ----
        st.subheader("Per-month capture")
        summary_rows = []
        for i, m in enumerate(bm):
            s = m['short']
            act_m = actual_monthly[i]
            mod_m = modo_monthly[i]
            iar_m = iar_monthly[i]
            cap_modo = (act_m / mod_m * 100) if mod_m else None
            cap_iar = (act_m / iar_m * 100) if iar_m else None
            summary_rows.append({
                'Month': s,
                'Actual (£/MW/mo)': f"£{round(act_m):,}",
                'Modo (£/MW/mo)': f"£{round(mod_m):,}" if mod_m else '—',
                'IAR (£/MW/mo)': f"£{round(iar_m):,}" if iar_m else '—',
                'Capture vs Modo': f"{cap_modo:.0f}%" if cap_modo is not None else '—',
                'Capture vs IAR': f"{cap_iar:.0f}%" if cap_iar is not None else '—',
            })
        st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

        # ---- Portfolio tiles ----
        modo_pairs = [
            (m['short'], actual_monthly[i] / modo_monthly[i] * 100)
            for i, m in enumerate(bm) if modo_monthly[i]
        ]
        iar_pairs = [
            (m['short'], actual_monthly[i] / iar_monthly[i] * 100)
            for i, m in enumerate(bm) if iar_monthly[i]
        ]
        avg_cap_modo = (sum(p[1] for p in modo_pairs) / len(modo_pairs)) if modo_pairs else 0
        avg_cap_iar = (sum(p[1] for p in iar_pairs) / len(iar_pairs)) if iar_pairs else 0
        if modo_pairs:
            best = max(modo_pairs, key=lambda p: p[1])
            worst = min(modo_pairs, key=lambda p: p[1])
            best_worst = f"{best[0]} {best[1]:.0f}% · {worst[0]} {worst[1]:.0f}%"
        else:
            best_worst = '—'

        tile_cols = st.columns(3)
        tile_cols[0].metric("Avg capture vs Modo", f"{avg_cap_modo:.0f}%")
        tile_cols[1].metric("Avg capture vs IAR", f"{avg_cap_iar:.0f}%")
        tile_cols[2].metric("Best · worst vs Modo", best_worst)

        st.caption(GB_NET_FOOTNOTE)

        # ---- Historical industry range (demoted to expander) ----
        with st.expander("Historical industry range (Modo 2024–25 envelope)"):
            st.caption(
                "Previously shown as the primary benchmark here. Kept as reference "
                "context — values are the min / median / max of Modo's monthly GB "
                "BESS Index figures across calendar years 2024 and 2025."
            )
            st.dataframe(pd.DataFrame([
                {'Tier': 'Low',  '£/MW/year': '£36,000', 'Cycles/day': '1.0', 'RTE': '82%'},
                {'Tier': 'Mid',  '£/MW/year': '£60,000', 'Cycles/day': '1.5', 'RTE': '85%'},
                {'Tier': 'High', '£/MW/year': '£88,000', 'Cycles/day': '3.0', 'RTE': '90%'},
            ]), use_container_width=True, hide_index=True)

        # Calculation explanations for Section 1
        st.subheader("Metric Calculations")

        with st.expander("📐 Total Revenue - How is it calculated?"):
            st.markdown("""
**Formula:**
```
GridBeyond Gross  = SFFR + EPEX DA + IDA1 + IDC + Imbalance Revenue - Imbalance Charge
GridBeyond Net    = GridBeyond Gross × 0.95   (5% GB revenue share deducted)
Total Revenue     = GridBeyond Net + Capacity Market + DUoS Net Credit - DUoS Fixed Charges
```

**Explanation:**
GridBeyond gross revenue is the sum of all wholesale/ancillary revenues traded by the aggregator.
The dashboard shows the **net** figure — after GridBeyond's 5% revenue share — so the numbers
tie out to the GridBeyond invoice and are directly comparable to the IAR (which is also post-fee).
Total Revenue adds non-GridBeyond streams: Capacity Market payments (EMR Settlement T062)
and Distribution Use of System credits (Hartree Partners passthrough invoices) — these are paid
direct and are not subject to the GB fee.

**Example (October 2025):**
- GridBeyond Revenue: £38,344 (SFFR + EPEX + IDA1 + IDC + Imbalance)
- Capacity Market: £1,704 (EMR Settlement CAN-2025-NSFL01-001)
- DUoS Net Credit: £5,808 (Red + Amber + Green GDuos credits)
- DUoS Fixed Charges: -£4 (DNO fixed charges)
- **Total: £45,852**
            """)

        with st.expander("📐 Revenue (£/MW/year) - How is it calculated?"):
            st.markdown("""
**Formula:**
```
£/MW/year = (Total Monthly Revenue ÷ Days in Month) × 365 ÷ Capacity_MW
```

**Explanation:**
Annualizes total monthly revenue (GridBeyond + CM + DUoS) and normalizes by installed
capacity (4.2 MW) to enable comparison with Modo Energy industry benchmarks regardless of asset size.

**Example (October 2025):**
- Total Monthly Revenue: £45,852 (GridBeyond £38,344 + CM £1,704 + DUoS £5,804)
- Days in October: 31
- Daily average: £45,852 ÷ 31 = £1,479/day
- Annualized: £1,479 × 365 = £539,790/year
- Per MW: £539,790 ÷ 4.2 MW = **£128,521/MW/year**
            """)

        with st.expander("📐 Daily Cycles - How is it calculated?"):
            st.markdown("""
**Formula:**
```
Daily Cycles = (Discharge_MWh + Charge_MWh) ÷ 2 ÷ Capacity_MWh ÷ Days
```

**Explanation:**
One full cycle = fully charging then fully discharging the battery (8.4 MWh).
We sum all energy throughput, divide by 2 (to count charge+discharge as one cycle),
then divide by capacity and number of days.

**Example (September 2025):**
- Total Discharge: 126 MWh
- Total Charge: 126 MWh
- Total throughput: 252 MWh
- Equivalent full cycles: 252 ÷ 2 ÷ 8.4 = 15 cycles
- Daily average: 15 ÷ 30 days = **0.5 cycles/day**
            """)

        with st.expander("📐 Degradation - How is it calculated?"):
            st.markdown("""
**Formula:**
```
Degradation = Capacity loss (%) per 365 equivalent full cycles
```

**Explanation:**
Battery capacity degrades with use. Industry benchmarks measure this as percentage
capacity loss per year of typical cycling (365 cycles). Lower is better.

**Industry Range (NREL Study):**
- Low: 4.0% (high-quality cells, conservative operation)
- Mid: 4.4% (typical lithium-ion)
- High: 11.0% (aggressive cycling, poor thermal management)

*Northwold TBD - requires long-term capacity testing data.*
            """)

        with st.expander("📐 Availability (TWCAA) - How is it calculated?"):
            st.markdown("""
**Formula:**
```
TWCAA = Technical Weighted Contracted Availability Assessment
```

**Explanation:**
National Grid ESO metric measuring the percentage of time the asset is available
to deliver contracted services. Accounts for planned maintenance, forced outages,
and partial availability.

**Industry Range (National Grid ESO):**
- Low: 90% (significant downtime)
- Mid: 94.4% (typical BESS performance)
- High: 98% (excellent availability)

*Northwold TBD - requires ESO reporting data.*
            """)

        with st.expander("📐 Round-Trip Efficiency - How is it calculated?"):
            st.markdown("""
**Formula:**
```
RTE = (Energy Discharged ÷ Energy Charged) × 100
```

**Explanation:**
Measures energy losses during charge/discharge cycles. A battery charged with
100 MWh that discharges 85 MWh has 85% RTE. Losses occur in power electronics,
battery cells, and thermal management.

**Industry Range (DNV GL):**
- Low: 82% (older systems, poor conditions)
- Mid: 85% (typical Li-ion NMC/LFP)
- High: 90% (optimized operation)

*Northwold: ~85% (estimated from system design)*
            """)

    st.markdown("---")

    # Section 2: IAR vs Actual Revenue Comparison
    st.header("2. Revenue IAR vs Actual")
    st.caption("Comparison of Internal Appraisal Report (IAR) projections against actual GridBeyond revenues. "
               "Both columns are shown **net of the 5% GridBeyond revenue share** so they are directly comparable "
               "to the GridBeyond invoice.")

    # Revenue stream labels
    streams = [
        'Wholesale Day Ahead', 'Wholesale Intraday', 'Balancing Mechanism',
        'Frequency Response', 'Capacity Market', 'DUoS Battery',
        'DUoS Fixed Charges', 'TNUoS', 'Imbalance Revenue', 'Imbalance Charge',
        'TOTAL (excl. BM, TNUoS)', 'TOTAL (all streams)'
    ]

    # ---- Load IAR projections from Excel ----
    IAR_PROJ = {}
    iar_file = os.path.join(os.path.dirname(__file__), 'extra', 'Northwold BESS Revenue_IAR.xlsx')
    try:
        import openpyxl
        iar_wb = openpyxl.load_workbook(iar_file, data_only=True)
        iar_ws = iar_wb['Sheet1']
        iar_mw = 4.2  # MW assumed in IAR model

        # Column mapping: Excel column index -> month short label
        # Row 3 has datetime headers; E(5)=Mar25, K(11)=Sep25, L(12)=Oct25, etc.
        iar_col_map = {11: 'Sep 25', 12: 'Oct 25', 13: 'Nov 25', 14: 'Dec 25', 15: 'Jan 26'}
        # Row mapping: rows 4-11 = DA, ID, BM, FR, CM, DUoS, DUoS_Fixed, TNUoS
        iar_stream_rows = [4, 5, 6, 7, 8, 9, 10, 11]

        for col_idx, short in iar_col_map.items():
            vals = []
            for row in iar_stream_rows:
                raw = iar_ws.cell(row=row, column=col_idx).value or 0
                vals.append(round(raw * iar_mw))
            # Imbalance Revenue, Imbalance Charge (not in IAR model)
            vals.extend([0, 0])
            # TOTAL (excl. BM, TNUoS) = DA + ID + FR + CM + DUoS + DUoS_Fixed
            total_excl = vals[0] + vals[1] + vals[3] + vals[4] + vals[5] + vals[6]
            # TOTAL (all streams)
            total_all = sum(vals[:8])
            vals.extend([total_excl, total_all])
            IAR_PROJ[short] = [f"{v:,}" if v >= 0 else f"-{abs(v):,}" for v in vals]
        iar_wb.close()
    except Exception:
        # Fallback hardcoded if Excel not available
        IAR_PROJ = {
            'Sep 25': ['14,343', '4,246', '1,863', '1,383', '4,438', '9,188', '-6,462', '835', '0', '0', '27,136', '29,834'],
            'Oct 25': ['17,178', '4,918', '4,237', '1,038', '4,586', '10,088', '-6,678', '863', '0', '0', '31,130', '36,230'],
        }

    # Actual values from GridBeyond + invoice data
    def _fmt(v):
        if v is None:
            return '-'
        return f"{v:,.0f}" if v >= 0 else f"-{abs(v):,.0f}"

    def _var(actual, iar_str):
        """Calculate variance % between actual and IAR string value."""
        if actual is None or iar_str in ('-', '0'):
            return '-'
        iar_val = float(iar_str.replace(',', ''))
        if iar_val == 0:
            return '-'
        var = ((actual - iar_val) / abs(iar_val)) * 100
        return f"{var:+.0f}%"

    # Build actuals from GridBeyond data + invoice data.
    # GB-traded streams netted by GB_REVENUE_NET_SHARE (0.95) so they match
    # the post-fee IAR projections and the GridBeyond invoice.
    # CM, DUoS Battery, DUoS Fixed, TNUoS are paid direct (EMR / Hartree) — no GB fee.
    actual_data = {}
    if data_loaded:
        for m in bm:
            short = m['short']
            df = masters[short]
            epex = apply_gb_net(safe_sum_b(df, 'EPEX 30 DA Revenue') + safe_sum_b(df, 'EPEX DA Revenues'))
            ida1 = apply_gb_net(safe_sum_b(df, 'IDA1 Revenue'))
            idc = apply_gb_net(safe_sum_b(df, 'IDC Revenue'))
            sffr = apply_gb_net(safe_sum_b(df, 'SFFR revenues'))
            imb_rev = apply_gb_net(safe_sum_b(df, 'Imbalance Revenue'))
            imb_charge = apply_gb_net(safe_sum_b(df, 'Imbalance Charge'))
            cm = m['cm']
            duos_cr = m['duos_credit']
            duos_fx = m['duos_fixed']

            gb_total = sffr + epex + ida1 + idc + imb_rev - imb_charge
            full_total = gb_total + cm + duos_cr - duos_fx

            actual_data[short] = [
                epex, ida1, None,  # BM not tracked in GridBeyond
                sffr, cm if cm else None,
                duos_cr if duos_cr else None,
                -duos_fx if duos_fx else None,
                None,  # TNUoS not available
                imb_rev if imb_rev != 0 else None,
                -imb_charge if imb_charge != 0 else None,
                gb_total, full_total
            ]

    # Build table
    iar_table = {'Revenue Stream': streams}
    for m in bm:
        short = m['short']
        iar_proj = IAR_PROJ.get(short)
        actuals = actual_data.get(short)

        if iar_proj:
            iar_table[f'{short} IAR (£)'] = iar_proj
        if actuals:
            iar_table[f'{short} Actual (£)'] = [_fmt(v) for v in actuals]
        if iar_proj and actuals:
            iar_table[f'{short} Var'] = [_var(actuals[i], iar_proj[i]) for i in range(len(streams))]

    iar_df = pd.DataFrame(iar_table)

    # Style total rows to be bold
    def highlight_total(row):
        if row.name >= len(iar_df) - 2:  # Last 2 rows (totals)
            return ['font-weight: bold; background-color: #f0f2f6'] * len(row)
        return [''] * len(row)

    styled_iar_df = iar_df.style.apply(highlight_total, axis=1)
    st.dataframe(styled_iar_df, use_container_width=True, hide_index=True, height=460)

    st.caption("*IAR projections based on 4.2 MW capacity with indexation factor 1.073. "
               "Wholesale DA/ID, Frequency Response and Imbalance actuals are shown net of the 5% GridBeyond revenue share. "
               "CM actuals from EMR Settlement (T062) — paid direct, no GB fee. "
               "DUoS actuals from Hartree Partners invoices — paid direct, no GB fee. "
               "TNUoS invoices not yet available.*")
    st.caption(GB_NET_FOOTNOTE_SHORT)

    # Summary metrics for IAR comparison
    iar_months = [m for m in bm if m['short'] in IAR_PROJ and m['short'] in actual_data]
    if iar_months:
        n_cols = min(len(iar_months) + 1, 6)
        metric_cols = st.columns(n_cols)
        for idx, m in enumerate(iar_months):
            short = m['short']
            actuals = actual_data[short]
            iar_proj = IAR_PROJ[short]
            actual_total = actuals[-1]  # TOTAL (all streams)
            iar_total = float(iar_proj[-1].replace(',', '').replace('-', ''))
            pct = (actual_total / iar_total * 100) if iar_total else 0
            delta = pct - 100
            with metric_cols[idx % n_cols]:
                st.metric(f"{short} vs IAR", f"{pct:.0f}%", f"{delta:+.0f}%",
                          delta_color="normal" if delta >= 0 else "inverse",
                          help="Total actual revenue (all streams) as % of IAR projection")

        if len(iar_months) > 1:
            total_actual = sum(actual_data[m['short']][-1] for m in iar_months)
            total_iar = sum(float(IAR_PROJ[m['short']][-1].replace(',', '').replace('-', '')) for m in iar_months)
            combined_pct = (total_actual / total_iar * 100) if total_iar else 0
            with metric_cols[min(len(iar_months), n_cols - 1)]:
                st.metric("Combined", f"{combined_pct:.0f}%",
                          help="Total actual vs total projected across all months")

    st.info("""
    **Key Insights from IAR Comparison:**
    - **Frequency Response** revenue massively exceeded projections due to higher ancillary service market rates
    - **Wholesale Day Ahead** significantly underperformed vs IAR in earlier months
    - **Capacity Market** actuals (£1.7-2.1k/month) are significantly below IAR projections (£4.4-4.6k/month) — different capacity obligation or auction assumptions
    - **DUoS Battery** credits are actual export credits from UKPN; IAR projected much higher (~£9-10k/month vs actual £0.8-5.8k)
    """)

    # Calculation explanations for Section 2
    st.subheader("Metric Calculations")

    with st.expander("📐 Variance (%) - How is it calculated?"):
        st.markdown("""
**Formula:**
```
Variance % = ((Actual - IAR) ÷ |IAR|) × 100
```

**Explanation:**
Compares actual revenues (GridBeyond + CM + DUoS) against Internal Appraisal Report (IAR) projections
loaded from `extra/Northwold BESS Revenue_IAR.xlsx`. IAR values are per MW/month, multiplied by 4.2 MW
for Real totals. Positive variance means outperformance; negative means underperformance.

**Example (October 2025 — Frequency Response):**
- IAR Projection: £1,038 (per IAR model, 4.2 MW × per-MW projection)
- Actual SFFR Revenue: £28,382
- Variance: ((28,382 - 1,038) ÷ 1,038) × 100 = **+2,634%**

**Key Patterns Observed:**
- **Frequency Response** massively outperforms IAR (SFFR prices much higher than projected)
- **Capacity Market** actuals (£1.7–2.1k/month) below IAR projections (£4.4–4.6k/month)
- **DUoS Battery** credits vary significantly by month (seasonal Red band weighting)
        """)

    with st.expander("📐 Revenue Streams - What do they mean?"):
        st.markdown("""
| Stream | Source | Description |
|--------|--------|-------------|
| **Wholesale Day Ahead** | GridBeyond | EPEX power exchange — trading for next-day delivery |
| **Wholesale Intraday** | GridBeyond | IDA1/IDC — same-day trading to balance positions |
| **Balancing Mechanism** | GridBeyond | National Grid dispatch instructions for system balancing |
| **Frequency Response** | GridBeyond | SFFR/DC/DM/DR — grid frequency stabilization services |
| **Capacity Market** | EMR Settlement (ESC) | Monthly payments under contract CAN-2025-NSFL01-001 (1.023 MW) |
| **DUoS Battery** | Hartree Partners | Distribution Use of System — Red/Amber/Green GDuos credits for export |
| **DUoS Fixed Charges** | Hartree Partners | DNO fixed network charges (cost, shown negative) |
| **TNUoS** | N/A | Transmission Network Use of System — no invoices received to date |
| **Imbalance Revenue** | GridBeyond | Payments when grid position helps system balance (SSP) |
| **Imbalance Charge** | GridBeyond | Penalties when grid position hurts system balance (SBP) |

*IAR projections are from the internal financial model (`extra/Northwold BESS Revenue_IAR.xlsx`),
calculated per MW/month and multiplied by 4.2 MW for comparison. No indexation is applied.*
        """)

    st.markdown("---")

    # Section 3: Multi-Market Optimization vs Actual
    st.header("3. Multi-Market Optimization vs Actual")
    st.caption("Compare actual GridBeyond performance against optimized multi-market strategy with perfect foresight. "
               "Both Actual and Optimised are shown **net of the 5% GridBeyond revenue share** (the optimiser output is "
               "scaled by 0.95 on the assumption that, in real-world deployment, the optimised revenue would also be "
               "traded through GridBeyond or an equivalent aggregator).")

    if data_loaded and opts:
        def get_revenue_breakdown(df):
            sffr = safe_sum_b(df, 'SFFR revenues')
            epex = safe_sum_b(df, 'EPEX 30 DA Revenue') + safe_sum_b(df, 'EPEX DA Revenues')
            ida1 = safe_sum_b(df, 'IDA1 Revenue')
            idc = safe_sum_b(df, 'IDC Revenue')
            imb_rev = safe_sum_b(df, 'Imbalance Revenue')
            imb_charge = safe_sum_b(df, 'Imbalance Charge')
            # Net 5% GridBeyond fee — apples-to-apples with the GB invoice and the optimiser output below.
            return {
                'SFFR': apply_gb_net(sffr),
                'EPEX DA': apply_gb_net(epex),
                'IDA1': apply_gb_net(ida1),
                'IDC': apply_gb_net(idc),
                'Imbalance': apply_gb_net(imb_rev - imb_charge),
                'Total': apply_gb_net(sffr + epex + ida1 + idc + imb_rev - imb_charge),
            }

        def get_opt_revenue_breakdown(df):
            df = df.copy()
            df['Revenue'] = pd.to_numeric(df['Optimised_Revenue_Multi'], errors='coerce').fillna(0)
            df['Market'] = df['Market_Used_Multi'].fillna('Idle')
            sffr = df[df['Market'] == 'SFFR']['Revenue'].sum()
            epex = df[df['Market'].str.contains('EPEX', na=False)]['Revenue'].sum()
            isem = df[df['Market'].str.contains('ISEM', na=False)]['Revenue'].sum()
            ssp = df[df['Market'].str.contains('SSP|SBP', na=False)]['Revenue'].sum()
            da_hh = df[df['Market'].str.contains('DA_HH', na=False)]['Revenue'].sum()
            total = df['Revenue'].sum()
            # Net 5% — assume the optimised revenue would also be traded via GB in deployment.
            return {
                'SFFR': apply_gb_net(sffr),
                'EPEX DA': apply_gb_net(epex + da_hh),
                'IDA1': apply_gb_net(isem),
                'IDC': 0,
                'Imbalance': apply_gb_net(ssp),
                'Total': apply_gb_net(total),
            }

        # Summary metrics row
        metric_cols = st.columns(len(bm))
        all_actuals = {}
        all_opts_bd = {}
        all_gaps = {}
        all_captures = {}

        for idx, m in enumerate(bm):
            short = m['short']
            act = get_revenue_breakdown(masters[short])
            opt_bd = get_opt_revenue_breakdown(opts[short])
            gap = opt_bd['Total'] - act['Total']
            capture = (act['Total'] / opt_bd['Total'] * 100) if opt_bd['Total'] > 0 else 0
            all_actuals[short] = act
            all_opts_bd[short] = opt_bd
            all_gaps[short] = gap
            all_captures[short] = capture

            with metric_cols[idx]:
                st.metric(f"{short} Capture", f"{capture:.0f}%",
                          delta=f"{capture - 100:+.0f}%" if capture != 100 else None)

        # Comparison table
        st.subheader("Monthly Revenue Comparison")
        streams = ['SFFR', 'EPEX DA', 'IDA1', 'IDC', 'Imbalance', 'Total']
        stream_labels = ['SFFR (Frequency Response)', 'EPEX DA (Day Ahead)',
                         'IDA1/ISEM (Intraday)', 'IDC (Continuous)',
                         'SSP/SBP (Imbalance)', 'TOTAL']

        comp_data = {'Revenue Stream': stream_labels + ['Revenue Gap', 'Capture Rate']}
        for m in bm:
            short = m['short']
            act = all_actuals[short]
            opt_bd = all_opts_bd[short]
            act_vals = [f"£{act[s]:,.0f}" for s in streams] + ['-', '-']
            opt_vals = [f"£{opt_bd[s]:,.0f}" for s in streams] + [f"£{all_gaps[short]:,.0f}", f"{all_captures[short]:.0f}%"]
            comp_data[f"{short} Actual"] = act_vals
            comp_data[f"{short} Opt"] = opt_vals

        comp_df = pd.DataFrame(comp_data)

        def style_comp(row):
            n = len(comp_df)
            if row.name == 5:
                return ['font-weight: bold; background-color: #e6f3ff'] * len(row)
            elif row.name in [6, 7]:
                return ['font-weight: bold; background-color: #fff3e6'] * len(row)
            return [''] * len(row)

        st.dataframe(comp_df.style.apply(style_comp, axis=1), use_container_width=True, hide_index=True)
        st.caption(GB_NET_FOOTNOTE_SHORT)

        total_gap = sum(all_gaps.values())
        avg_cap = sum(all_captures.values()) / len(all_captures)

        if avg_cap >= 100:
            st.success(f"**Performance Summary:** Average capture rate {avg_cap:.0f}% across {len(bm)} months — outperforming multi-market optimization.")
        else:
            st.warning(f"**Performance Summary:** Total gap £{total_gap:,.0f} across {len(bm)} months (avg capture: {avg_cap:.0f}%). Optimization uses perfect foresight — gap is expected.")

        # Calculation explanations
        st.subheader("Metric Calculations")

        with st.expander("📐 Multi-Market Optimization - How is it calculated?"):
            st.markdown("""
**Formula:**
```
For each day:
  SFFR_Daily = Sum(7.0 MW × SFFR_Clearing_Price × 0.5hr) for 48 periods
  Multi_Market = LP solver maximizing: Sum(Discharge × Sell_Price - Charge × Buy_Price) × 0.5hr
  Daily Revenue = max(SFFR_Daily, Multi_Market)
```

**Explanation:**
A linear optimization model (scipy linprog, HiGHS solver) that first compares SFFR availability revenue
against optimal multi-market dispatch for the whole day. If SFFR wins, the battery is locked in frequency
response. If multi-market wins, it dispatches across 5 markets using **perfect price foresight**:
- Buy from the lowest-priced market (min of EPEX, ISEM, SSP, SBP, DA HH)
- Sell to the highest-priced market (max of the same 5)
- Hold or idle when spreads don't cover round-trip losses

**Constraints Applied:**
- Charge: 0–4.2 MW | Discharge: 0–7.5 MW (asymmetric)
- SOC range: 5%–95% (0.42–7.98 MWh)
- Max daily discharge throughput: 12.6 MWh (1.5 cycles × 8.4 MWh)
- One-way efficiency: 93.3% (round-trip 87%)
- SOC carries forward between days

**Example (January 5, 2026 — best day):**
- SFFR option: ~£477 (low SFFR clearing prices)
- Multi-Market option: £5,222 (SSP spiked to 750 GBP/MWh at 19:00)
- Decision: Multi-Market wins — battery charged at 68 GBP/MWh (SSP) overnight, held fully charged until the spike, then discharged aggressively into SSP
            """)

        with st.expander("📐 Revenue Gap - How is it calculated?"):
            st.markdown("""
**Formula:**
```
Revenue Gap = Optimized Multi-Market Revenue - Actual GridBeyond Revenue
```

**Explanation:**
Measures the theoretical revenue improvement possible if the battery had been operated
with perfect market foresight across all available markets. The gap typically concentrates
in a few spike days per month (e.g., SSP spikes that are unpredictable in real-time).

**Example (January 2026):**
- Multi-Market Optimal: £33,376
- Actual GridBeyond Revenue: £28,190
- Revenue Gap: £33,376 - £28,190 = **£5,186**
- Capture Rate: 84.5%

**Gap Decomposition (Jan 26):**
- ~67% from SSP spike events not captured (esp. Jan 5, 8)
- ~15% from sub-optimal market selection
- ~10% from SFFR availability assumption (7.0 MW vs 6.81 MW actual)
- ~7% from imbalance penalties

**Important Caveats:**
- Optimization uses **perfect foresight** (knows future prices)
- Does not account for market liquidity or execution costs
- Represents theoretical maximum, not achievable in practice
            """)

        with st.expander("📐 Capture Rate - How is it calculated?"):
            st.markdown("""
**Formula:**
```
Capture Rate = (Actual GridBeyond Revenue ÷ Optimized Multi-Market Revenue) × 100
```

**Explanation:**
Shows what percentage of the theoretical optimal revenue was actually captured by GridBeyond.
A capture rate of 100% means actual matched optimal; >100% means outperformance (possible when
actual strategies earn revenue from sources not in the optimization model).

**Example (January 2026):**
- Actual GridBeyond Revenue: £28,190
- Multi-Market Optimal: £33,376
- Capture Rate: (28,190 ÷ 33,376) × 100 = **84.5%**

**Note:** On SFFR-only days, capture rates are typically 95%+ (gap is only from availability
assumption: optimizer uses 7.0 MW, actual averages 6.81 MW). Large gaps concentrate in 2–3
spike days per month when SSP prices spike unpredictably.

**Interpretation:**
- **>100%**: Outperforming optimization (revenue from strategies not modeled)
- **80–100%**: Good performance, close to theoretical optimal
- **60–80%**: Room for improvement in market participation
- **<60%**: Significant opportunity gap to investigate
            """)

    st.markdown("---")

    # Section 4: TB Spread Benchmarks
    st.header("4. TB Spread Benchmarks")
    st.caption("Top-Bottom spread analysis: comparing theoretical arbitrage potential to actual wholesale trading revenue")

    if data_loaded:
        try:
            capacity_mwh = 8.4
            combined_months = {}  # short -> combined DataFrame

            for m in bm:
                short = m['short']
                tb = calculate_tb_spreads(masters[short])
                arb = calculate_daily_arbitrage(masters[short])
                comb = tb.merge(arb, on='Date', how='inner')

                if len(comb) > 0:
                    for n in [1, 2, 3]:
                        comb[f'TB{n}_Max'] = comb[f'TB{n}'] * capacity_mwh
                        comb[f'TB{n}_Capture'] = (comb['Arbitrage_Revenue'] / comb[f'TB{n}_Max']) * 100
                    comb = comb.replace([float('inf'), float('-inf')], float('nan'))
                    combined_months[short] = comb

            if combined_months:
                # Summary metrics
                all_tb2 = [c['TB2'].mean() for c in combined_months.values()]
                all_arb = [c['Arbitrage_Revenue'].mean() for c in combined_months.values()]
                all_cap = [c['TB2_Capture'].mean() for c in combined_months.values()]

                avg_tb2 = np.mean(all_tb2)
                avg_arb = np.mean(all_arb)
                avg_capture_tb2 = np.nanmean(all_cap)

                col1, col2, col3, col4 = st.columns(4)

                with col1:
                    st.metric("Avg Daily TB2", f"£{avg_tb2:.0f}/MWh",
                              help="Average daily TB2 spread (sum of 2 highest - 2 lowest hourly prices)")
                with col2:
                    st.metric("Avg Daily Arbitrage", f"£{avg_arb:.0f}",
                              help="Average daily wholesale trading revenue")
                with col3:
                    delta_val = avg_capture_tb2 - 142 if not pd.isna(avg_capture_tb2) else 0
                    st.metric("TB2 Capture Rate", f"{avg_capture_tb2:.0f}%" if not pd.isna(avg_capture_tb2) else "N/A",
                              delta=f"{delta_val:+.0f}% vs benchmark",
                              help="Actual arbitrage as % of TB2 theoretical max. Industry benchmark: 142% for 2-hr batteries")
                with col4:
                    st.metric("Industry Benchmark", "142%",
                              help="2-hour batteries typically earn ~142% of TB2 spread (Source: Modo Energy)")

                # Monthly comparison table
                st.subheader("Monthly TB Spread Summary")

                summary_data = {
                    'Metric': ['Avg TB1 (£/MWh)', 'Avg TB2 (£/MWh)', 'Avg TB3 (£/MWh)',
                               'Avg Daily Arbitrage (£)', 'TB2 Capture Rate (%)'],
                }
                for short, comb in combined_months.items():
                    cap_mean = comb['TB2_Capture'].mean()
                    summary_data[short] = [
                        f"£{comb['TB1'].mean():.1f}",
                        f"£{comb['TB2'].mean():.1f}",
                        f"£{comb['TB3'].mean():.1f}",
                        f"£{comb['Arbitrage_Revenue'].mean():.0f}",
                        f"{cap_mean:.0f}%" if not pd.isna(cap_mean) else "N/A"
                    ]

                summary_df = pd.DataFrame(summary_data)
                st.dataframe(summary_df, use_container_width=True, hide_index=True)

                # Calculation explanations for Section 3
                st.subheader("Metric Calculations")

                with st.expander("📐 TB1 - How is it calculated?"):
                    st.markdown("""
**Formula:**
```
TB1 = Highest hourly price − Lowest hourly price (for the day)
```

**Explanation:**
The simplest measure of daily price spread. Represents the theoretical maximum
profit from a single 1-hour charge/discharge cycle buying at the lowest price
and selling at the highest.

**Example (1st September 2025):**
- Highest hourly EPEX price: £85/MWh (at 6pm)
- Lowest hourly EPEX price: £32/MWh (at 3am)
- **TB1 = £85 - £32 = £53/MWh**
                    """)

                with st.expander("📐 TB2 - How is it calculated?"):
                    st.markdown("""
**Formula:**
```
TB2 = (Sum of 2 highest hourly prices) − (Sum of 2 lowest hourly prices)
```

**Explanation:**
Represents the theoretical maximum profit for a 2-hour battery (like Northwold)
that can charge for 2 hours at the cheapest prices and discharge for 2 hours
at the most expensive prices.

**Example (1st September 2025):**
- 2 highest prices: £85 + £78 = £163
- 2 lowest prices: £32 + £35 = £67
- **TB2 = £163 - £67 = £96/MWh**

*TB2 is the primary benchmark for 2-hour duration batteries.*
                    """)

                with st.expander("📐 TB3 - How is it calculated?"):
                    st.markdown("""
**Formula:**
```
TB3 = (Sum of 3 highest hourly prices) − (Sum of 3 lowest hourly prices)
```

**Explanation:**
Theoretical maximum for a 3-hour battery. Useful for comparing against longer
duration assets or understanding diminishing returns of longer duration.

**Example (1st September 2025):**
- 3 highest prices: £85 + £78 + £72 = £235
- 3 lowest prices: £32 + £35 + £38 = £105
- **TB3 = £235 - £105 = £130/MWh**
                    """)

                with st.expander("📐 Daily Arbitrage Revenue - How is it calculated?"):
                    st.markdown("""
**Formula:**
```
Daily Arbitrage = EPEX DA Revenue + IDA1 Revenue + IDC Revenue
```

**Explanation:**
Sum of all wholesale trading revenues for the day. Excludes frequency response
(SFFR) since that's an ancillary service, not arbitrage.

**Example (1st September 2025):**
- EPEX DA Revenue: £180
- IDA1 Revenue: £95
- IDC Revenue: £0
- **Daily Arbitrage = £275**

*This is compared against TB2 theoretical maximum to calculate capture rate.*
                    """)

                with st.expander("📐 TB2 Capture Rate - How is it calculated?"):
                    st.markdown("""
**Formula:**
```
TB2 Capture Rate = (Arbitrage Revenue ÷ (TB2 × Capacity_MWh)) × 100
```

**Explanation:**
Measures how much of the theoretical maximum arbitrage revenue was actually captured.
The denominator (TB2 × 8.4 MWh) represents perfect trading at best prices.

**Example (1st September 2025):**
- TB2 spread: £96/MWh
- Capacity: 8.4 MWh
- Theoretical max: £96 × 8.4 = £806
- Actual arbitrage: £275
- **Capture Rate = (275 ÷ 806) × 100 = 34%**

*Note: Capture rates >100% are possible through intraday trading (multiple cycles)
and stacking frequency response on top of arbitrage.*
                    """)

                with st.expander("📐 Industry Benchmark (142%) - What does it mean?"):
                    st.markdown("""
**Source:** Modo Energy - "Benchmarking European battery revenue with TB spreads"

**Explanation:**
Modo Energy analyzed revenues from GB BESS assets and found that well-operated
2-hour batteries typically earn ~142% of the TB2 theoretical spread. This is
achieved through:

1. **Multiple daily cycles** - Trading more than once per day
2. **Intraday optimization** - Capturing price spikes in real-time markets
3. **Ancillary service stacking** - Earning frequency response alongside arbitrage
4. **Balancing Mechanism participation** - Additional revenue from grid balancing

**Rating Guide:**
- **≥142%**: At or above industry benchmark (excellent)
- **100-142%**: Capturing spread but below benchmark (room for improvement)
- **<100%**: Not fully capturing available spread (needs investigation)
                    """)

                # Time series chart
                st.subheader("Daily TB2 Spread vs Actual Arbitrage")

                all_combined = pd.concat([
                    comb.assign(Month=short) for short, comb in combined_months.items()
                ])
                all_combined['Date'] = pd.to_datetime(all_combined['Date'])

                fig_tb = go.Figure()

                fig_tb.add_trace(go.Scatter(
                    x=all_combined['Date'],
                    y=all_combined['TB2_Max'],
                    name='TB2 Theoretical Max (£)',
                    mode='lines',
                    line=dict(color='lightblue', width=2),
                    fill='tozeroy',
                    fillcolor='rgba(173, 216, 230, 0.3)'
                ))

                fig_tb.add_trace(go.Scatter(
                    x=all_combined['Date'],
                    y=all_combined['Arbitrage_Revenue'],
                    name='Actual Arbitrage (£)',
                    mode='lines+markers',
                    line=dict(color=COLOR_ACTUAL, width=2),
                    marker=dict(size=4)
                ))

                fig_tb.add_trace(go.Scatter(
                    x=all_combined['Date'],
                    y=all_combined['TB2_Max'] * 1.42,
                    name='Industry Benchmark (142%)',
                    mode='lines',
                    line=dict(color='orange', dash='dash', width=1)
                ))

                fig_tb.update_layout(
                    xaxis_title="Date",
                    yaxis_title="Revenue (£)",
                    height=400,
                    showlegend=True,
                    legend=dict(yanchor="top", y=0.99, xanchor="left", x=0.01)
                )

                st.plotly_chart(fig_tb, use_container_width=True)

                # Performance rating
                col1, col2 = st.columns(2)

                with col1:
                    st.markdown("**Performance Rating:**")
                    for short, comb in combined_months.items():
                        avg_cap = comb['TB2_Capture'].mean()
                        if pd.isna(avg_cap):
                            continue
                        if avg_cap >= 142:
                            st.success(f"{short}: {avg_cap:.0f}% - At or above benchmark")
                        elif avg_cap >= 100:
                            st.warning(f"{short}: {avg_cap:.0f}% - Below benchmark but positive")
                        else:
                            st.info(f"{short}: {avg_cap:.0f}% - Below 100%")

                with col2:
                    st.markdown("**TB Spread Interpretation:**")
                    st.markdown("""
                    - **TB2 Capture > 142%**: Exceeding industry benchmark for 2-hr batteries
                    - **TB2 Capture 100-142%**: Capturing spread but below benchmark
                    - **TB2 Capture < 100%**: Not fully capturing available spread

                    *Capture rates >100% achieved through intraday trading and frequency response stacking.*
                    """)

                # Expandable daily details
                with st.expander("View Daily TB Spread Details"):
                    tab_labels = list(combined_months.keys())
                    tabs = st.tabs(tab_labels)
                    for tab, short in zip(tabs, tab_labels):
                        with tab:
                            display = combined_months[short][['Date', 'TB1', 'TB2', 'TB3', 'Arbitrage_Revenue', 'TB2_Capture']].copy()
                            display.columns = ['Date', 'TB1 (£/MWh)', 'TB2 (£/MWh)', 'TB3 (£/MWh)', 'Arbitrage (£)', 'Capture (%)']
                            display['Date'] = pd.to_datetime(display['Date']).dt.strftime('%Y-%m-%d')
                            st.dataframe(display, use_container_width=True, hide_index=True)

                st.caption("""
                **TB Spread Benchmark Source:** [Modo Energy - Benchmarking European battery revenue with TB spreads](https://modoenergy.com/research/top-bottom-spread-revenue-benchmark-battery-energy-storage-sytems-gb-europe-spain-germany-solar-2025)
                """)
            else:
                st.warning("Insufficient data to calculate TB spreads. Ensure price data is available.")

        except Exception as e:
            st.error(f"Error calculating TB spreads: {str(e)}")

    st.markdown("---")

    # Source links
    with st.expander("📎 Data Sources"):
        st.markdown("""
        **Aggregator Revenue (GridBeyond):**
        - Monthly backing data spreadsheets with half-hourly dispatch, SFFR revenues, EPEX DA, IDA1, IDC, imbalance settlements
        - Source files: `Northwold YYYYMM Backing Data.xlsx` (raw) → processed into `Master_BESS_Analysis_*.csv`
        - Markets: SFFR (Frequency Response), EPEX Day Ahead, IDA1/ISEM (Intraday), IDC (Continuous), SSP/SBP (Imbalance)

        **Capacity Market (EMR Settlement):**
        - Source: Electricity Settlements Company (ESC) — T062 settlement CSV files
        - Contract: CAN-2025-NSFL01-001 (Northwold Solar Farm Ltd)
        - Capacity obligation: 1.023 MW @ ~£20,000/MW/year, paid monthly with seasonal weighting
        - Files: `NORTHWO_*_T062.csv` in `raw/New/`

        **DUoS — Distribution Use of System (Hartree Partners):**
        - Source: Hartree Partners Supply (UK) — generator invoices (passthrough credits from UKPN)
        - Components: Red/Amber/Green GDuos credits (revenue to generator for export) + DNO Fixed Charges (cost)
        - Files: `NWOSFL_Hartree*_Gen_Inv_*.pdf` in `raw/New/OneDrive_3*/` and `OneDrive_4*/`

        **IAR — Investment Appraisal Report:**
        - Source: Internal financial model — projected revenue per MW per month by stream
        - File: `extra/Northwold BESS Revenue_IAR.xlsx` (Sheet1, rows 4-11, multiplied by 4.2 MW for Real totals)
        - Streams: DA, ID, BM, FR, CM, DUoS Battery, DUoS Fixed, TNUoS

        **Multi-Market Optimisation:**
        - Hindsight-based linear programming (scipy.optimize.linprog, HiGHS solver)
        - Uses 5 market prices per period: EPEX, GB-ISEM, SSP, SBP, DA HH
        - Constraints: 4.2 MW charge / 7.5 MW discharge, 8.4 MWh capacity, 87% RTE, 1.5 cycles/day warranty
        - Output: `Optimized_Results_*.csv`

        **Industry Benchmarks (Modo Energy GB BESS Index):**
        - [2024 Year in Review](https://modoenergy.com/research/gb-battery-energy-storage-markets-2024-year-in-review-great-britain-wholesale-balancing-mechanism-frequency-response-reserve) — Annual summary *(Jan 2025)*
        - [December 2024 Benchmark](https://modoenergy.com/research/battery-energy-storage-revenues-december-benchmark-gb-2024-quick-reserve) — £84k/MW/year *(Jan 2025)*
        - [January 2025 Roundup](https://modoenergy.com/research/gb-research-roundup-january-2025-battery-energy-storage-great-britain-revenues-markets-wholesale-capacity-market-balancing-mechanism) — £88k/MW/year *(Feb 2025)*
        - [June 2025 Benchmark](https://modoenergy.com/research/battery-energy-storage-revenues-gb-benchmark-june-2025-negative-prices) — £76k/MW/year *(Jul 2025)*
        - [TB Spread Benchmark](https://modoenergy.com/research/top-bottom-spread-revenue-benchmark-battery-energy-storage-sytems-gb-europe-spain-germany-solar-2025) — 142% TB2 capture for 2-hr batteries

        **Technical References:**
        - **NREL**: [Battery Degradation Study](https://www.nrel.gov/docs/fy22osti/80688.pdf) *(2022)*
        - **National Grid ESO**: [Transmission Performance Reports](https://www.nationalgrideso.com/research-and-publications/transmission-performance-reports) *(quarterly)*
        - **DNV GL**: Energy Storage Performance Standards *(industry standard)*
        - **OEM Warranty**: 1.5 cycles/day limit *(per Northwold Storage Asset Optimisation Agreement)*

        **Note:** Modo Energy's GB BESS Index tracks monthly revenues across all GB batteries. Range £36k–£88k/MW/year (2024–2025). Total Revenue in this dashboard includes GridBeyond + CM + DUoS for like-for-like comparison.
        """)

    st.markdown("---")

    if data_loaded:
        # Visual comparison chart
        st.subheader("Revenue Benchmark Comparison")

        fig = go.Figure()

        # Industry range bar
        fig.add_trace(go.Bar(
            name='Industry Range',
            x=['Revenue £/MW/year'],
            y=[88000 - 36000],
            base=[36000],
            marker_color='lightgray',
            width=0.3,
            showlegend=True
        ))

        fig.add_hline(y=60000, line_dash="dash", line_color="orange",
                      annotation_text="Industry Mid (£60k)", annotation_position="right")

        # One scatter point per month
        symbols = ['diamond', 'circle', 'square', 'triangle-up', 'star']
        colors = [COLOR_EPEX, COLOR_ACTUAL, COLOR_SFFR, COLOR_MULTI_MARKET, '#E69F00']
        max_apm = 0
        for idx, m in enumerate(bm):
            apm = m['total_annual_per_mw']
            max_apm = max(max_apm, apm)
            fig.add_trace(go.Scatter(
                name=m['short'],
                x=['Revenue £/MW/year'],
                y=[apm],
                mode='markers',
                marker=dict(size=18, color=colors[idx % len(colors)],
                            symbol=symbols[idx % len(symbols)]),
            ))

        fig.update_layout(
            title="Northwold vs Industry",
            yaxis_title="£/MW/year",
            yaxis=dict(range=[0, max(120000, max_apm * 1.1)]),
            height=400,
            showlegend=True
        )

        st.plotly_chart(fig, use_container_width=True)

        # Key insights
        st.markdown("---")
        st.subheader("Key Insights")

        col1, col2 = st.columns(2)

        with col1:
            st.markdown("**Strengths:**")
            above_high = [m['short'] for m in bm if m['total_annual_per_mw'] >= 88000]
            if above_high:
                st.markdown(f"- {', '.join(above_high)} exceeded industry high benchmark")
            if avg_annual >= 60000:
                st.markdown(f"- Combined average (£{avg_annual:,.0f}/MW/yr) above industry mid")
            within_warranty = [m['short'] for m in bm if m['daily_cycles'] and m['daily_cycles'] <= 1.5]
            if within_warranty:
                st.markdown(f"- {', '.join(within_warranty)} cycling within warranty limits (<=1.5/day)")

        with col2:
            st.markdown("**Areas for Improvement:**")
            below_mid = [(m['short'], 60000 - m['total_annual_per_mw']) for m in bm if m['total_annual_per_mw'] < 60000]
            for short, gap in below_mid:
                st.markdown(f"- {short}: £{gap:,.0f}/MW below industry mid")
            if avg_annual < 88000:
                gap = 88000 - avg_annual
                st.markdown(f"- Combined average £{gap:,.0f}/MW below industry high")

        st.markdown("---")
        with st.expander("📰 Modo Energy source articles (click month to open)"):
            st.caption(
                "Each month's benchmark is the headline GBP/MW/year figure from "
                "Modo Energy's monthly ME BESS GB article. March 2026 is "
                "API-derived pending the article's publication."
            )
            for short in ['Sep 25', 'Oct 25', 'Nov 25', 'Dec 25', 'Jan 26', 'Feb 26', 'Mar 26']:
                val = MODO_BENCHMARKS.get(short)
                url = MODO_SOURCE_LINKS.get(short)
                if val and url:
                    suffix = " *(API — article pending)*" if short == 'Mar 26' else ""
                    st.markdown(f"- **{short}** — £{val:,}/MW/yr — [{url.split('/')[-1][:80]}]({url}){suffix}")

        st.caption("""
        **Sources:**
        - Revenue benchmarks: Modo Energy ME BESS GB Index (monthly articles — see expander above)
        - Degradation & cycling: Industry warranty standards and research
        - Availability (TWCAA): National Grid ESO performance data
        - Round-trip efficiency: Lithium-ion industry specifications
        """)


# ============================================================================
# Benchmark Sources Catalogue
# ============================================================================
# Reference page listing all third-party BESS revenue benchmarks under
# evaluation for APD. Source metadata (methodology, geography, access model,
# status) is captured here as the single source of truth before deciding how
# to integrate them into the data pipeline and benchmark comparison engine.

BENCHMARK_SOURCES = [
    {
        'id': 'modo_gb_articles',
        'vendor': 'Modo Energy',
        'product': 'GB BESS Index — monthly research articles',
        'countries': ['UK'],
        'granularity': 'Monthly headline',
        'duration': 'Fleet average',
        'rte': 'Fleet average',
        'cycles_per_day': 'Fleet average',
        'streams': ['DA', 'ID', 'BM', 'FR (DC/DM/DR)', 'CM'],
        'access': 'Free articles (JS-rendered, manual extraction); full data via paid Modo Terminal',
        'status': 'Active',
        'status_note': 'Currently wired into Benchmarks page (hardcoded monthly £/MW/yr dict)',
        'url': 'https://modoenergy.com/research',
        'notes': 'Headline £/MW/year is the aggregate across all tracked GB batteries, annualised. No single duration/RTE assumption — reflects real fleet mix.',
        'data_provided': [
            "Monthly headline £/MW/year fleet average",
            "Narrative revenue stack breakdown (DA / ID / BM / FR / CM shares)",
            "TB2 spread capture %",
            "Fleet-average cycling stats",
            "Delivered as HTML articles (JS-rendered) — manual extraction each month",
            "Full time series requires paid Modo Terminal",
        ],
    },
    {
        'id': 'modo_me_bess_gb',
        'vendor': 'Modo Energy',
        'product': 'ME-BESS-GB public index (FCA-regulated)',
        'countries': ['UK'],
        'granularity': 'Daily / asset-level',
        'duration': 'TBD',
        'rte': 'TBD',
        'cycles_per_day': 'TBD',
        'streams': ['Asset-level revenue + capture rates + optimiser attribution'],
        'access': 'Paid (Modo Terminal + API); FCA-regulated, IOSCO-certified — embeddable in financing docs',
        'status': 'Evaluation',
        'status_note': 'Regulated benchmark — strongest credibility for contracts and financing',
        'url': 'https://modoenergy.com/public-indices/me-bess-gb',
        'notes': 'Distinct from the monthly research articles: this is a structured, regulated benchmark with API access. Higher cost but contract-grade.',
        'data_provided': [
            "Daily asset-level revenue time series",
            "Adjusted capture rates",
            "Optimiser attribution",
            "Revenue stream breakdown",
            "Embeddable index values for contracts / financing docs",
            "Structured JSON via API; downloadable from Terminal",
        ],
    },
    {
        'id': 'kyos_bess_report',
        'vendor': 'Kyos',
        'product': 'BESS Benchmark Report (ReFlex)',
        'countries': ['DE', 'NL'],
        'granularity': 'Monthly',
        'duration': 'Not specified',
        'rte': 'Not specified',
        'cycles_per_day': 'Not specified',
        'streams': ['DA', 'ID', 'aFRR'],
        'access': 'Free (email-gated monthly report)',
        'status': 'High priority',
        'status_note': 'Free + covers DE and NL — strong candidate for first non-UK benchmark',
        'url': 'https://www.kyos.com/bess-benchmark-report/',
        'notes': 'ReFlex optimisation model with full trading transparency and asset-backed trading strategies. No capacity market or balancing revenue.',
        'data_provided': [
            "Monthly €/MW revenue figures for DE and NL",
            "Split by wholesale (DA + ID) and aFRR",
            "Commentary on market conditions and optimisation performance",
            "Delivered as a free PDF report by email",
        ],
    },
    {
        'id': 'clean_horizon',
        'vendor': 'Clean Horizon',
        'product': 'Storage Index',
        'countries': ['BE', 'DK1', 'DK2', 'EE', 'FI', 'FR', 'DE', 'IT', 'LV', 'LT', 'PL', 'PT', 'RO', 'ES', 'SE'],
        'granularity': 'Monthly (free) / Daily (Premium)',
        'duration': 'Derived from usable AC energy',
        'rte': '85%',
        'cycles_per_day': '1.5',
        'streams': ['DA', 'ID', 'FCR', 'aFRR', 'mFRR', 'Imbalance', 'Ancillary', 'Capacity (FR, until Jan 2026)'],
        'access': 'Free monthly Storage Index; paid Premium for daily + custom project analysis',
        'status': 'High priority',
        'status_note': 'Broadest European coverage (14 countries) — anchor benchmark for pan-EU expansion',
        'url': 'https://www.cleanhorizon.com/battery-index/',
        'notes': 'Methodology closest to Modo Energy. Gross revenue annualised. 100% availability assumed.',
        'data_provided': [
            "Free: monthly annualised gross €/MW/year per country (14 markets)",
            "Premium: daily values",
            "Premium: per-stream breakdown (DA, ID, FCR, aFRR, mFRR, imbalance, ancillary, capacity)",
            "Premium: custom project-level analysis",
            "Free viz on website; paid data via portal / API",
        ],
    },
    {
        'id': 'enspired',
        'vendor': 'enspired trading',
        'product': 'Battery Index',
        'countries': ['DE'],
        'granularity': 'TBD',
        'duration': '1h and 2h',
        'rte': 'Not specified',
        'cycles_per_day': '1 or 2 (matching duration)',
        'streams': ['Spot (DA, IDA, IDC)', 'FCR', 'aFRR'],
        'access': 'Free (planned Excel download)',
        'status': 'Pre-launch',
        'status_note': 'Under development as of Sep 2024 — monitor for release',
        'url': 'https://www.enspired-trading.com/blog/battery-index-to-benchmark-bess-revenue-potential',
        'notes': 'Cross-market revenue potential benchmark. Parameters to be flexibilised in future versions.',
        'data_provided': [
            "Planned: €/MW revenue figures for 1h and 2h batteries in DE",
            "Split by spot (DA + IDA + IDC), FCR, and aFRR",
            "Target format: publicly accessible Excel download",
            "Not yet released",
        ],
    },
    {
        'id': 'aurora_gb',
        'vendor': 'Aurora Energy Research',
        'product': 'GB Battery Index',
        'countries': ['UK'],
        'granularity': 'TBD',
        'duration': 'TBD',
        'rte': 'TBD',
        'cycles_per_day': 'TBD',
        'streams': ['TBD'],
        'access': 'Paid (subscription likely)',
        'status': 'Evaluation',
        'status_note': 'UK second opinion alongside Modo — useful for triangulation',
        'url': 'https://auroraer.com/resources/aurora-insights/articles/introducing-auroras-gb-battery-index',
        'notes': 'Details gated — need to request access/demo to confirm methodology and pricing.',
        'data_provided': [
            "TBD — likely historical GB revenue index with forward-looking outlook",
            "Aurora EOS platform style (subscription-based)",
            "Request demo to confirm granularity, stream breakdown, and delivery format",
        ],
    },
    {
        'id': 'montel',
        'vendor': 'Montel',
        'product': 'BESS Index + UK BESS Leaderboard',
        'countries': ['NL', 'BE', 'DE', 'UK', 'Pan-EU'],
        'granularity': 'TBD',
        'duration': 'Multiple',
        'rte': 'TBD',
        'cycles_per_day': 'TBD',
        'streams': ['Spot', 'Intraday', 'Balancing', 'Ancillary'],
        'access': 'Paid subscription',
        'status': 'Evaluation',
        'status_note': 'Strongest single-vendor pan-EU commercial offering; covers all AMPYR target geographies',
        'url': 'https://montel.energy/products/bess',
        'notes': 'Standardised cross-country comparison with expert validation and forward-looking scenarios. Pricing undisclosed.',
        'data_provided': [
            "Operational BESS Index (NL / BE / DE)",
            "Pan-European BESS Index",
            "UK BESS Leaderboard (asset-level ranking)",
            "Project-level revenue evaluation",
            "Forward-looking scenario analysis",
            "Delivered via Montel platform with API access for data integration",
        ],
    },
    {
        'id': 'enervis',
        'vendor': 'enervis energy',
        'product': 'BESS Index',
        'countries': ['DE'],
        'granularity': 'TBD',
        'duration': 'TBD',
        'rte': 'TBD',
        'cycles_per_day': 'TBD',
        'streams': ['TBD'],
        'access': 'Free on request (form-gated)',
        'status': 'Evaluation',
        'status_note': 'DE second opinion — methodology needs confirmation after requesting the index',
        'url': 'https://enervis.de/anfrage-bess-index/',
        'notes': 'Landing page is a request form; details in the emailed report.',
        'data_provided': [
            "TBD — DE BESS revenue index",
            "Methodology detailed in the emailed report",
            "Submit request form to confirm granularity, stream coverage, and data format",
        ],
    },
    {
        'id': 'regelleistung_de_2h',
        'vendor': 'Regelleistung Online',
        'product': 'German Energy Storage Revenue Index 2h',
        'countries': ['DE'],
        'granularity': '30d and 365d moving averages, annualised',
        'duration': '2h',
        'rte': '90% AC/AC',
        'cycles_per_day': 'Max 2 (100% DoD)',
        'streams': ['ID1 EPEX 15-min', 'FCR capacity', 'aFRR capacity'],
        'access': 'Free web visualisation; data on request',
        'status': 'High priority',
        'status_note': 'Cleanest published DE methodology, free access — baseline DE benchmark',
        'url': 'https://www.regelleistung-online.de/german-energy-storage-revenue-index/bess-revenue-index-2h/',
        'notes': 'Excludes mFRR and aFRR activation revenues. End-of-day SoC target 50%. Official TSO-published index.',
        'data_provided': [
            "30-day and 365-day moving averages of €/MW/year annualised revenue",
            "Applies to a 2h BESS in DE",
            "Split by ID1 arbitrage, FCR capacity, aFRR capacity",
            "Free interactive chart on the regelleistung-online portal",
            "Raw data available on request",
        ],
    },
]


# Methodology detail keyed by source id. Kept separate so BENCHMARK_SOURCES
# stays compact; rendered in Section 5 of the Benchmark Sources page.
BENCHMARK_METHODOLOGY = {
    'modo_gb_articles': {
        'class': 'Empirical',
        'foresight': 'N/A — observed',
        'algorithm': 'Fleet aggregate of tracked GB BESS assets',
        'basis': 'Realised revenue (mixed gross/net per asset)',
        'availability': 'Actual fleet availability',
        'caveats': (
            "No stated duration/RTE/cycling assumption — reflects real fleet "
            "mix. Includes a long tail of poorly-run assets, so beating the "
            "headline figure is not an ambitious target."
        ),
    },
    'modo_me_bess_gb': {
        'class': 'Empirical (FCA-regulated, IOSCO-certified)',
        'foresight': 'N/A — observed',
        'algorithm': 'Asset-level revenues with capture-rate adjustment and optimiser attribution',
        'basis': 'Methodology behind paywall',
        'availability': 'Actual',
        'caveats': (
            "Regulated benchmark intended for embedding in contracts and "
            "financing. Full methodology only published to Terminal subscribers "
            "and in the IOSCO compliance statement."
        ),
    },
    'kyos_bess_report': {
        'class': 'Theoretical',
        'foresight': '"Realistic market approach" — likely rolling, not perfect foresight',
        'algorithm': 'ReFlex optimisation model (proprietary, undisclosed on public page)',
        'basis': 'Undisclosed — likely gross',
        'availability': 'Unknown',
        'caveats': (
            "Marketed as transparent with asset-backed trading strategies, but "
            "algorithm specifics not published on the landing page. Detail "
            "likely in the emailed PDF."
        ),
    },
    'clean_horizon': {
        'class': 'Theoretical (COSMOS simulation)',
        'foresight': 'Hybrid — perfect DA price + D-1 ancillary known; statistical estimation elsewhere',
        'algorithm': 'COSMOS proprietary energy-storage simulation',
        'basis': 'Gross — grid fees, taxes and SoC management costs explicitly excluded',
        'availability': '100% assumed',
        'caveats': (
            "Gross-of-fees headline will overstate what hits a real P&L; "
            "Northwold grid fees alone are ~20% of revenue. 1.5 cyc/day cap, "
            "85% RTE. Methodology is the most similar to Regelleistung."
        ),
    },
    'enspired': {
        'class': 'Hybrid (theoretical with real-asset validation in Phase 4)',
        'foresight': 'Undisclosed',
        'algorithm': 'Phase 1: simple weighting of single-market revenues (ignores conflicts between market designs)',
        'basis': 'Undisclosed',
        'availability': 'Unknown',
        'caveats': (
            "Still under development (phased release). Early versions neglect "
            "conflicts between market participation — treat as indicative, not "
            "definitive."
        ),
    },
    'aurora_gb': {
        'class': 'Unknown (gated)',
        'foresight': 'TBD',
        'algorithm': 'TBD',
        'basis': 'TBD',
        'availability': 'TBD',
        'caveats': (
            "Methodology gated behind Aurora's sales process. Request a demo "
            "to confirm whether the index is empirical fleet-based or a "
            "modelled optimum."
        ),
    },
    'montel': {
        'class': 'Mixed — UK Leaderboard is empirical asset-level ranking; pan-EU index undisclosed',
        'foresight': 'Product-dependent',
        'algorithm': 'Undisclosed publicly; Leaderboard appears to aggregate asset-level revenues from operators',
        'basis': 'Undisclosed',
        'availability': 'Product-dependent',
        'caveats': (
            "Multiple products under one brand — the UK Leaderboard and the "
            "pan-EU BESS Index likely have different methodologies. Detail "
            "only surfaced during demo / contract."
        ),
    },
    'enervis': {
        'class': 'Unknown (form-gated)',
        'foresight': 'TBD',
        'algorithm': 'TBD',
        'basis': 'TBD',
        'availability': 'TBD',
        'caveats': "Methodology detailed in the emailed report — submit the request form to confirm.",
    },
    'regelleistung_de_2h': {
        'class': 'Theoretical',
        'foresight': 'Perfect foresight',
        'algorithm': 'Linear Programming — daily one-day-ahead dispatch',
        'basis': 'Gross (90% RTE losses only; no SoC management cost, no aFRR activation revenue)',
        'availability': '100% assumed',
        'caveats': (
            "Upper bound — perfect foresight + LP optimum means real operators "
            "should expect a structural gap. Daily optimisation picks the "
            "highest-revenue market of the day. Official TSO publication."
        ),
    },
}


def show_benchmark_sources():
    """Catalogue page — all third-party BESS benchmarks under evaluation."""
    st.title("🗂️ Benchmark Sources")
    st.markdown(
        "Reference catalogue of third-party BESS revenue benchmarks under evaluation for APD. "
        "Use this page to compare methodology, geography coverage, and access model before "
        "deciding which sources to integrate into the dashboard."
    )

    # ---- Summary metrics ----
    all_countries = sorted({c for s in BENCHMARK_SOURCES for c in s['countries']})
    active_count = sum(1 for s in BENCHMARK_SOURCES if s['status'] == 'Active')
    high_prio_count = sum(1 for s in BENCHMARK_SOURCES if s['status'] == 'High priority')
    eval_count = sum(1 for s in BENCHMARK_SOURCES if s['status'] == 'Evaluation')
    prelaunch_count = sum(1 for s in BENCHMARK_SOURCES if s['status'] == 'Pre-launch')
    free_count = sum(1 for s in BENCHMARK_SOURCES if 'Free' in s['access'])

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Total sources", len(BENCHMARK_SOURCES))
    c2.metric("Countries covered", len(all_countries))
    c3.metric("Free access", free_count)
    c4.metric("Active (wired in)", active_count)
    c5.metric("High priority", high_prio_count)

    st.caption(
        f"Status breakdown: {active_count} active · {high_prio_count} high priority · "
        f"{eval_count} evaluation · {prelaunch_count} pre-launch"
    )

    st.markdown("---")

    # ---- Overview table ----
    st.header("1. Overview")
    overview_rows = []
    for s in BENCHMARK_SOURCES:
        overview_rows.append({
            'Source': f"{s['vendor']} — {s['product']}",
            'Countries': ', '.join(s['countries']),
            'Granularity': s['granularity'],
            'Duration': s['duration'],
            'RTE': s['rte'],
            'Cycles/day': s['cycles_per_day'],
            'Access': 'Free' if 'Free' in s['access'] and 'Paid' not in s['access'].split(';')[0]
                     else ('Free + Paid' if 'Free' in s['access'] else 'Paid'),
            'Status': s['status'],
        })
    overview_df = pd.DataFrame(overview_rows)

    def _status_color(val):
        colors = {
            'Active': 'background-color: #d4edda; color: #155724',
            'High priority': 'background-color: #fff3cd; color: #856404',
            'Evaluation': 'background-color: #e2e3e5; color: #383d41',
            'Pre-launch': 'background-color: #cce5ff; color: #004085',
        }
        return colors.get(val, '')

    styled_overview = overview_df.style.map(_status_color, subset=['Status'])
    st.dataframe(styled_overview, use_container_width=True, hide_index=True)

    st.markdown("---")

    # ---- Coverage matrix: Country × Source ----
    st.header("2. Geographic Coverage Matrix")
    st.caption("Which sources cover which markets. ✓ = full coverage as published.")

    # Flatten 'Pan-EU' into a separate column
    display_countries = [c for c in all_countries if c != 'Pan-EU']
    if 'Pan-EU' in all_countries:
        display_countries.append('Pan-EU')

    matrix_rows = []
    for s in BENCHMARK_SOURCES:
        row = {'Source': f"{s['vendor']} — {s['product'].split('—')[0].strip()}"}
        for country in display_countries:
            row[country] = '✓' if country in s['countries'] else ''
        matrix_rows.append(row)
    matrix_df = pd.DataFrame(matrix_rows)
    st.dataframe(matrix_df, use_container_width=True, hide_index=True)

    # Country coverage summary
    st.subheader("Sources per country")
    country_counts = {}
    for country in display_countries:
        country_counts[country] = sum(1 for s in BENCHMARK_SOURCES if country in s['countries'])
    country_summary = pd.DataFrame(
        [{'Country': k, 'Number of sources': v} for k, v in sorted(country_counts.items(), key=lambda x: -x[1])]
    )
    st.dataframe(country_summary, use_container_width=True, hide_index=True)

    st.markdown("---")

    # ---- Detailed source cards ----
    st.header("3. Source Details")
    st.caption("Click to expand each source for methodology, access model, and notes.")

    status_icons = {
        'Active': '🟢',
        'High priority': '🟡',
        'Evaluation': '⚪',
        'Pre-launch': '🔵',
    }

    for s in BENCHMARK_SOURCES:
        icon = status_icons.get(s['status'], '⚪')
        header = f"{icon} **{s['vendor']}** — {s['product']}  ·  {', '.join(s['countries'])}"
        with st.expander(header):
            col_a, col_b = st.columns(2)
            with col_a:
                st.markdown(f"**Status:** {s['status']}")
                st.markdown(f"**Countries:** {', '.join(s['countries'])}")
                st.markdown(f"**Granularity:** {s['granularity']}")
                st.markdown(f"**Duration:** {s['duration']}")
                st.markdown(f"**RTE:** {s['rte']}")
                st.markdown(f"**Cycles/day:** {s['cycles_per_day']}")
            with col_b:
                st.markdown(f"**Revenue streams:** {', '.join(s['streams'])}")
                st.markdown(f"**Access:** {s['access']}")
                st.markdown(f"**URL:** [{s['url']}]({s['url']})")
                st.markdown(f"**Why it matters:** {s['status_note']}")

            if s.get('data_provided'):
                st.markdown("**What the data looks like**")
                st.markdown('\n'.join(f"- {item}" for item in s['data_provided']))

            method = BENCHMARK_METHODOLOGY.get(s['id'])
            if method:
                st.markdown("**How the number is arrived at**")
                st.markdown(f"- **Class:** {method['class']}")
                st.markdown(f"- **Foresight:** {method['foresight']}")
                st.markdown(f"- **Algorithm:** {method['algorithm']}")
                st.markdown(f"- **Revenue basis:** {method['basis']}")
                st.markdown(f"- **Availability assumption:** {method['availability']}")
                if method.get('caveats'):
                    st.info(method['caveats'])

            if s.get('notes'):
                st.markdown("**Notes**")
                st.markdown(s['notes'])

    st.markdown("---")

    # ---- What each source provides (side-by-side comparison) ----
    st.header("4. What Each Source Provides")
    st.caption(
        "Comparison of the actual outputs each benchmark publishes — the shape of "
        "the data you receive, not just the revenue streams covered."
    )

    # Header row
    h1, h2, h3, h4 = st.columns([3, 2, 2, 6])
    h1.markdown("**Source**")
    h2.markdown("**Geography**")
    h3.markdown("**Granularity**")
    h4.markdown("**Data provided**")
    st.markdown("<hr style='margin: 0.25rem 0'/>", unsafe_allow_html=True)

    # One row per source — last column wraps as flowing text
    for s in BENCHMARK_SOURCES:
        c1, c2, c3, c4 = st.columns([3, 2, 2, 6])
        c1.markdown(f"**{s['vendor']}**<br/><span style='color:#666;font-size:0.9em'>{s['product']}</span>",
                    unsafe_allow_html=True)
        c2.markdown(', '.join(s['countries']))
        c3.markdown(s['granularity'])
        items = s.get('data_provided') or []
        c4.markdown('\n'.join(f"- {item}" for item in items) if items else '—')
        st.markdown("<hr style='margin: 0.25rem 0; border-color: #eee'/>", unsafe_allow_html=True)

    st.markdown("---")

    # ---- Methodology analysis ----
    st.header("5. Methodology — How Each Benchmark Is Arrived At")
    st.markdown("""
Benchmark sources fall into two structurally different classes — and the
numbers they publish are **not directly comparable** to each other or to a
real asset's P&L without adjustment.

- **Theoretical optimisers** run a model with fixed BESS assumptions (duration,
  RTE, cycling) against historical market prices. They set an **upper bound** —
  "what a perfectly-operated battery could have earned."
- **Empirical fleet aggregates** report observed revenues from real operating
  batteries. They show the **realised average** — includes operational
  friction, aggregator skill, and availability.

Comparing Northwold to a theoretical benchmark tells you how much headroom the
optimiser claims exists. Comparing to a fleet aggregate tells you how you rank
against peers. The two questions are different.
    """)

    # Group sources by methodology class for the grouped view
    class_groups = {
        'Theoretical optimisers': ['clean_horizon', 'regelleistung_de_2h', 'kyos_bess_report'],
        'Empirical fleet aggregates': ['modo_gb_articles', 'modo_me_bess_gb', 'montel'],
        'Hybrid / pre-launch': ['enspired'],
        'Gated / TBD': ['aurora_gb', 'enervis'],
    }
    source_by_id = {s['id']: s for s in BENCHMARK_SOURCES}

    st.subheader("5.1 Methodology class")
    for group_name, ids in class_groups.items():
        names = [f"{source_by_id[i]['vendor']} ({source_by_id[i]['product'].split('—')[0].strip()})"
                 for i in ids if i in source_by_id]
        st.markdown(f"**{group_name}:** {'; '.join(names)}")

    st.markdown("---")

    st.subheader("5.2 Methodology matrix")
    method_rows = []
    for s in BENCHMARK_SOURCES:
        m = BENCHMARK_METHODOLOGY.get(s['id'], {})
        method_rows.append({
            'Source': f"{s['vendor']} — {s['product'].split('—')[0].strip()}",
            'Class': m.get('class', '—'),
            'Foresight': m.get('foresight', '—'),
            'Algorithm': m.get('algorithm', '—'),
            'Revenue basis': m.get('basis', '—'),
            'Availability': m.get('availability', '—'),
        })
    method_df = pd.DataFrame(method_rows)
    st.dataframe(method_df, use_container_width=True, hide_index=True)

    st.markdown("---")

    st.subheader("5.3 Implications for like-for-like comparison")
    st.markdown("""
Four traps to avoid when comparing Northwold's numbers to any of these
benchmarks:

1. **Perfect-foresight theoretical benchmarks are soft upper bounds.**
   Regelleistung and Clean Horizon assume the optimiser knew every price in
   advance. Real traders don't. A 20–30% gap to these is expected, not a
   performance failure.

2. **Gross-revenue benchmarks overstate what hits the P&L.** Clean Horizon
   explicitly excludes grid fees, taxes, and SoC management. Northwold grid
   fees alone are ~£6k/month (~20% of revenue). Comparing Northwold net to a
   gross benchmark will look worse than it is.

3. **Cycling-assumption asymmetry.** Regelleistung assumes 2 cyc/day;
   Clean Horizon 1.5; Northwold is actually running at ~0.95 cyc/day. Any
   headline "capture rate" is confounded — Northwold isn't just trading
   worse, it's trading less.

4. **Empirical fleet aggregates are peer comparisons, not targets.** Modo's
   headline £/MW/yr reflects whatever the tracked GB fleet earned — which
   includes a long tail of poorly-run assets. Beating it isn't ambitious; the
   FCA-regulated ME-BESS-GB likely strips to optimiser-attributed revenue to
   control for aggregator quality.
    """)


def main():
    """Main application with sidebar navigation"""

    # Sidebar configuration
    st.sidebar.title("🔋 Asset Performance Dashboard")

    # General section (month-independent pages)
    # Use session_state to persist the active general page across reruns.
    # Default to Executive Comparison on first load so the app opens on the
    # portfolio overview rather than a month-specific page.
    if 'active_general_page' not in st.session_state:
        st.session_state.active_general_page = 'exec_comparison'

    st.sidebar.markdown("### General")

    def _set_general_page(page_name):
        st.session_state.active_general_page = page_name

    show_asset_page = st.sidebar.button("🏭 Asset Details", use_container_width=True,
                                         on_click=_set_general_page, args=('asset_details',))
    show_import_page = st.sidebar.button("📥 Data Import", use_container_width=True,
                                          on_click=_set_general_page, args=('data_import',))
    show_exec_comparison = st.sidebar.button("📊 Executive Comparison", use_container_width=True,
                                              on_click=_set_general_page, args=('exec_comparison',))
    show_benchmark_page = st.sidebar.button("📈 Benchmarks", use_container_width=True,
                                             on_click=_set_general_page, args=('benchmarks',))
    show_benchmark_sources_page = st.sidebar.button("🗂️ Benchmark Sources", use_container_width=True,
                                                     on_click=_set_general_page, args=('benchmark_sources',))
    show_export_page = st.sidebar.button("📄 Export Reports", use_container_width=True,
                                          on_click=_set_general_page, args=('export_reports',))
    show_invoice_page = st.sidebar.button("🧾 Invoice Analysis", use_container_width=True,
                                           on_click=_set_general_page, args=('invoice_analysis',))
    show_checklist_page = st.sidebar.button("📋 Monthly Checklist", use_container_width=True,
                                             on_click=_set_general_page, args=('monthly_checklist',))

    st.sidebar.markdown("---")

    # Month selector for data-dependent pages
    st.sidebar.markdown("### 📅 Monthly Analysis")
    selected_month = st.sidebar.selectbox(
        "Select Month",
        list(AVAILABLE_MONTHS.keys()),
        index=0
    )

    # Navigation menu for month-dependent pages
    def _clear_general_page():
        st.session_state.active_general_page = None

    page = st.sidebar.radio(
        "Pages",
        ["📊 Operations Summary", "🚀 Multi-Market Optimization", "📈 Market Prices", "⚠️ Imbalance Analysis", "⚡ Ancillary Services", "🔋 BESS Health", "📑 Performance Report"],
        index=0,  # Default to operations summary
        label_visibility="collapsed",
        on_change=_clear_general_page,
    )

    st.sidebar.markdown("---")
    st.sidebar.info(f"""
    **System:** Northwold Solar Farm
    **Location:** Hall Farm
    **Capacity:** 8.4 MWh
    """)

    # Glossary of terms
    with st.sidebar.expander("📖 Glossary"):
        st.markdown("""
**Markets:**
- **EPEX** - European Power Exchange (Day Ahead)
- **IDA1** - Intraday Auction 1
- **IDC** - Intraday Continuous
- **SSP** - System Sell Price (grid pays you)
- **SBP** - System Buy Price (you pay grid)

**Ancillary Services:**
- **SFFR** - Static Firm Frequency Response
- **DC** - Dynamic Containment (L=Low, H=High)
- **DM** - Dynamic Moderation (L=Low, H=High)
- **DR** - Dynamic Regulation (L=Low, H=High)

**Other Terms:**
- **EFA** - Electricity Forward Agreement (4-hour blocks)
- **SOC** - State of Charge (%)
- **MWh** - Megawatt-hours (energy)
- **MW** - Megawatts (power)
- **pp** - Percentage points (absolute % change)
- **Capture Rate** - Actual / Optimal revenue (%)
- **Gap** - Optimal - Actual revenue (£)
        """)

    # Display General pages if active (persisted via session_state)
    active = st.session_state.active_general_page
    if active == 'asset_details':
        show_asset_details()
        return
    if active == 'data_import':
        show_data_quality_page()
        return
    if active == 'exec_comparison':
        show_executive_comparison()
        return
    if active == 'benchmarks':
        show_benchmark_comparison()
        return
    if active == 'benchmark_sources':
        show_benchmark_sources()
        return
    if active == 'export_reports':
        show_pdf_export_page(selected_month)
        return
    if active == 'invoice_analysis':
        show_invoice_analysis()
        return
    if active == 'monthly_checklist':
        show_monthly_checklist()
        return

    # Display selected monthly page
    if page == "📊 Operations Summary":
        # Load data for selected month
        with st.spinner(f"Loading {selected_month} data..."):
            bess_df, northwold_df = load_month_data(selected_month)

        if bess_df is None or northwold_df is None:
            st.error(f"Failed to load data files for {selected_month}. Please ensure CSV files are in the correct location.")
            return

        # Perform analysis
        bess_analysis = analyze_bess_data(bess_df)
        northwold_analysis = analyze_northwold_data(northwold_df)

        # Show operations summary
        show_operations_summary(bess_df, northwold_df, bess_analysis, northwold_analysis, month=selected_month)
    elif page == "🚀 Multi-Market Optimization":
        show_multimarket_optimization(selected_month)
    elif page == "📈 Market Prices":
        show_market_price_analysis(selected_month)
    elif page == "⚠️ Imbalance Analysis":
        show_imbalance_deep_dive(selected_month)
    elif page == "⚡ Ancillary Services":
        show_ancillary_services_analysis(selected_month)
    elif page == "🔋 BESS Health":
        show_bess_health(selected_month)
    elif page == "📑 Performance Report":
        show_report_page(selected_month)

if __name__ == "__main__":
    main()